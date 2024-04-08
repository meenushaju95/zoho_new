#Zoho Final
from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout
from django.contrib import messages
from django.conf import settings
from datetime import date
from datetime import datetime, timedelta
from Company_Staff.models import *
from django.db import models
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from django.template.loader import get_template
from bs4 import BeautifulSoup
import io,os
import csv
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse,HttpResponseRedirect
from io import BytesIO
from django.db.models import Max
from django.db.models import Q
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import render,redirect,get_object_or_404
from . models import *
from decimal import Decimal
from Company_Staff.models import Vendor, Vendor_comments_table, Vendor_doc_upload_table, Vendor_mail_table,Vendor_remarks_table,VendorContactPerson,VendorHistory,Delivery_challan,Delivery_challan_item,Delivery_challan_reference,Delivery_challan_history,Delivery_challan_comment
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest, HttpResponseNotFound, JsonResponse
from email.message import EmailMessage
from django.core.exceptions import ObjectDoesNotExist
import re
from .models import payroll_employee,Attendance,Attendance_History,Holiday,Attendance_comment,Bloodgroup
from calendar import monthrange
from collections import defaultdict
import calendar
from Company_Staff.models import *
import openpyxl
from django.shortcuts import get_object_or_404
import calendar
from django.core.mail import EmailMultiAlternatives  
from django.http.response import JsonResponse, HttpResponse
from django.core.mail import send_mail, EmailMultiAlternatives
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from Company_Staff.models import BankAccount
from Company_Staff.models import loan_account
from Company_Staff.models import LoanRepayemnt
from Company_Staff.models import LoanAccountHistory
from Company_Staff.models import LoanRepaymentHistory
from Company_Staff.models import BankAccountHistory
from Company_Staff.models import Comments
from Company_Staff.models import Banking
from django.shortcuts import render, get_object_or_404
from datetime import date as dt
from django.db.models import Sum
from django.utils.timezone import now
# Create your views here.



# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Calculate the date 20 days before the end date for payment term renew and 10 days before for trial period renew
        if dash_details.payment_term:
            reminder_date = dash_details.End_date - timedelta(days=20)
        else:
            reminder_date = dash_details.End_date - timedelta(days=10)
        current_date = date.today()
        alert_message = current_date >= reminder_date
        
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False

        # Calculate the number of days between the reminder date and end date
        days_left = (dash_details.End_date - current_date).days
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'alert_message':alert_message,
            'days_left':days_left,
            'payment_request':payment_request,
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company staff request for login approval
def company_staff_request(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
       
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request, pk):
    """
    Sets the company approval status to 2 for the specified staff member, effectively canceling staff approval.

    This function is designed to be used for canceling staff approval, and the company approval value is set to 2.
    This can be useful for identifying resigned staff under the company in the future.

    """
    staff = StaffDetails.objects.get(id=pk)
    staff.company_approval = 2
    staff.save()
    return redirect('company_all_staff')


# company profile, profile edit
def company_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        terms=PaymentTerms.objects.all()
        payment_history=dash_details.previous_plans.all()

        # Calculate the date 20 days before the end date
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        renew_button = current_date >= reminder_date

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'renew_button': renew_button,
            'terms':terms,
            'payment_history':payment_history,
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

def company_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile_editpage.html', context)
    else:
        return redirect('/')

def company_profile_basicdetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            messages.success(request,'Updated')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
    
def company_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('company_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
       
def company_profile_companydetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            gstno = request.POST.get('gstno')
            profile_pic = request.FILES.get('image')

            # Update the CompanyDetails object with form data
            dash_details.company_name = request.POST.get('cname')
            dash_details.contact = request.POST.get('phone')
            dash_details.address = request.POST.get('address')
            dash_details.city = request.POST.get('city')
            dash_details.state = request.POST.get('state')
            dash_details.country = request.POST.get('country')
            dash_details.pincode = request.POST.get('pincode')
            dash_details.pan_number = request.POST.get('pannumber')

            if gstno:
                dash_details.gst_no = gstno
            else:
                dash_details.gst_no = ''

            if profile_pic:
                dash_details.profile_pic = profile_pic

            dash_details.save()

            messages.success(request, 'Updated')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/')    

# company modules editpage
def company_module_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_module_editpage.html', context)
    else:
        return redirect('/')

def company_module_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Check for any previous module update request
        if ZohoModules.objects.filter(company=dash_details,status='Pending').exists():
            messages.warning(request,'You have a pending update request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            # Retrieve values
            items = request.POST.get('items', 0)
            price_list = request.POST.get('price_list', 0)
            stock_adjustment = request.POST.get('stock_adjustment', 0)
            godown = request.POST.get('godown', 0)

            cash_in_hand = request.POST.get('cash_in_hand', 0)
            offline_banking = request.POST.get('offline_banking', 0)
            upi = request.POST.get('upi', 0)
            bank_holders = request.POST.get('bank_holders', 0)
            cheque = request.POST.get('cheque', 0)
            loan_account = request.POST.get('loan_account', 0)

            customers = request.POST.get('customers', 0)
            invoice = request.POST.get('invoice', 0)
            estimate = request.POST.get('estimate', 0)
            sales_order = request.POST.get('sales_order', 0)
            recurring_invoice = request.POST.get('recurring_invoice', 0)
            retainer_invoice = request.POST.get('retainer_invoice', 0)
            credit_note = request.POST.get('credit_note', 0)
            payment_received = request.POST.get('payment_received', 0)
            delivery_challan = request.POST.get('delivery_challan', 0)

            vendors = request.POST.get('vendors', 0)
            bills = request.POST.get('bills', 0)
            recurring_bills = request.POST.get('recurring_bills', 0)
            vendor_credit = request.POST.get('vendor_credit', 0)
            purchase_order = request.POST.get('purchase_order', 0)
            expenses = request.POST.get('expenses', 0)
            recurring_expenses = request.POST.get('recurring_expenses', 0)
            payment_made = request.POST.get('payment_made', 0)

            projects = request.POST.get('projects', 0)

            chart_of_accounts = request.POST.get('chart_of_accounts', 0)
            manual_journal = request.POST.get('manual_journal', 0)

            eway_bill = request.POST.get('ewaybill', 0)

            employees = request.POST.get('employees', 0)
            employees_loan = request.POST.get('employees_loan', 0)
            holiday = request.POST.get('holiday', 0)
            attendance = request.POST.get('attendance', 0)
            salary_details = request.POST.get('salary_details', 0)

            reports = request.POST.get('reports', 0)

            update_action=1
            status='Pending'

            # Create a new ZohoModules instance and save it to the database
            data = ZohoModules(
                company=dash_details,
                items=items, price_list=price_list, stock_adjustment=stock_adjustment, godown=godown,
                cash_in_hand=cash_in_hand, offline_banking=offline_banking, upi=upi, bank_holders=bank_holders,
                cheque=cheque, loan_account=loan_account,
                customers=customers, invoice=invoice, estimate=estimate, sales_order=sales_order,
                recurring_invoice=recurring_invoice, retainer_invoice=retainer_invoice, credit_note=credit_note,
                payment_received=payment_received, delivery_challan=delivery_challan,
                vendors=vendors, bills=bills, recurring_bills=recurring_bills, vendor_credit=vendor_credit,
                purchase_order=purchase_order, expenses=expenses, recurring_expenses=recurring_expenses,
                payment_made=payment_made,
                projects=projects,
                chart_of_accounts=chart_of_accounts, manual_journal=manual_journal,
                eway_bill=eway_bill,
                employees=employees, employees_loan=employees_loan, holiday=holiday,
                attendance=attendance, salary_details=salary_details,
                reports=reports,update_action=update_action,status=status    
            )
            data.save()
            messages.success(request,"Request sent successfully. Please wait for approval.")
            return redirect('company_profile')
        else:
            return redirect('company_module_editpage')  
    else:
        return redirect('/')


def company_renew_terms(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        # Check for any previous  extension request
        if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists():
            messages.warning(request,'You have a pending request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            select=request.POST['select']
            terms=PaymentTerms.objects.get(id=select)
            update_action=1
            status='Pending'
            newterms=PaymentTermsUpdates(
               company=dash_details,
               payment_term=terms,
               update_action=update_action,
               status=status 
            )
            newterms.save()
            messages.success(request,'Request sent successfully, Please wait for approval...')
            return redirect('company_profile')
        else:
            return redirect('company_profile')
    else:
        return redirect('/')

# company notifications and messages
def company_notifications(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        notifications = dash_details.notifications.filter(is_read=0).order_by('-date_created','-time')
        end_date = dash_details.End_date
        company_days_remaining = (end_date - date.today()).days
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False
        
        print(company_days_remaining)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'notifications':notifications,
            'days_remaining':company_days_remaining,
            'payment_request':payment_request,
        }

        return render(request,'company/company_notifications.html',context)
        
    else:
        return redirect('/')
        
        
def company_message_read(request,pk):
    '''
    message read functions set the is_read to 1, 
    by default it is 0 means not seen by user.

    '''
    notification=Notifications.objects.get(id=pk)
    notification.is_read=1
    notification.save()
    return redirect('company_notifications')
    
    
def company_payment_history(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        payment_history=dash_details.previous_plans.all()

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'payment_history':payment_history,
            
        }
        return render(request,'company/company_payment_history.html', context)
    else:
        return redirect('/')
        
def company_trial_feedback(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        trial_instance = TrialPeriod.objects.get(company=dash_details)
        if request.method == 'POST':
            interested = request.POST.get('interested')
            feedback=request.POST.get('feedback') 
            
            trial_instance.interested_in_buying=1 if interested =='yes' else 2
            trial_instance.feedback=feedback
            trial_instance.save()

            if interested =='yes':
                return redirect('company_profile')
            else:
                return redirect('company_dashboard')
        else:
            return redirect('company_dashboard')
    else:
        return redirect('/')
# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')


def staff_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'staff/staff_profile_editpage.html', context)
    else:
        return redirect('/')

def staff_profile_details_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            dash_details.contact = request.POST.get('phone')
            old=dash_details.image
            new=request.FILES.get('profile_pic')
            print(new,old)
            if old!=None and new==None:
                dash_details.image=old
            else:
                print(new)
                dash_details.image=new
            dash_details.save()
            messages.success(request,'Updated')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

def staff_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('staff_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

# -------------------------------Zoho Modules section--------------------------------

#--------------------------------------------------- TINTO VIEW ITEMS START-------------------------------------------

# items llist
    
def items_list(request):                                                                
     if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'item':item,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/items/items_list.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            context = {
                    'details': dash_details,
                    'item': item,
                    'allmodules': allmodules,
            }
        return render(request,'zohomodules/items/items_list.html',context)

   
   
# create Items

def new_items(request):                                                              
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
                }
                return render(request,'zohomodules/items/newitem.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
            }
    
            return render(request, 'zohomodules/items/newitem.html',context)
# create Items
def create_item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.Date=date.today()
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track = request.POST.get("trackState",None)
            track_state_value = request.POST.get("trackstate", None)

# Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0

            
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if Items.objects.filter(item_name=item_name, company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('items_list')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track_state_value = request.POST.get("trackState", None)

            track_state_value = request.POST.get("trackstate", None)

            # Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0
            minstock=request.POST.get("minimum_stock",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
        
        

        
            if Items.objects.filter(item_name=item_name,company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('items_list')
    return redirect('items_list')

# create unit
def add_unit(request):                                                                #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        if request.method == 'POST':
            c = CompanyDetails.objects.get(login_details=login_id)
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    elif log_user.user_type == 'Staff':
        if request.method == 'POST':
            staff = LoginDetails.objects.get(id=login_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c = sf.company
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    return JsonResponse({"message": "success"})
# create unit


    
def unit_dropdown(request):                                                               
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
      

    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details.company)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
             



def add_account(request):                                                              
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method == 'POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            a=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        
      
        
    return redirect('newitems')

def account_dropdown(request):                                                                
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                account_type=option.account_type
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)
    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
       
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details.company) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)


    
    
def itemsoverview(request,pk):                                                                 
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
              
                items=Items.objects.filter(company=dash_details.company)
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                context = {
                     'details': dash_details,
                
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                }
                return render(request, 'zohomodules/items/itemsoverview.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            items=Items.objects.filter(company=dash_details)
            selitem=Items.objects.get(id=pk)
            est_comments=Items_comments.objects.filter(Items=pk)
            stock_value=selitem.opening_stock*selitem.purchase_price  
            latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
            context = {
                    'details': dash_details,
                   
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
            }
    
            return render(request, 'zohomodules/items/itemsoverview.html',context)


    return render(request, 'zohomodules/items/itemsoverview.html')


def edititems(request, pr):                                                                #new by tinto mt
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    item = get_object_or_404(Items, id=pr)
    

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
      
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
        units = Unit.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        item = get_object_or_404(Items, id=pr)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details)
        units = Unit.objects.filter(company=dash_details)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                   'accounts': accounts,
                    'allmodules': allmodules,
            }
       
    
        
        if request.method=='POST':
   
            b=Item_Transaction_History()
            # c = CompanyDetails.objects.get(login_details=company_id)
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            item.login_details=log_user
            item.company=dash_details
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            if request.POST.get("radio",None) == 'taxable':

                item.intrastate_tax = request.POST.get("intra",None)
                item.interstate_tax= request.POST.get("inter",None)
            elif request.POST.get("radio",None) == 'None-Taxable':
                item.intrastate_tax = 0
                item.interstate_tax= 0
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                item.minimum_stock_to_maintain = 0
            # item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            # item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.opening_stock_per_unit = request.POST.get("rate",None)
            item.current_stock= request.POST.get("openstock",None)
            track_state_value = request.POST.get("trackstate", None)
            if track_state_value == "on":
                item.track_inventory = 1
            else:
                item.track_inventory = 0
            
            # Save the changes
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()
            # Redirect to another page after successful update
            return redirect('itemsoverview', pr)
        return render(request, 'zohomodules/items/edititems.html',context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        item = get_object_or_404(Items, id=pr)
        units = Unit.objects.filter(company=dash_details.company)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                    'accounts': accounts,
                   
                    'allmodules': allmodules,
            }
 
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()

            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            item.intrastate_tax = request.POST.get("intra",None)
            item.interstate_tax= request.POST.get("inter",None)
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                item.minimum_stock_to_maintain = 0
            # item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.current_stock= request.POST.get("openstock",None)
            item.opening_stock_per_unit = request.POST.get("rate",None)
            track_state_value = request.POST.get("trackstate", None)
            if track_state_value == "on":
                item.track_inventory = 1
            else:
                item.track_inventory = 0
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()

            return redirect('itemsoverview', pr)
 
        return render(request, 'zohomodules/items/edititems.html', context)
   
def item_status_edit(request, pv):                                                                #new by tinto mt
    
    selitem = Items.objects.get(id=pv)

    if selitem.activation_tag == 'Active':
        selitem.activation_tag = 'inactive'
        selitem.save()
    elif selitem.activation_tag != 'Active':
        selitem.activation_tag = 'Active'
        selitem.save()

    selitem.save()

    return redirect('itemsoverview',pv)


def shareItemToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                item = Items.objects.get(id=pt)
                context = {
                
                    'selitem':item,
                }
                print('2')
                template_path = 'zohomodules/items/itememailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Item Transactions.pdf'
                subject = f"Transactipns"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Item transactions. \n{email_message}\n\n--\nRegards,\n{item.item_name}\n{item.item_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(itemsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(itemsoverview,pt)   
        
def deleteitem(request,pl):                                                                #new by tinto mt
    items=Items.objects.filter(id=pl)
    items.delete()
    
    return redirect(items_list)

def delete_item_comment(request,ph,pr):                                                                #new by tinto mt
    items=Items_comments.objects.filter(id=ph)
    items.delete()
    ac=Items.objects.get(id=pr)
    
    return redirect(itemsoverview,ac.id)


def add_item_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=Items_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    item=Items.objects.get(id=pc)
                    com.Items=item
                    
                    com.save()
                    return redirect('itemsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=Items_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            item=Items.objects.get(id=pc)
            com.Items=item
                    
            com.save()
            return redirect('itemsoverview',pc)
    return redirect('itemsoverview',pc)
        




         
def downloadItemSampleImportFile(request):                                                                  #new by tinto mt
    estimate_table_data = [['No.','ITEM TYPE','ITEM NAME','HSN','TAX REFERENCE','INTRASTATE TAX','INTERSTATE TAX','SELLING PRICE','SALES ACCOUNT','SALES DESCRIPTION','PURCHASE PRICE','PURCHASE ACCOUNT','PURCHASE DESCRIPTION','MINIMUM STOCK TO MAINTAIN','ACTIVATION TAG','OPENING STOCK','CURRENT STOCK','OPENING STOCK PER UNIT']]      
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    

    # Populate the sheets with data
    for row in estimate_table_data:
        sheet1.append(row)  
    
    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expense_sample_file.xlsx'
     # Save the workbook to the response
    wb.save(response)
    return response





def import_item(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        company_id = request.session['login_id']

        if request.method == 'POST' and 'excel_file' in request.FILES:
            company = CompanyDetails.objects.get(login_details=company_id)
            excel_file = request.FILES.get('excel_file')
            wb = load_workbook(excel_file)

            try:
                ws = wb["Sheet1"]
                header_row = ws[1]
                column_names = [cell.value for cell in header_row]
                print("Column Names:", column_names)
            except KeyError:
                print('Sheet not found')
                messages.error(request, '`Sheet1` not found in the Excel file. Please check.')
                return redirect('expensepage')

            expected_columns = ['No.', 'ITEM TYPE', 'ITEM NAME', 'HSN', 'TAX REFERENCE', 'INTRASTATE TAX', 'INTERSTATE TAX',
                                'SELLING PRICE', 'SALES ACCOUNT', 'SALES DESCRIPTION', 'PURCHASE PRICE',
                                'PURCHASE ACCOUNT', 'PURCHASE DESCRIPTION', 'MINIMUM STOCK TO MAINTAIN', 'ACTIVATION TAG',
                                'OPENING STOCK', 'CURRENT STOCK', 'OPENING STOCK PER UNIT']

            if column_names != expected_columns:
                print('Invalid sheet columns or order')
                messages.error(request, 'Sheet column names or order is not in the required format. Please check.')
                return redirect("comapny_items")

            for row in ws.iter_rows(min_row=2, values_only=True):
                _, item_type, item_name, hsn, tax_reference, intrastate_tax, interstate_tax, selling_price, sales_account, \
                sales_description, purchase_price, purchase_account, purchase_description, min_stock, activation_tag, \
                opening_stock, current_stock, opening_stock_per_unit = row

                # Fetching the 'Unit' instance with id=1 (you may adjust this based on your 'Unit' model)
                unit_instance = Unit.objects.get(pk=1)

                # Creating an instance of the 'Items' model and saving it
                item = Items(
                    login_details=log_user,
                    company=company,
                    unit=unit_instance,  # Use the fetched 'Unit' instance
                    item_type=item_type,
                    item_name=item_name,
                    hsn_code=hsn,
                    tax_reference=tax_reference,
                    intrastate_tax=intrastate_tax,
                    interstate_tax=interstate_tax,
                    selling_price=selling_price,
                    sales_account=sales_account,
                    sales_description=sales_description,
                    purchase_price=purchase_price,
                    purchase_account=purchase_account,
                    purchase_description=purchase_description,
                    minimum_stock_to_maintain=min_stock,
                    activation_tag=activation_tag,
                    inventory_account="Inventory Account",
                    opening_stock=opening_stock,
                    opening_stock_per_unit=opening_stock_per_unit
                )
                item.save()

            messages.success(request, 'Data imported successfully!')
            return redirect("items_list")
        else:
            messages.error(request, 'Invalid request. Please check the file and try again.')
            return redirect("items_list")
    else:
        messages.error(request, 'Invalid user type. Please check your user type.')
        return redirect("items_list")


def item_view_sort_by_name(request, pk):    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
                items = list(Items.objects.filter(company=dash_details.company).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items': sorted_items, 
                    'selitem': selitem, 
                    'stock_value': stock_value, 
                    'latest_item_id': filtered_data, 
                    'est_comments': est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items = list(Items.objects.filter(company=dash_details).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'allmodules': allmodules,
                        'items': sorted_items, 
                        'selitem': selitem, 
                        'stock_value': stock_value, 
                        'latest_item_id': filtered_data, 
                        'est_comments': est_comments
                        
                }  
                return render(request,'zohomodules/items/itemsoverview.html',content)

def item_view_sort_by_hsn(request, pk):      
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
                items = list(Items.objects.filter(company=dash_details.company).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['hsn_code'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items': sorted_items, 
                    'selitem': selitem, 
                    'stock_value': stock_value, 
                    'latest_item_id': filtered_data, 
                    'est_comments': est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items = list(Items.objects.filter(company=dash_details).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'allmodules': allmodules,
                        'items': sorted_items, 
                        'selitem': selitem, 
                        'stock_value': stock_value, 
                        'latest_item_id': filtered_data, 
                        'est_comments': est_comments
                        
                }  
                return render(request,'zohomodules/items/itemsoverview.html',content)

def filter_item_view_Active(request,pk):          
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
           

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='Active',company=dash_details.company)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
         

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='Active',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content) 

def filter_item_view_inActive(request,pk):         
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
           

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='inactive',company=dash_details.company)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
         

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='inactive',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content) 

    
    #--------------------------------------------------- TINTO VIEW ITEMS END-------------------------------------------


        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS START-------------------------------------------
def addchartofaccounts(request):                                                                #new by tinto mt
        if 'login_id' in request.session:
            login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
        
                    'allmodules': allmodules,
         
                }
                return render(request,'zohomodules/chartofaccounts/addchartofaccounts.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
          
                    'allmodules': allmodules,
           
            }
    
            return render(request,'zohomodules/chartofaccounts/addchartofaccounts.html',context)


def chartofaccounts(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
  

def create_account(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            account=Chart_of_Accounts.objects.all()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       

            
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits !!!')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            account_type=request.POST.get("account_type",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')

    return redirect('addchartofaccounts')

def chartofaccountsoverview(request,pk):                                                                #new by tinto mt
       if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                    dash_details = StaffDetails.objects.get(login_details=log_details)

                    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                
                    acc=Chart_of_Accounts.objects.filter(company=dash_details.company)  
                    selacc=Chart_of_Accounts.objects.get(id=pk)  
                    est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
                    latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                    filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                    context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
                    return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            acc=Chart_of_Accounts.objects.filter(company=dash_details)  
            selacc=Chart_of_Accounts.objects.get(id=pk)  
            est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
            latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
            context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
    
            return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)



   
        
    # acc=Chart_of_Accounts.objects.all()  
    # selacc=Chart_of_Accounts.objects.get(id=pk)  
    # est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
    # latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
    # filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
    # return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',{'acc':acc,'selacc':selacc,'latest_item_id':filtered_data,'est_comments':est_comments})


from django.shortcuts import render, redirect

def editchartofaccounts(request, pr):                                                                #new by tinto mt
    # Retrieve the chart of accounts entry
    

    # Check if 'company_id' is in the session
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    acc = get_object_or_404(Chart_of_Accounts, id=pr)

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
       
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
       
    
        
        

        if request.method == 'POST':
        
            b=Chart_of_Accounts_History()
       
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=dash_details
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.account_description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccounts/editchartofaccounts.html', context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
        if request.method=='POST':
         
            b=Chart_of_Accounts_History()
         
            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=c
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.account_description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccounts/editchartofaccounts.html', context)

def deleteaccount(request,pl):                                                                #new by tinto mt
    acc=Chart_of_Accounts.objects.filter(id=pl)
    acc.delete()
    
    return redirect(chartofaccounts)


def acc_status_edit(request, pv):                                                                #new by tinto mt
    
    selacc = Chart_of_Accounts.objects.get(id=pv)

    if selacc.status == 'Active':
        selacc.status = 'inactive'
        selacc.save()
    elif selacc.status != 'Active':
        selacc.status = 'Active'
        selacc.save()

    selacc.save()

    return redirect('chartofaccountsoverview',pv)


def add_account_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=chart_of_accounts_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    acc=Chart_of_Accounts.objects.get(id=pc)
                    com.chart_of_accounts=acc
                    
                    com.save()
                    return redirect('chartofaccountsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=chart_of_accounts_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            acc=Chart_of_Accounts.objects.get(id=pc)
            com.chart_of_accounts=acc
                    
            com.save()
            return redirect('chartofaccountsoverview',pc)


def delete_account_comment(request,ph,pr):                                                                #new by tinto mt
    acc=chart_of_accounts_comments.objects.filter(id=ph)
    acc.delete()
    ac=Chart_of_Accounts.objects.get(id=pr)
    
    return redirect(chartofaccountsoverview,ac.id)

def account_view_sort_by_name(request,pk):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc = Chart_of_Accounts.objects.filter(company=dash_details.company).order_by('account_name')
                selacc = Chart_of_Accounts.objects.get(id=pk)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc = Chart_of_Accounts.objects.filter(company=dash_details).order_by('account_name')
                selacc = Chart_of_Accounts.objects.get(id=pk)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)


def shareaccountToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                acc = Chart_of_Accounts.objects.get(id=pt)
                context = {
                
                    'selacc':acc,
                }
                print('2')
                template_path = 'zohomodules/chartofaccounts/accountemailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Account Details.pdf'
                subject = f"Account"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Account Details. \n{email_message}\n\n--\nRegards,\n{acc.account_name}\n{acc.account_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(chartofaccountsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(chartofaccountsoverview,pt)
        
        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS END-------------------------------------------
        
        
def chartofaccountsActive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="active")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
def chartofaccountsInactive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="inactive")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details,status="inactive")
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
            
#------------------------------------payroll employee--------------------------------
#------------------------------------------------GEORGE MATHEW---------------------------------------
def payroll_employee_create(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    blood=Bloodgroup.objects.all()
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'log_id':log_details,
            'blood':blood
            
    }
    return render(request,'zohomodules/payroll-employee/payroll_create_employee.html',content)
    
def employee_list(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
        
def employee_overview(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type =='Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=payroll_employee_comment.objects.filter(employee=pk)
        history=employee_history.objects.filter(employee=pk)
    if log_details.user_type =='Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=payroll_employee_comment.objects.filter(employee=pk)
        history=employee_history.objects.filter(employee=pk)
    content = {
                'details': dash_details,
                'pay':pay,
                'p':p,
                'allmodules': allmodules,
                'comment':comment_data,
                'history':history,
                'log_id':log_details,
        }
    return render(request,'zohomodules/payroll-employee/overview_page.html',content)
    
def create_employee(request):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':    
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                            emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                            amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                            UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details,login_details=log_details)
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='CREATED')
                history.save()
                messages.info(request,'employee created')
                return redirect('employee_list')
        if log_details.user_type == 'Staff':
            company_details = StaffDetails.objects.get(login_details=log_details)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details.company):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'emerency phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details.company,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details.company):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            dash_details = StaffDetails.objects.get(login_details=log_details)
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                         emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                         amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                         UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details.company,login_details=log_details)
            payroll.save()
            history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            messages.info(request,'employee created')
            return redirect('employee_list')
    return redirect('payroll_employee_create')
    
def payroll_employee_edit(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    blood=Bloodgroup.objects.all()
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        
    print(p)
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'p':p,
            'log_id':log_details,
            'blood':blood
    }
    return render(request,'zohomodules/payroll-employee/edit_employee.html',content)
    
def do_payroll_edit(request,pk):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type =='Company':
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)    
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed' or saltype =='Temporary'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours']
            empnum=request.POST['empnum']
            result_set2 = payroll_employee.objects.filter(company=company_details,emp_number=empnum).exclude(id=pk)
            if result_set2:
                messages.error(request,'employee number  already exists')
                return redirect('payroll_employee_edit',pk)
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone).exclude(id=pk)
            result_set3 = payroll_employee.objects.filter(company=company_details,emergency_phone=phone).exclude(id=pk)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_edit',pk)
            if result_set3:
                messages.error(request,'emergency phone no already exists')
                return redirect('payroll_employee_edit',pk)
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email).exclude(id=pk)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_edit',pk)
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an'] 
            if payroll_employee.objects.filter(Aadhar=an,company=company_details).exclude(id=pk):
                messages.error(request,'Aadhra number already exists')
                return redirect('payroll_employee_edit',pk)
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
        if log_details.user_type == 'Staff':
            if log_details.user_type =='Staff':
                company_details = StaffDetails.objects.get(login_details=log_details)    
                title=request.POST['title']
                fname=request.POST['fname']
                lname=request.POST['lname']
                alias=request.POST['alias']
                joindate=request.POST['joindate']
                salarydate=request.POST['salary']
                saltype=request.POST['saltype']
                if (saltype == 'Fixed' or saltype =='Temporary'):
                    salary=request.POST['fsalary']
                else:
                    salary=request.POST['vsalary']
                image=request.FILES.get('file')
                amountperhr=request.POST['amnthr']
                workhr=request.POST['hours']
                empnum=request.POST['empnum']
                result_set2 = payroll_employee.objects.filter(company=company_details.company,emp_number=empnum).exclude(id=pk)
                if result_set2:
                    messages.error(request,'employee number  already exists')
                    return redirect('payroll_employee_edit',pk)
                designation = request.POST['designation']
                location=request.POST['location']
                gender=request.POST['gender']
                dob=request.POST['dob']
                blood=request.POST['blood']
                fmname=request.POST['fm_name']
                sname=request.POST['s_name']        
                add1=request.POST['address']
                add2=request.POST['address2']
                address=add1+" "+add2
                padd1=request.POST['paddress'] 
                padd2=request.POST['paddress2'] 
                paddress= padd1+padd2
                phone=request.POST['phone']
                ephone=request.POST['ephone']
                result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone).exclude(id=pk)
                result_set3 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone).exclude(id=pk)
                if result_set1:
                    messages.error(request,'phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                if result_set3:
                    messages.error(request,'emergency phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                email=request.POST['email']
                result_set = payroll_employee.objects.filter(company=company_details.company,email=email).exclude(id=pk)
                if result_set:
                    messages.error(request,'email already exists')
                    return redirect('payroll_employee_edit',pk)
                isdts=request.POST['tds']
                attach=request.FILES.get('attach')
                if isdts == '1':
                    istdsval=request.POST['pora']
                    if istdsval == 'Percentage':
                        tds=request.POST['pcnt']
                    elif istdsval == 'Amount':
                        tds=request.POST['amnt']
                else:
                    istdsval='No'
                    tds = 0
                itn=request.POST['itn']
                an=request.POST['an'] 
                if payroll_employee.objects.filter(Aadhar=an,company=company_details.company).exclude(id=pk):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_edit',pk)
                uan=request.POST['uan'] 
                pfn=request.POST['pfn']
                pran=request.POST['pran']
                age=request.POST['age']
                bank=request.POST['bank']
                accno=request.POST['acc_no']       
                ifsc=request.POST['ifsc']       
                bname=request.POST['b_name']       
                branch=request.POST['branch']
                ttype=request.POST['ttype']
                dash_details = StaffDetails.objects.get(login_details=log_details)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details.company
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
    
def add_comment(request,pk):
    if request.method =='POST':
        comment_data=request.POST['comments']
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        payroll= payroll_employee.objects.get(id=pk) 
        data=payroll_employee_comment(comment=comment_data,login_details=log_details,employee=payroll)
        data.save()
        return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
    
def delete_commet(request,pk,pi):
    data=payroll_employee_comment.objects.get(id=pk)
    data.delete()
    return redirect('employee_overview',pi)
    
def delete_employee(request,pk):
    data=payroll_employee.objects.get(id=pk)
    data.delete()
    return redirect('employee_list')
    
def employee_status(request,pk):
    data=payroll_employee.objects.get(id=pk)
    if data.status == 'Active':
        data.status ='Inactive'
    elif data.status == 'Inactive':
        data.status ='Active'
    data.save()
    return redirect('employee_overview',pk)
    
def add_blood(request):
    if request.method == 'POST':
        blood = request.POST.get('blood')
        print(blood)

        # Validate input
        if not blood:
            return JsonResponse({'message': 'Invalid or missing blood group'})

        # Use get_or_create for simplicity
        if Bloodgroup.objects.filter(Blood_group=blood):
            return JsonResponse({'message': 'Blood group already exists'})
        Bloodgroup.objects.create(Blood_group=blood)
        data=Bloodgroup.objects.all()
        return JsonResponse({'message': 'Blood group added','blood' : blood})
        
def import_payroll_excel(request):
    print(1)
    print('hello')
    if request.method == 'POST' :
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=datetime.date.fromisoformat(billsheet[4]),salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=datetime.date.fromisoformat(billsheet[14]),blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details.company,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=billsheet[4],salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=billsheet[14],blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
    messages.error(request,'File upload Failed!11')
    return redirect('employee_list')
    
def add_file(request,pk):
    if request.method == 'POST':
        data=request.FILES.get('file')
        payroll=payroll_employee.objects.get(id=pk)
        if payroll.uploaded_file:
            try:
                                # Check if the file exists before removing it
                if os.path.exists(payroll.uploaded_file.path):
                    os.remove(payroll.uploaded_file.path)
            except Exception as e:
                messages.error(request,'file upload error')
                return redirect('employee_overview',pk)

                            # Assign the new file to payroll.image
            payroll.uploaded_file = data
            payroll.save()
            messages.info(request,'fil uploaded')
            return redirect('employee_overview',pk)
        else:
            payroll.uploaded_file = data
            payroll.save()
        messages.info(request,'fil uploaded')
        return redirect('employee_overview',pk)
        
def shareemail(request,pk):
    try:
            if request.method == 'POST':
                emails_string = request.POST['email']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                print(emails_list)
                p=payroll_employee.objects.get(id=pk)
                        
                context = {'p':p}
                template_path = 'zohomodules/payroll-employee/mailoverview.html'
                template = get_template(template_path)
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'{p.first_name}details - {p.id}.pdf'
                subject = f"{p.first_name}{p.last_name}  - {p.id}-details"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached employee details - File-{p.first_name}{p.last_name} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)
                messages.success(request, 'over view page has been shared via email successfully..!')
                return redirect('employee_overview',pk)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('employee_overview',pk)
#----------------------------------------------------------end----------------------------------------


def accounts_asset_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_liability_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)


def accounts_equity_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Equity"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Equity"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_income_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Income","Other Income"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Income","Other Income"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
def accounts_expense_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
            
def account_view_sort_by_namelist(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                # acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                acc = Chart_of_Accounts.objects.filter(company=dash_details.company).order_by('account_name')
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            # acc=Chart_of_Accounts.objects.filter(company=dash_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details).order_by('account_name')
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
            
def account_view_filterActive(request,ph):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="active")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details,status="active")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
                
                
def account_view_filterinActive(request,ph):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="inactive")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details,status="inactive")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
                
                
#----------------- Banking -----------------------------#

def bank_list(request):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    return render(request, 'bank_list.html',{'bnk':bnk, 'allmodules':allmodules, 'details':dash_details})

def load_bank_create(request):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    tod = datetime.now().strftime('%Y-%m-%d')
    return render(request, 'bank_create.html',{'tod':tod, 'allmodules':allmodules, 'details':dash_details})

def bank_create(request):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        date = request.POST.get('date')
        name = request.POST.get('name')
        opn_bal = request.POST.get('opn_bal')
        if opn_bal != '':
            opn_bal = float(opn_bal)
        else:
            opn_bal = 0
        bal_type=request.POST.get('bal_type')
        branch= request.POST.get('branch')
        ac_no= request.POST.get('ac_no')
        ifsc=request.POST.get('ifsc')

        if Banking.objects.filter(company = cmp, bnk_acno = ac_no).exists():
            messages.info(request, 'Bank with this account number already exsist !!')
            return redirect('bank_list')

        bank = Banking.objects.create(
            login_details = log_details,
            company = cmp,
            bnk_name=name,
            bnk_bal_type = bal_type,
            bnk_opnbal=opn_bal,
            bnk_bal=opn_bal,
            bnk_branch=branch,
            bnk_acno=ac_no,
            bnk_ifsc=ifsc,
            date=date
        )

        bank.save()

        BankTransaction.objects.create( login_details=log_details, company=cmp, banking=bank, trans_amount=opn_bal, trans_adj_date=date, 
                                       trans_type='Opening Balance', trans_adj_type='', trans_adj_amount = opn_bal)

        BankingHistory.objects.create( login_details=log_details, company=cmp, banking=bank, hist_adj_amount = 0, hist_adj_date=date, hist_action='Created')

        return redirect('bank_list')
    
def bank_view(request, id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    allmodules= ZohoModules.objects.get(company = cmp)
    tod = datetime.now().strftime('%Y-%m-%d')
    bnk_list = Banking.objects.filter(company=cmp)
    bnk = Banking.objects.get(id=id, company=cmp)
    show_div = request.GET.get('Transaction', False)
    trans = BankTransaction.objects.filter(company=cmp, banking = bnk)
    hist_f = BankingHistory.objects.get(company = cmp, banking = bnk, hist_action = 'Created')
    hist_l = BankingHistory.objects.filter(company = cmp, banking = bnk).order_by('-id').first()
    context = {'cmp': cmp, 'bnk': bnk, 'bnk_list':bnk_list, 'trans':trans, 'tod':tod, 'id':id, 'show_div':show_div,
               'allmodules':allmodules, 'details':dash_details, 'hist_f':hist_f, 'hist_l':hist_l}
    return render(request, 'bank_view.html', context)

def bank_transaction_create(request, id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        date = request.POST.get('date')
        origin = request.POST.get('origin')
        dest = request.POST.get('dest')
        amount = float(request.POST.get('amount', 0.0))
        description = request.POST.get('description')
        ttype = request.POST.get('type')
        adjtype = request.POST.get('adjtype') 
        adjacname = request.POST.get('adjacname')
                    
        if ttype == 'Bank To Cash Transfer':
            origin_bnk = Banking.objects.get(id=origin)
            bal = float(origin_bnk.bnk_bal) - float(amount)
            thist =  BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = origin_bnk,
                trans_cur_amount = origin_bnk.bnk_opnbal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Cash',
                trans_adj_type = 'Balance Decrease'
            )
            BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=thist, hist_cur_amount=origin_bnk.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            origin_bnk.bnk_bal = bal
            origin_bnk.save()

        if ttype == 'Cash To Bank Transfer':
            dest_bnk = Banking.objects.get(id=dest)
            bal = float(dest_bnk.bnk_bal) + float(amount)
            thist = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = dest_bnk,
                trans_cur_amount = dest_bnk.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Cash to Bank',
                trans_adj_type = 'Balance Increase'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist, hist_cur_amount=dest_bnk.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            dest_bnk.bnk_bal = bal
            dest_bnk.save()
            
        if ttype == 'Bank To Bank Transfer':
            from_bank = Banking.objects.get(id=origin)
            to_bank = Banking.objects.get(id=dest)
            bal = float(from_bank.bnk_bal) - float(amount)
            thist1 = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = from_bank,
                trans_cur_amount = from_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Bank',
                trans_adj_type = 'Balance Decrease'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist1, hist_cur_amount=from_bank.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            from_bank.bnk_bal = bal
            from_bank.save()

            bal = float(to_bank.bnk_bal) + float(amount)
            thist2 = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = to_bank,
                trans_cur_amount = to_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Bank',
                trans_adj_type = 'Balance Increase'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist2, hist_cur_amount=to_bank.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            to_bank.bnk_bal = bal
            to_bank.save()

            thist1.bank_to_bank_no = thist2.id
            thist2.bank_to_bank_no = thist1.id
            thist1.save()
            thist2.save()
                    
        if ttype == 'Adjust Bank Balance':
            adj_bank = Banking.objects.get(id=adjacname)
            if adjtype == 'Reduce Balance':
                adj_type = 'Balance Decrease'
                bal = float(adj_bank.bnk_bal) - float(amount)
            else:
                adj_type = 'Balance Increase'
                bal = float(adj_bank.bnk_bal) + float(amount)
            thist = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = adj_bank,
                trans_cur_amount = adj_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank Adjustment',
                trans_adj_type = adj_type
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist, hist_cur_amount = adj_bank.bnk_bal, hist_amount = amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            adj_bank.bnk_bal = bal
            adj_bank.save()
        url = reverse('bank_view', kwargs={'id': id}) + '?Transaction=True'
        return redirect(url)
    
def load_trans_details(request):
    id = request.POST.get('id')
    trans = BankTransaction.objects.get(id=id)
    bnk_org_name = ''
    bnk_dest_name = ''
    if trans.trans_type == 'Bank to Bank':
        dest_trans = BankTransaction.objects.get(id=trans.bank_to_bank_no)
        bnk_org_name = trans.banking.id
        bnk_dest_name = dest_trans.banking.id
    else:
        bnk_org_name = trans.banking.id
        bnk_dest_name = trans.banking.id

    return JsonResponse({'message':'success', 
                         'ttype':trans.trans_type,
                         'origin':bnk_org_name,
                         'destination':bnk_dest_name,
                         'trans_id':trans.id,
                         'trans_adj_type':trans.trans_adj_type,
                         'trans_desc':trans.trans_desc,
                         'trans_adj_date':trans.trans_adj_date,
                         'trans_amount':trans.trans_amount})

def bank_transaction_edit(request):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        id = request.POST.get('id')
        origin = request.POST.get('origin')
        dest = request.POST.get('dest')
        amount = float((request.POST.get('amount',0.0)))
        edit_date = request.POST.get('edit_date')
        desc = request.POST.get('desc')
        ttype = request.POST.get('type')
        adjtype = request.POST.get('adjtype') 
        adjacname = request.POST.get('adjacname')
                    
        # If it is Bank to Cash Transfer
        if ttype == 'Bank To Cash Transfer':
            htrans = BankTransaction.objects.get(id=id)
            origin_bnk = Banking.objects.get(id=origin)
            # If we did not change the Bank while editing
            if htrans.banking.bnk_acno == origin_bnk.bnk_acno:
                # Changing balance using edit amount
                bal = float(htrans.trans_cur_amount) - float(amount)
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.save()
                # Creating a transaction history for the edit
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                # Obtaning all transaction done after the edited transaction for that particular bank
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                # Creating transaction history for all the transactions done after the edited transaction by calculating new balance
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                    
                # Changing final balance for the bank
                origin_bnk.bnk_bal = newbal
                origin_bnk.save()
            # If we changed the Bank while editing
            else:
                origin_bnk = Banking.objects.get(id=origin)
                htrans = BankTransaction.objects.get(id=id)
                # Deleting Transaction history of the transaction
                BankTransactionHistory.objects.filter(transaction = htrans).delete()

                # Obtaning all transaction done after the edited transaction for that particular bank
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                # Creating transaction history for all the transactions done after the edited transaction by calculating new balance
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                    hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                # Changing final balance for the bank
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(origin_bnk.bnk_bal) - float(amount)
                # Creating a transaction for the new bank
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = origin_bnk,
                    trans_type = htrans.trans_type,
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = origin_bnk.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                origin_bnk.bnk_bal = bal
                origin_bnk.save()
                # Creating a transaction history for the new bank
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                # Deleting old bank transaction
                htrans.delete()

        # If it is Cash to Bank Transfer
        if ttype == 'Cash To Bank Transfer':
            htrans = BankTransaction.objects.get(id=id)
            dest_bnk = Banking.objects.get(id=dest)
            if htrans.banking.bnk_acno == dest_bnk.bnk_acno:
                bal = float(htrans.trans_cur_amount) + float(amount)
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                dest_bnk.bnk_bal = newbal
                dest_bnk.save()
            else:
                dest_bnk = Banking.objects.get(id=dest)
                htrans = BankTransaction.objects.get(id=id)
                BankTransactionHistory.objects.filter(transaction = htrans).delete()

                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(dest_bnk.bnk_bal) + float(amount)
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = dest_bnk,
                    trans_type = htrans.trans_type,
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = dest_bnk.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                dest_bnk.bnk_bal = bal
                dest_bnk.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                htrans.delete()

        # If it is Bank to Bank Transfer   
        if ttype == 'Bank To Bank Transfer':
            htrans_in = BankTransaction.objects.get(id=id)
            htrans_de = BankTransaction.objects.get(id=htrans_in.bank_to_bank_no)

            if htrans_in.trans_adj_type == 'Balance Decrease':
                red_bal = float(htrans_in.trans_cur_amount) - float(amount)
                add_bal = float(htrans_de.trans_cur_amount) + float(amount)
            else:
                red_bal = float(htrans_in.trans_cur_amount) + float(amount)
                add_bal = float(htrans_de.trans_cur_amount) - float(amount)

            origin_bnk = Banking.objects.get(id=origin)
            dest_bnk = Banking.objects.get(id=dest)

            # If Origin Bank is not changed
            if htrans_in.banking.bnk_acno == origin_bnk.bnk_acno:
                # If destination bank is not changed
                if htrans_in.banking.bnk_acno == dest_bnk.bnk_acno:
                    htrans_in.trans_amount = amount
                    htrans_in.trans_adj_amount = red_bal
                    htrans_in.trans_desc = desc
                    htrans_in.trans_adj_date = edit_date
                    htrans_in.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans_in, hist_cur_amount = htrans_in.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = red_bal, hist_action='Updated')
                    newbal = red_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=origin_bnk, id__gt=htrans_in.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    origin_bnk.bnk_bal = newbal
                    origin_bnk.save()

                    htrans_de.trans_amount = amount
                    htrans_de.trans_adj_amount = add_bal
                    htrans_de.trans_desc = desc
                    htrans_de.trans_adj_date = edit_date
                    htrans_de.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans_de, hist_cur_amount = htrans_de.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = add_bal, hist_action='Updated')
                    newbal = add_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=dest_bnk, id__gt=htrans_de.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    dest_bnk.bnk_bal = newbal
                    dest_bnk.save()
                #if destination bank is changed
                else:
                    htrans_in.trans_amount = amount
                    htrans_in.trans_adj_amount = red_bal
                    htrans_in.trans_desc = desc
                    htrans_in.trans_adj_date = edit_date
                    htrans_in.save()
                    BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=htrans_in, hist_cur_amount=htrans_in.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = red_bal, hist_action='Updated')
                    newbal = red_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=origin_bnk, id__gt=htrans_in.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    origin_bnk.bnk_bal = newbal
                    origin_bnk.save()


                    BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                    newbal = htrans_de.trans_cur_amount
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    htrans_de.banking.bnk_bal = newbal
                    htrans_de.banking.save()

                    if htrans_de.trans_adj_type == 'Balance Decrease':
                        bal = float(dest_bnk.bnk_bal) - float(amount)
                    else:
                        bal = float(dest_bnk.bnk_bal) + float(amount)

                    newtrans = BankTransaction.objects.create(
                        login_details=log_details, 
                        company=cmp,
                        banking = dest_bnk,
                        trans_type = htrans_de.trans_type,
                        trans_adj_type = htrans_de.trans_adj_type,
                        origin = origin_bnk.bnk_name,
                        origin_accno = origin_bnk.bnk_acno,
                        destination = dest_bnk.bnk_name,
                        destination_accno = dest_bnk.bnk_acno,
                        trans_cur_amount = dest_bnk.bnk_bal,
                        trans_amount = amount,
                        trans_adj_amount = bal,
                        trans_desc = desc,
                        trans_adj_date = edit_date,
                        bank_to_bank_no = htrans_in.id
                    )
                    dest_bnk.bnk_bal = bal
                    dest_bnk.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                            hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                    htrans_de.delete()
            # if origin bank is changed 
            else:
                # if destination bank is not changed
                if htrans_in.banking.bnk_acno == dest_bnk.bnk_acno:
                    htrans_de.trans_amount = amount
                    htrans_de.trans_adj_amount = add_bal
                    htrans_de.trans_desc = desc
                    htrans_de.trans_adj_date = edit_date
                    htrans_de.save()
                    BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=htrans_de, hist_cur_amount=htrans_de.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = add_bal, hist_action='Updated')
                    newbal = add_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=dest_bnk, id__gt=htrans_de.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    dest_bnk.bnk_bal = newbal
                    dest_bnk.save()

                    BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                    newbal = htrans_in.trans_cur_amount
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    htrans_in.banking.bnk_bal = newbal
                    htrans_in.banking.save()

                    if htrans_in.trans_adj_type == 'Balance Decrease':
                        bal = float(origin_bnk.bnk_bal) - float(amount)
                    else:
                        bal = float(origin_bnk.bnk_bal) + float(amount)

                    newtrans = BankTransaction.objects.create(
                        login_details=log_details, 
                        company=cmp,
                        banking = origin_bnk,
                        trans_type = htrans_in.trans_type,
                        trans_adj_type = htrans_in.trans_adj_type,
                        trans_cur_amount = origin_bnk.bnk_bal,
                        trans_amount = amount,
                        trans_adj_amount = bal,
                        trans_desc = desc,
                        trans_adj_date = edit_date,
                        bank_to_bank_no = htrans_de.id
                    )
                    origin_bnk.bnk_bal = bal
                    origin_bnk.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                            hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                    htrans_in.delete()
                # if destination bank is changed
                else:
                    # if new origin bank is not previous destination bank and new destiantion bank is not previous origin bank
                    if origin_bnk.bnk_acno != htrans_in.banking.bnk_acno and dest_bnk.bnk_acno != htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(origin_bnk.bnk_bal) - float(amount)
                        else:
                            bal = float(origin_bnk.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = origin_bnk.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=newtrans_or, hist_amount=amount, 
                                                              hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')
                        htrans_in.delete()

                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(dest_bnk.bnk_bal) - float(amount)
                        else:
                            bal = float(dest_bnk.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = dest_bnk.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                        htrans_de.delete()

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                    
                    # if new origin bank is equal to previous destination bank and new destiantion bank is not previous origin bank
                    elif origin_bnk.bnk_acno == htrans_in.banking.bnk_acno and dest_bnk.bnk_acno != htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                                                        
                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_or, hist_cur_amount=newtrans_or.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

                    # if new origin bank is not previous destination bank and new destiantion bank is equal to previous origin bank
                    elif origin_bnk.bnk_acno != htrans_in.banking.bnk_acno and dest_bnk.bnk_acno == htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                                                        
                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_or, hist_cur_amount=newtrans_or.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

                    # if new origin bank is equal to previous destination bank and new destiantion bank is equal to previous origin bank
                    else:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                           
                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = htrans_in.banking,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=newtrans_or, hist_amount=amount, 
                                                              hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = htrans_de.banking,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_amount=amount, 
                                                                hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

        # If it is Adjust Bank Balance           
        if ttype == 'Adjust Bank Balance':
            htrans = BankTransaction.objects.get(id=id)
            adj_bank = Banking.objects.get(id=adjacname)
            if adjtype == 'Reduce Balance':
                adj_type = 'Balance Decrease'
                bal = float(htrans.trans_cur_amount) - float(amount)
            else:
                adj_type = 'Balance Increase'
                bal = float(htrans.trans_cur_amount) + float(amount)
            if htrans.banking.bnk_acno == adj_bank.bnk_acno:
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.trans_adj_type = adj_type
                htrans.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                adj_bank.bnk_bal = newbal
                adj_bank.save()
            else:
                htrans = BankTransaction.objects.get(id=id)
                adj_bank = Banking.objects.get(id=adjacname)
                BankTransactionHistory.objects.filter(transaction = htrans).delete()
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(adj_bank.bnk_bal) - float(amount)
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = adj_bank,
                    trans_type = 'Bank Adjustment',
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = adj_bank.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                adj_bank.bnk_bal = bal
                adj_bank.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                htrans.delete()
        return JsonResponse({'message':'success'})
    
def load_bank_edit(request, id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    bnk = Banking.objects.get(id=id)
    return render(request, 'bank_edit.html', {'bnk':bnk, 'allmodules':allmodules, 'details':dash_details})

def bank_edit(request,id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        bnk = Banking.objects.get(id=id)
        bnk.bnk_name = request.POST.get('name')
        bnk.bnk_branch = request.POST.get('branch')
        bnk.bnk_acno = request.POST.get('ac_no')
        bnk.bnk_ifsc = request.POST.get('ifsc')
        bnk.bnk_bal_type = request.POST.get('bal_type')
        bnk.date = request.POST.get('date')
        newbal = request.POST.get('opn_bal')
        if newbal != '':
            newbal = float(newbal)
        else:
            newbal = 0
        bnk.bnk_opnbal = newbal
        bnk.save()
        BankingHistory.objects.create(login_details=log_details, company=cmp, banking=bnk, hist_adj_amount=request.POST.get('opn_bal', 0.0), hist_action='Updated')

        trans_list = BankTransaction.objects.filter(company=cmp, banking=bnk)
        for t in trans_list:
            hist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal,
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
            if t.trans_type == 'Opening Balance':
                t.trans_amount = newbal
                t.trans_adj_amount = newbal
                t.trans_adj_date = bnk.date
            else:
                t.trans_cur_amount = newbal
                t.trans_adj_date = bnk.date
                if t.trans_adj_type == 'Balance Increase':
                    newbal = float(t.trans_amount) + newbal
                else:
                    newbal = newbal - float(t.trans_amount)  
                t.trans_adj_amount = newbal
            t.save()
            hist.hist_adj_amount = newbal
            hist.save()
        bnk.bnk_bal = newbal
        bnk.save()
        return redirect('bank_view',id)
    return redirect('bank_list')

def load_bank_history(request,id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    bhis = BankingHistory.objects.filter(company=cmp, banking=id)
    return render(request, 'bank_history.html', {'allmodules':allmodules, 'id':id, 'bhis':bhis, 'details':dash_details})

def load_bank_trans_history(request,id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    thist = BankTransactionHistory.objects.filter(transaction=id)
    bnk_id = thist[0].transaction.banking.id
    return render(request, 'bank_trans_history.html', {'allmodules':allmodules, 'thist':thist, 'id':bnk_id, 'details':dash_details})

def delete_banking(request, id):
    bnk = Banking.objects.get(id=id)
    if BankTransaction.objects.filter(banking=bnk).count() > 1:
        messages.info(request,'This bank cannot be deleted as it has done some transactions !!')
        return redirect('bank_view',id)
    bnk.delete()
    return redirect ('bank_list')

def delete_transaction(request, id):
    cp = company_details.objects.get(user=request.user)
    transaction = get_object_or_404(transactions, id=id)
    
    bank = transaction.bank
    from_bank_id = transaction.fromB
    to_bank_id=transaction.toB
    print(f"Original Bank : {bank.balance}")
    
    print(f"From Bank ID: {from_bank_id}")
    print(f"To Bank ID: {to_bank_id}")

    # Update the bank balance based on the type of transaction
    if transaction.adjtype == 'Increase Balance':
        bank.balance -= transaction.amount
    elif transaction.adjtype == 'Reduce Balance':
        bank.balance += transaction.amount
    elif transaction.type == 'Bank To Cash Transfer':
        bank.balance += transaction.amount
    elif transaction.type == 'Cash To Bank Transfer':
        bank.balance -= transaction.amount
    # elif transaction.type == 'Bank To Bank Transfer':
        # from_bank = Bankcreation.objects.get(id=transaction.fromB,user=request.user)
        # to_bank = Bankcreation.objects.get(id=transaction.toB,user=request.user)
        # from_bank = transaction.fromB
        # to_bank =transaction.toB
        # from_bank = Bankcreation.objects.get(id=from_bank_id)
        # to_bank= Bankcreation.objects.get(id=to_bank_id)
        # print( from_bank)
        # print( to_bank)

        
        # from_bank.balance += transaction.amount
        # to_bank.balance -= transaction.amount
        # # bank.balance+= transaction.amount
       
        # from_bank_id.save()
        # to_bank_id.save()
        # try:
            # from_bank = Bankcreation.objects.get(id=from_bank_id.id,user=request.user)
            # to_bank = Bankcreation.objects.get(id=to_bank_id.id,user=request.user)

            # print(f"From Bank Balance Before Transfer: {from_bank.balance}")
            # print(f"To Bank Balance Before Transfer: {to_bank.balance}")
        # print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        # print(f"To Bank Balance After Transfer: {to_bank_id.balance}")
        # from_bank_id.balance += transaction.amount
        # to_bank_id.balance -= transaction.amount


        # print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        # print(f"To Bank Balance After Transfer: {to_bank_id.balance}")

        # from_bank_id.save()
        # to_bank_id.save()
        # print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        # print(f"To Bank Balance After Transfer: {to_bank_id.balance}")
        # except Bankcreation.DoesNotExist:
        #     raise Http404("Bank does not exist")


        # print(f"After Transfer - From Bank Balance: {from_bank.balance}, To Bank Balance: {to_bank.balance}")
    bank.save()
   

    previous_transaction = transactions.objects.filter(bank=bank, id__lt=transaction.id).order_by('-id').first()
    subsequent_transactions = transactions.objects.filter(bank=bank, id__gt=transaction.id).order_by('id')
    cumulative_balance_change = previous_transaction.balance if previous_transaction else 0

    for sub_transaction in subsequent_transactions:
        if sub_transaction.type == 'Bank To Bank Transfer':
            # try:
            #     to_bank = Bankcreation.objects.get(name=sub_transaction.toB.name, user=request.user)
            #     print(f"To Bank: {sub_transaction.toB.name}")
            # except Bankcreation.DoesNotExist:
            #     # Handle the case where to_bank is not found
            #     to_bank = None
            # print("Before try-except block")
            try:
                to_bank = Bankcreation.objects.get(id=sub_transaction.toB.id, user=request.user)
                print(f"To Bank: {sub_transaction.toB.id}")
                print(to_bank)
            except Bankcreation.DoesNotExist:
                print("Bankcreation.DoesNotExist exception caught")
                to_bank = None

            print("After try-except block")

            if to_bank:
                # if sub_transaction.toB.id == to_bank:
                cumulative_balance_change -= sub_transaction.amount
            else:
                # Handle the case where to_bank is not found
                pass
            
        
        elif sub_transaction.type == 'Bank To Cash Transfer':
            cumulative_balance_change -= sub_transaction.amount
        elif sub_transaction.adjtype == 'Reduce Balance':
            cumulative_balance_change -= sub_transaction.amount
        else:
            cumulative_balance_change += sub_transaction.amount

        sub_transaction.balance = cumulative_balance_change
        sub_transaction.save()

    bank_id = transaction.bank.id
    if transaction.type == 'Bank To Bank Transfer':
      
        print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        print(f"To Bank Balance After Transfer: {to_bank_id.balance}")
        from_bank_id.balance += transaction.amount
        to_bank_id.balance -= transaction.amount


        print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        print(f"To Bank Balance After Transfer: {to_bank_id.balance}")

        from_bank_id.save()
        to_bank_id.save()
        print(f"From Bank Balance After Transfer: {from_bank_id.balance}")
        print(f"To Bank Balance After Transfer: {to_bank_id.balance}")
    transaction.delete()
    
    return redirect('bank_listout', id=bank_id)

def banking_status(request,id):
    bnk = Banking.objects.get(id=id)
    if bnk.status == 'Active':
        bnk.status = 'Inactive'
    else:
        bnk.status = 'Active'
    bnk.save()
    return redirect('bank_view',id)

def bank_attachfile(request,id):
    if request.method == 'POST':
        bnk = Banking.objects.get(id=id) 
        bnk.document = request.POST.get('file')
        if len(request.FILES) != 0:
            bnk.document = request.FILES['file']
        bnk.save()
        return redirect('bank_view',id)
    
def send_bank_transaction(request,id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            dash_details = StaffDetails.objects.get(login_details=log_details)
        bnk = Banking.objects.get(id=id, company=cmp)
        trans = BankTransaction.objects.filter(company=cmp, banking=bnk)
        context = { 'bnk':bnk, 'trans':trans}

        emails_string = request.POST['mail']
        cemail = [email.strip() for email in emails_string.split(',')]
        template_path = 'bank_trans_template.html'
        template = get_template(template_path)
        html  = template.render(context)
        
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Transactions.pdf'
        subject = f"Bank Transactions"
        email = EmailMessage(subject, f"Hi,\nPlease find below the attached bank transaction on the bank {bnk.bnk_name} with account number {bnk.bnk_acno}.\n", 
                             from_email=settings.EMAIL_HOST_USER, to=cemail)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)
        
        messages.success(request, 'Bill has been shared via email successfully..!')
        return redirect('bank_view', id)
    
def company_gsttype_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            
            gstno = request.POST.get('gstno')
            gsttype = request.POST.get('gsttype')

            # Check if gsttype is one of the specified values
            if gsttype in ['unregistered Business', 'Overseas', 'Consumer']:
                dash_details.gst_no = None
            else:
                if gstno:
                    dash_details.gst_no = gstno
                else:
                    messages.error(request,'GST Number is not entered*')
                    return redirect('company_profile_editpage')


            dash_details.gst_type = gsttype

            dash_details.save()
            messages.success(request,'GST Type changed')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/') 
        
        
#------------------- PRICE LIST MODULE ------------
#  display all price lists
def all_price_lists(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)



def import_price_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('price_list_file') and request.FILES.get('items_file'):
            price_list_file = request.FILES['price_list_file']
            items_file = request.FILES['items_file']

            try:
                # Read PriceList Excel file(price_list_file)
                price_list_df = pd.read_excel(price_list_file)

                # Create PriceList and PriceListItem instances
                for index, row in price_list_df.iterrows():
                    # Check if a PriceList with the same name already exists
                    existing_price_list = PriceList.objects.filter(name=row['NAME'], company=dash_details).first()
                    if existing_price_list:
                        messages.error(request, f'Error importing data: PriceList with name "{row["NAME"]}" already exists.')
                        continue

                    new_price_list = PriceList.objects.create(
                        name=row['NAME'],
                        type=row['TYPE'],
                        item_rate_type=row['ITEM_RATE_TYPE'], 
                        description=row['DESCRIPTION'],
                        percentage_type=row['PERCENTAGE_TYPE'],
                        percentage_value=row['PERCENTAGE_VALUE'],
                        round_off=row['ROUNDING'],
                        currency=row['CURRENCY'],
                        company=dash_details, 
                        login_details=log_details,
                    )
                    PriceListTransactionHistory.objects.create(
                        company=dash_details,
                        login_details=log_details,
                        price_list=new_price_list,
                        action='Created',
                    )

                    # Read Items Excel file(items_file) for each PriceList
                    items_df = pd.read_excel(items_file)
                    for item_index, item_row in items_df.iterrows():
                        item = Items.objects.filter(item_name=item_row['ITEM_NAME'], company=dash_details, activation_tag='active').first()
                        if item:
                            standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                            custom_rate1 = item_row.get('SELLING_CUSTOM_RATE') if new_price_list.type == 'Sales' else item_row.get('PURCHASE_CUSTOM_RATE')
                            custom_rate = standard_rate if new_price_list.item_rate_type == 'Percentage' else custom_rate1
                            if custom_rate is None or math.isnan(custom_rate) or custom_rate == '':
                                custom_rate = Decimal(standard_rate)
                            
                            PriceListItem.objects.create(
                                company=dash_details,
                                login_details=log_details,
                                price_list=new_price_list,
                                item=item,
                                standard_rate=standard_rate,
                                custom_rate=custom_rate,
                            )
                        else:
                            messages.warning(request, f'Skipping item "{item_row["ITEM_NAME"]}" in PriceList "{row["NAME"]}": Item is not active.')

                messages.success(request, 'Price List data imported successfully.')
                return redirect('all_price_lists')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('price_list_file') and request.FILES.get('items_file'):
            price_list_file = request.FILES['price_list_file']
            items_file = request.FILES['items_file']

            try:
                # Read PriceList Excel file(price_list_file)
                price_list_df = pd.read_excel(price_list_file)

                # Create PriceList and PriceListItem instances
                for index, row in price_list_df.iterrows():
                    # Check if a PriceList with the same name already exists
                    existing_price_list = PriceList.objects.filter(name=row['NAME'], company=dash_details.company).first()
                    if existing_price_list:
                        messages.error(request, f'Error importing data: PriceList with name "{row["NAME"]}" already exists.')
                        continue

                    new_price_list = PriceList.objects.create(
                        name=row['NAME'],
                        type=row['TYPE'],
                        item_rate_type=row['ITEM_RATE_TYPE'], 
                        description=row['DESCRIPTION'],
                        percentage_type=row['PERCENTAGE_TYPE'],
                        percentage_value=row['PERCENTAGE_VALUE'],
                        round_off=row['ROUNDING'],
                        currency=row['CURRENCY'],
                        company=dash_details.company, 
                        login_details=log_details,
                    )
                    PriceListTransactionHistory.objects.create(
                        company=dash_details.company,
                        login_details=log_details,
                        price_list=new_price_list,
                        action='Created',
                    )

                    # Read Items Excel file(items_file) for each PriceList
                    items_df = pd.read_excel(items_file)
                    for item_index, item_row in items_df.iterrows():
                        item = Items.objects.filter(item_name=item_row['ITEM_NAME'], company=dash_details.company, activation_tag='active').first()
                        if item:
                            standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                            custom_rate1 = item_row.get('SELLING_CUSTOM_RATE') if new_price_list.type == 'Sales' else item_row.get('PURCHASE_CUSTOM_RATE')
                            custom_rate = standard_rate if new_price_list.item_rate_type == 'Percentage' else custom_rate1
                            if custom_rate is None or math.isnan(custom_rate) or custom_rate == '':
                                custom_rate = Decimal(standard_rate)
                            
                            PriceListItem.objects.create(
                                company=dash_details.company,
                                login_details=log_details,
                                price_list=new_price_list,
                                item=item,
                                standard_rate=standard_rate,
                                custom_rate=custom_rate,
                            )
                        else:
                            messages.warning(request, f'Skipping item "{item_row["ITEM_NAME"]}" in PriceList "{row["NAME"]}": Item is not active.')

                messages.success(request, 'Price List data imported successfully.')
                return redirect('all_price_lists')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    else:
        return redirect('/')

    return redirect('all_price_lists')

def create_price_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details, status='New')
        items = Items.objects.filter(company=dash_details,activation_tag='active')

        if request.method == 'POST':
            name = request.POST['name']
            if PriceList.objects.filter(name=name, company=dash_details).exists():
                messages.error(request, f"A Price List with the name '{name}' already exists.")
                return redirect('create_price_list')
            
            new_price_list = PriceList.objects.create(
                name=name,
                type=request.POST['type'],
                item_rate_type=request.POST['item_rate_type'],
                description=request.POST['description'],
                percentage_type=request.POST['percentage_type'],
                percentage_value=request.POST['percentage_value'],
                round_off=request.POST['round_off'],
                currency=request.POST['currency'],
                company=dash_details,
                login_details=log_details,
            )

            PriceListTransactionHistory.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=new_price_list,
                action='Created',
            )
            custom_rates = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rates):
                custom_rate = custom_rate if custom_rate else (item.selling_price if new_price_list.type == 'Sales' else item.purchase_price)
                standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                PriceListItem.objects.create(
                    company=dash_details,
                    login_details=log_details,
                    price_list=new_price_list,
                    item=item,
                    standard_rate=standard_rate,
                    custom_rate=custom_rate,
                )
            return redirect('all_price_lists')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/create_price_list.html', context)

    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        items = Items.objects.filter(company=dash_details.company,activation_tag='active')

        if request.method == 'POST':
            name = request.POST['name']
            if PriceList.objects.filter(name=name, company=dash_details.company).exists():
                messages.error(request, f"A Price List with the name '{name}' already exists.")
                return redirect('create_price_list')
            
            new_price_list = PriceList.objects.create(
                name=name,
                type=request.POST['type'],
                item_rate_type=request.POST['item_rate_type'],
                description=request.POST['description'],
                percentage_type=request.POST['percentage_type'],
                percentage_value=request.POST['percentage_value'],
                round_off=request.POST['round_off'],
                currency=request.POST['currency'],
                company=dash_details.company,
                login_details=log_details
            )
            
            PriceListTransactionHistory.objects.create(
                company=dash_details.company,
                login_details=log_details,
                price_list=new_price_list,
                action='Created',
            )
            
            custom_rates = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rates):
                custom_rate = custom_rate if custom_rate else (item.selling_price if new_price_list.type == 'Sales' else item.purchase_price)
                standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                PriceListItem.objects.create(
                    company=dash_details.company,
                    login_details=log_details,
                    price_list=new_price_list,
                    item=item,
                    standard_rate=standard_rate,
                    custom_rate=custom_rate,
                )

            return redirect('all_price_lists')

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/create_price_list.html', context)
 

def edit_price_list(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        allmodules = ZohoModules.objects.get(company=dash_details, status='New')
        price_list = get_object_or_404(PriceList, id=price_list_id)
        items = PriceListItem.objects.filter(price_list=price_list)

        
        if request.method == 'POST':
            price_list.name = request.POST['name']
            price_list.type = request.POST['type']
            price_list.item_rate_type = request.POST['item_rate_type']
            price_list.description = request.POST['description']
            price_list.percentage_type = request.POST['percentage_type']
            price_list.percentage_value = request.POST['percentage_value']
            price_list.round_off = request.POST['round_off']
            price_list.currency = request.POST['currency']
            price_list.save()
            
            PriceListTransactionHistory.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=price_list,
                action='Edited',
            )
            
            # edit PriceListItem
            custom_rate = request.POST.getlist('custom_rate')
            items = PriceListItem.objects.filter(price_list=price_list)
            for item, custom_rate in zip(items, custom_rate):
                standard_rate = item.item.selling_price if price_list.type == 'Sales' else item.item.purchase_price
                item.standard_rate = standard_rate
                item.custom_rate = custom_rate
                item.save()
            
            
            return redirect('price_list_details', price_list_id=price_list_id)
        context = {
            'log_details': log_details,
            'details': dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/edit_price_list.html', context)
    elif log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        price_list = get_object_or_404(PriceList, id=price_list_id)
        items = PriceListItem.objects.filter(price_list=price_list)
        
        if request.method == 'POST':
            price_list.name = request.POST['name']
            price_list.type = request.POST['type']
            price_list.item_rate_type = request.POST['item_rate_type']
            price_list.description = request.POST['description']
            price_list.percentage_type = request.POST['percentage_type']
            price_list.percentage_value = request.POST['percentage_value']
            price_list.round_off = request.POST['round_off']
            price_list.currency = request.POST['currency']
            price_list.save()
            PriceListTransactionHistory.objects.create(
                    company=dash_details.company,
                    login_details=log_details,
                    price_list=price_list,
                    action='Edited',
                )
            
            # edit PriceListItem
            custom_rate = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rate):
                    standard_rate = item.item.selling_price if price_list.type == 'Sales' else item.item.purchase_price
                    item.standard_rate = standard_rate
                    item.custom_rate = custom_rate
                    item.save()
            
            return redirect('price_list_details', price_list_id=price_list_id)

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'items':items,
        }
        return render(request, 'zohomodules/price_list/edit_price_list.html', context)


# display details of selected price list
def price_list_details(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        comments = PriceListComment.objects.filter(price_list=price_list)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        transaction_history = PriceListTransactionHistory.objects.filter(price_list=price_list)
        items = PriceListItem.objects.filter(company=dash_details, price_list=price_list)
        latest_transaction = PriceListTransactionHistory.objects.filter(price_list=price_list)

        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'comments': comments,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'latest_transaction': latest_transaction,
            'transaction_history': transaction_history,
            'items':items,
        }
        return render(request,'zohomodules/price_list/price_list_details.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        comments = PriceListComment.objects.filter(price_list=price_list)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        transaction_history = PriceListTransactionHistory.objects.filter(price_list=price_list)
        items = PriceListItem.objects.filter(company=dash_details.company, price_list=price_list)
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'comments': comments,
            'price_list': price_list,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'transaction_history': transaction_history,
            'items':items,
        }
        return render(request,'zohomodules/price_list/price_list_details.html',context)
    

def delete_price_list(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        price_list.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        price_list.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)


def toggle_price_list_status(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details)
        if price_list.status == 'Active':
            price_list.status = 'Inactive'
        else:
            price_list.status = 'Active'
        price_list.save()
        PriceListTransactionHistory.objects.create(
            company=dash_details,
            login_details=log_details,
            price_list=price_list,
            action='Edited' 
        )
        return redirect('price_list_details', price_list_id=price_list_id)
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details.company)
        if price_list.status == 'Active':
            price_list.status = 'Inactive'
        else:
            price_list.status = 'Active'
        price_list.save()
        PriceListTransactionHistory.objects.create(
            company=dash_details.company,
            login_details=log_details,
            price_list=price_list,
            action='Edited'  
        )
        return redirect('price_list_details', price_list_id=price_list_id)

def add_pricelist_comment(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            PriceListComment.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=price_list,
                comment=comment
            )
            
        return redirect('price_list_details', price_list_id=price_list_id)
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details.company)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            PriceListComment.objects.create(
                company=dash_details.company,
                login_details=log_details,
                price_list=price_list,
                comment=comment
            )
        return redirect('price_list_details', price_list_id=price_list_id)

def delete_pricelist_comment(request, comment_id, price_list_id):
    comment = get_object_or_404(PriceListComment, id=comment_id)
    comment.delete()
    return redirect('price_list_details', price_list_id=price_list_id)





def whatsapp_pricelist(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_items = PriceListItem.objects.filter(price_list=price_list)

        context = {
            'price_list': price_list,
            'price_list_items': price_list_items,
        }

        # Render the template
        html = render(request, 'zohomodules/price_list/pdf_price_list.html', context).content

        # Create a PDF file
        pdf_file = BytesIO()
        pisa.pisaDocument(BytesIO(html), pdf_file)

        # Check if PDF generation was successful
        if pdf_file.tell():
            pdf_file.seek(0)

            # Save the PDF to the server's media folder
            pdf_filename = f"{price_list.name}_price_list.pdf"
            pdf_path = os.path.join('media', pdf_filename)
            with open(pdf_path, 'wb') as pdf_file_on_server:
                pdf_file_on_server.write(pdf_file.read())

            # Create a direct link to the saved PDF
            pdf_link = f"{request.scheme}://{request.get_host()}/{pdf_path}"

            # Create the WhatsApp message with a link to download the PDF
            whatsapp_message = f"Check out this price list: [Download PDF]{pdf_link}"

            # Create the WhatsApp link
            whatsapp_link = f"https://wa.me/?text={whatsapp_message}"

            # Return the WhatsApp link
            return redirect(whatsapp_link)

    except Exception as e:
        print(e)
        messages.error(request, f'{e}')

    # If there is an error or PDF generation fails, redirect to 'all_price_lists'
    return redirect('all_price_lists')



# email pricelist details(overview)
def email_pricelist(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_item = PriceListItem.objects.filter( price_list=price_list)

        if request.method == 'POST':
            emails_string = request.POST['email_ids']
            emails_list = [email.strip() for email in emails_string.split(',')]
            email_message = request.POST['email_message']

            context = {
                'price_list': price_list,
                'price_list_item': price_list_item,
            }

            template_path = 'zohomodules/price_list/pdf_price_list.html'
            template = get_template(template_path)
            html = template.render(context)
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
            pdf = result.getvalue()

            filename = f'Price_List_Details.pdf'
            subject = f"Price List Details: {price_list.name}"
            email = EmailMessage(subject, f"Hi,\nPlease find the attached Price List Details. \n{email_message}\n\n--\nRegards,\n{price_list.name}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
            email.attach(filename, pdf, "application/pdf")
            email.send(fail_silently=False)

            msg = messages.success(request, 'Details have been shared via email successfully..!')
            return redirect('price_list_details', price_list_id=price_list_id)  

    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect('all_price_lists')  

# dwnld pdf
def price_list_pdf(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_item = PriceListItem.objects.filter(price_list=price_list)

        context = {
            'price_list': price_list,
            'price_list_item': price_list_item,
        }

        template_path = 'zohomodules/price_list/pdf_price_list.html'
        template = get_template(template_path)
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{price_list.name}_Details.pdf"'
        response.write(pdf)
        return response
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect('all_price_lists')

# upload attachment
def attach_file(request, price_list_id):
    price_list = PriceList.objects.get(pk=price_list_id)
    if request.method == 'POST':
        attachment = request.FILES.get('attachment')
        price_list.attachment = attachment
        price_list.save()
        return redirect('price_list_details', price_list_id=price_list.id)
    return HttpResponse("Invalid request method.")


#----------------------------------------------------------akshay--end--------------------------------------------------------


#-------------------------------Arya E.R-----------------------------------------------

 ##### Vendor #####
    
def vendor(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        # net_30_exists = Company_Payment_Term.objects.filter(company=dash_details, term_name='NET 30').exists()

        # net_60_exists = Company_Payment_Term.objects.filter(company=dash_details, term_name='NET 60').exists()

        # if not net_30_exists:
        #     normalized_data='NET30'
        #     pay_tm = add_space_before_first_digit(normalized_data)

        #     Company_Payment_Term.objects.create(company=dash_details, term_name='NET 30',days=30)

        # if not net_60_exists:
        #     normalized_data='NET60'
        #     pay_tm = add_space_before_first_digit(normalized_data)
        #     Company_Payment_Term.objects.create(company=dash_details, term_name='NET 60',days=60)
        comp_payment_terms=Company_Payment_Term.objects.filter(company=dash_details)
        if log_details.user_type=='Staff':

            return render(request,'zohomodules/vendor/create_vendor.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details}) 
        else:
            return render(request,'zohomodules/vendor/create_vendor.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details}) 
    else:
        return redirect('/')

def view_vendor_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

        allmodules= ZohoModules.objects.get(company=dash_details,status='New')  

        data=Vendor.objects.filter(company=dash_details)

         # Pagination
        
        # page = request.GET.get('page', 1)
        # paginator = Paginator(data, 5)

        # try:
        #     items = paginator.page(page)
        # except PageNotAnInteger:
        #     items = paginator.page(1)
        # except EmptyPage:
        #     items = paginator.page(paginator.num_pages)

        return render(request,'zohomodules/vendor/vendor_list.html',{'details':dash_details,'allmodules': allmodules,'data':data,'log_details':log_details}) 


    else:
        return redirect('/')


# @login_required(login_url='login')
def add_vendor(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

        

       
        if request.method=="POST":
            vendor_data=Vendor()
            vendor_data.login_details=log_details
            vendor_data.company=dash_details
            vendor_data.title = request.POST.get('salutation')
            vendor_data.first_name=request.POST['first_name']
            vendor_data.last_name=request.POST['last_name']
            vendor_data.company_name=request.POST['company_name']
            vendor_data.vendor_display_name=request.POST['v_display_name']
            vendor_data.vendor_email=request.POST['vendor_email']
            vendor_data.phone=request.POST['w_phone']
            vendor_data.mobile=request.POST['m_phone']
            vendor_data.skype_name_number=request.POST['skype_number']
            vendor_data.designation=request.POST['designation']
            vendor_data.department=request.POST['department']
            vendor_data.website=request.POST['website']
            vendor_data.gst_treatment=request.POST['gst']
            vendor_data.vendor_status="Active"
            vendor_data.remarks=request.POST['remark']
            vendor_data.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                vendor_data.pan_number=request.POST['pan_number']
                vendor_data.gst_number="null"
            else:
                vendor_data.gst_number=request.POST['gst_number']
                vendor_data.pan_number=request.POST['pan_number']

            vendor_data.source_of_supply=request.POST['source_supply']
            vendor_data.currency=request.POST['currency']
            print(vendor_data.currency)
            op_type=request.POST.get('op_type')
            if op_type is not None:
                vendor_data.opening_balance_type=op_type
            else:
                vendor_data.opening_balance_type='Opening Balance not selected'
    
            vendor_data.opening_balance=request.POST['opening_bal']
            vendor_data.payment_term=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])

           
            vendor_data.billing_attention=request.POST['battention']
            vendor_data.billing_country=request.POST['bcountry']
            vendor_data.billing_address=request.POST['baddress']
            vendor_data.billing_city=request.POST['bcity']
            vendor_data.billing_state=request.POST['bstate']
            vendor_data.billing_pin_code=request.POST['bzip']
            vendor_data.billing_phone=request.POST['bphone']
            vendor_data.billing_fax=request.POST['bfax']
            vendor_data.shipping_attention=request.POST['sattention']
            vendor_data.shipping_country=request.POST['s_country']
            vendor_data.shipping_address=request.POST['saddress']
            vendor_data.shipping_city=request.POST['scity']
            vendor_data.shipping_state=request.POST['sstate']
            vendor_data.shipping_pin_code=request.POST['szip']
            vendor_data.shipping_phone=request.POST['sphone']
            vendor_data.shipping_fax=request.POST['sfax']
            vendor_data.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=VendorHistory()
            vendor_history_obj.company=dash_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.vendor=vendor_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
            rdata=Vendor_remarks_table()
            rdata.remarks=request.POST['remark']
            rdata.company=dash_details
            rdata.vendor=vdata
            rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = VendorContactPerson.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=dash_details,vendor=vendor)
                
        
            messages.success(request, 'Data saved successfully!')   

            return redirect('view_vendor_list')
        
        else:
            messages.error(request, 'Some error occurred !')   

            return redirect('view_vendor_list')
    

    
def sort_vendor_by_name(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
  
        data=Vendor.objects.filter(login_details=log_details).order_by('first_name')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
            return redirect('/')   

def sort_vendor_by_amount(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details).order_by('opening_balance')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 

def view_vendor_active(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details,vendor_status='Active').order_by('-id')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 

    
    
def view_vendor_inactive(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details,vendor_status='Inactive').order_by('-id')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 
    
def delete_vendor(request,pk):
    if Vendor_comments_table.objects.filter(vendor=pk).exists():
        user2=Vendor_comments_table.objects.filter(vendor=pk)
        user2.delete()
    if Vendor_mail_table.objects.filter(vendor=pk).exists():
        user3=Vendor_mail_table.objects.filter(vendor=pk)
        user3.delete()
    if Vendor_doc_upload_table.objects.filter(vendor=pk).exists():
        user4=Vendor_doc_upload_table.objects.filter(vendor=pk)
        user4.delete()
    if VendorContactPerson.objects.filter(vendor=pk).exists():
        user5=VendorContactPerson.objects.filter(vendor=pk)
        user5.delete()
    if Vendor_remarks_table.objects.filter(vendor=pk).exists():
        user6=Vendor_remarks_table.objects.filter(vendor=pk)
        user6.delete()
    
    user1=Vendor.objects.get(id=pk)
    user1.delete()
    return redirect("view_vendor_list")



def view_vendor_details(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        vendor_obj=Vendor.objects.get(id=pk)

        # Getting all vendor to disply on the left side of vendor_detailsnew page
        vendor_objs=Vendor.objects.filter(company=dash_details)

        vendor_comments=Vendor_comments_table.objects.filter(vendor=vendor_obj)
        vendor_history=VendorHistory.objects.filter(vendor=vendor_obj)
    
    content = {
                'details': dash_details,
               
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/vendor/vendor_detailsnew.html',content)

def import_vendor_excel(request):
   if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)
            

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method == 'POST' :
       
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                            
                    vendorsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    comp_term=vendorsheet[16]
                    pay_tm = add_space_before_first_digit(comp_term)
                    try:
                        com_term_obj=Company_Payment_Term.objects.get(company=dash_details,term_name=pay_tm)
                    except Company_Payment_Term.DoesNotExist:
                        com_term_obj= None
                    opn_blc_str = vendorsheet[15]  # Assuming vendorsheet[15] is a string representing a decimal
                    opn_blc = Decimal(opn_blc_str)
                    Vendor_object=Vendor(title=vendorsheet[0],first_name=vendorsheet[1],last_name=vendorsheet[2],company_name=vendorsheet[3],vendor_email=vendorsheet[4],phone=vendorsheet[5],mobile=vendorsheet[6],skype_name_number=vendorsheet[7],designation=vendorsheet[8],department=vendorsheet[9],website=vendorsheet[10],
                                         gst_treatment=vendorsheet[11],source_of_supply=vendorsheet[12],currency=vendorsheet[13],opening_balance_type=vendorsheet[14],
                                         opening_balance=opn_blc,payment_term=com_term_obj,billing_attention=vendorsheet[17],billing_address=vendorsheet[18],
                                         billing_city=vendorsheet[19],billing_state=vendorsheet[20],billing_country=vendorsheet[21],billing_pin_code=vendorsheet[22],
                                         billing_phone=vendorsheet[23],billing_fax=vendorsheet[24],shipping_attention=vendorsheet[25],shipping_address=vendorsheet[26],shipping_city=vendorsheet[27],
                                         shipping_state=vendorsheet[28],shipping_country=vendorsheet[29],shipping_pin_code=vendorsheet[30],
                                         shipping_phone=vendorsheet[31], shipping_fax=vendorsheet[32], remarks=vendorsheet[33],vendor_status="Active",company=dash_details,login_details=log_details)
                    Vendor_object.save()

    
                   
                messages.warning(request,'file imported')
                return redirect('view_vendor_list')    

    
            messages.error(request,'File upload Failed!11')
            return redirect('view_vendor_list')
        else:
            messages.error(request,'File upload Failed!11')
            return redirect('view_vendor_list')
        
def Vendor_edit(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)

   

    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

        allmodules= ZohoModules.objects.get(company=dash_details,status='New') 

        vendor_obj=Vendor.objects.get(id=pk)

    vendor_contact_obj=VendorContactPerson.objects.filter(vendor=vendor_obj)  
    comp_payment_terms=Company_Payment_Term.objects.filter(company=dash_details)
   
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'vendor_obj':vendor_obj,
            'log_details':log_details,
            'vendor_contact_obj':vendor_contact_obj,
            'comp_payment_terms':comp_payment_terms,
    }
   

    return render(request,'zohomodules/vendor/edit_vendor.html',content)

def do_vendor_edit(request,pk):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method=="POST":
            vendor_data=Vendor.objects.get(id=pk)
            vendor_data.login_details=log_details
            vendor_data.company=dash_details
            vendor_data.title = request.POST.get('salutation')
            vendor_data.first_name=request.POST['first_name']
            vendor_data.last_name=request.POST['last_name']
            vendor_data.company_name=request.POST['company_name']
            vendor_data.vendor_display_name=request.POST['v_display_name']
            vendor_data.vendor_email=request.POST['vendor_email']
            vendor_data.phone=request.POST['w_phone']
            vendor_data.mobile=request.POST['m_phone']
            vendor_data.skype_name_number=request.POST['skype_number']
            vendor_data.designation=request.POST['designation']
            vendor_data.department=request.POST['department']
            vendor_data.website=request.POST['website']
            vendor_data.gst_treatment=request.POST['gst']
            vendor_data.vendor_status="Active"
            vendor_data.remarks=request.POST['remark']
            
            cob=Decimal(request.POST['opening_bal'])
            oc=Decimal(vendor_data.current_balance) 
            ob=Decimal(vendor_data.opening_balance) 

            if cob > ob:
                diffadd=cob-ob
                oc=oc + diffadd
                vendor_data.current_balance=oc
                vendor_data.opening_balance=cob
            elif cob < ob:
                diffadd=ob-cob
                oc=oc-diffadd
                vendor_data.current_balance=oc
                vendor_data.opening_balance=cob

            else:
                vendor_data.current_balance=request.POST['opening_bal']   
       
            

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                vendor_data.pan_number=request.POST['pan_number']
                vendor_data.gst_number="null"
            else:
                vendor_data.gst_number=request.POST['gst_number']
                vendor_data.pan_number=request.POST['pan_number']

            vendor_data.source_of_supply=request.POST['source_supply']
            vendor_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                vendor_data.opening_balance_type=op_type
            else:
                vendor_data.opening_balance_type='Opening Balance not selected'
            vendor_data.opening_balance=request.POST['opening_bal']
            vendor_data.payment_term=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            
           
            vendor_data.billing_attention=request.POST['battention']
            vendor_data.billing_country=request.POST['bcountry']
            vendor_data.billing_address=request.POST['baddress']
            vendor_data.billing_city=request.POST['bcity']
            vendor_data.billing_state=request.POST['bstate']
            vendor_data.billing_pin_code=request.POST['bzip']
            vendor_data.billing_phone=request.POST['bphone']
            vendor_data.billing_fax=request.POST['bfax']
            vendor_data.shipping_attention=request.POST['sattention']
            vendor_data.shipping_country=request.POST['s_country']
            vendor_data.shipping_address=request.POST['saddress']
            vendor_data.shipping_city=request.POST['scity']
            vendor_data.shipping_state=request.POST['sstate']
            vendor_data.shipping_pin_code=request.POST['szip']
            vendor_data.shipping_phone=request.POST['sphone']
            vendor_data.shipping_fax=request.POST['sfax']
            vendor_data.save()


              # ................ Adding to History table...........................
            
            vendor_history_obj=VendorHistory()
            vendor_history_obj.company=dash_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.vendor=vendor_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Edited'
            vendor_history_obj.save()
    # .......................................................adding to remaks table.....................
            vdata=Vendor.objects.get(id=vendor_data.id)
            try:

                rdata=Vendor_remarks_table.objects.get(vendor=vdata)
                rdata.remarks=request.POST['remark']
                rdata.company=dash_details
                rdata.vendor=vdata
                rdata.save()
            except Vendor_remarks_table.DoesNotExist:
                remarks_obj= Vendor_remarks_table()   
                remarks_obj.remarks=request.POST['remark']
                remarks_obj.company=dash_details
                remarks_obj.vendor=vdata
                remarks_obj.save()


    #  ...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            person = request.POST.getlist('contact_person_id[]')
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
            print(person)
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department)==len(person):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department,person)
                    mapped2=list(mapped2)
                    for ele in mapped2:
                       
                        existing_instance = VendorContactPerson.objects.filter(id=ele[9], company=dash_details, vendor=vendor).first()
                        if existing_instance:
                            # Update the existing instance
                            existing_instance.title = ele[0]
                            existing_instance.first_name = ele[1]
                            existing_instance.last_name = ele[2]
                            existing_instance.email = ele[3]
                            existing_instance.work_phone  = ele[4]
                            existing_instance.mobile = ele[5]
                            existing_instance.skype_name_number = ele[6]
                            existing_instance.designation = ele[7]
                            existing_instance.department = ele[8]

                            # Update other fields

                            existing_instance.save()
                        else:
                            # Create a new instance
                            new_instance = VendorContactPerson.objects.create(
                                title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=dash_details,vendor=vendor
                            )
            return redirect('view_vendor_details',pk)
    

def delete_vendors(request, pk):
    try:
        vendor_obj = Vendor.objects.get(id=pk)

        vendor_obj.delete()
        return redirect('view_vendor_list')  
    except Vendor.DoesNotExist:
        return HttpResponseNotFound("Vendor not found.")
    
def vendor_status(request,pk):
    vendor_obj = Vendor.objects.get(id=pk)
    if vendor_obj.vendor_status == 'Active':
        vendor_obj.vendor_status ='Inactive'
    elif vendor_obj.vendor_status == 'Inactive':
        vendor_obj.vendor_status ='Active'
    vendor_obj.save()
    return redirect('view_vendor_details',pk)   

def vendor_add_comment(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
  
        if request.method =='POST':
            comment_data=request.POST['comments']
       
            vendor_id= Vendor.objects.get(id=pk) 
            vendor_obj=Vendor_comments_table()
            vendor_obj.comment=comment_data
            vendor_obj.vendor=vendor_id
            vendor_obj.company=dash_details
            vendor_obj.login_details= LoginDetails.objects.get(id=log_id)

            vendor_obj.save()
            return redirect('view_vendor_details',pk)
    return redirect('view_vendor_details',pk) 


def vendor_delete_comment(request, pk):
    try:
        vendor_comment =Vendor_comments_table.objects.get(id=pk)
        vendor_id=vendor_comment.vendor.id
        vendor_comment.delete()
        return redirect('view_vendor_details',vendor_id)  
    except Vendor_comments_table.DoesNotExist:
        return HttpResponseNotFound("comments not found.")
    

def add_vendor_file(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method == 'POST':
            data=request.FILES.getlist('file')
            try:
                for doc in data:

                    vendor_obj=Vendor_doc_upload_table()
                    
                    vendor_obj.document = doc
                    vendor_obj.login_details = log_details
                    vendor_obj.company = dash_details
                    vendor_obj.vendor = Vendor.objects.get(id=pk)
                    vendor_obj.save()
                
                messages.success(request,'File uploaded')
                return redirect('view_vendor_details',pk) 
            except Vendor_doc_upload_table.DoesNotExist:
                return redirect('view_vendor_details',pk) 


from django.core.mail import EmailMultiAlternatives    
def vendor_shareemail(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
    
        vendor_obj=Vendor.objects.get(id=pk)

        context = {'vendor_obj':vendor_obj}

        emails_string = request.POST.get('email', '').strip()
        cemail = [email.strip() for email in emails_string.split(',')]
        
        
        subject = "Transaction Details"
       
        message = 'Hi,\nPlease find the attached transaction details - File-{vendor_obj.first_name} {vendor_obj.last_name} .\n--\nRegards,\n",'
       
        email_from = settings.EMAIL_HOST_USER
            
        recipient_list = [emails_string, ]
        
       
        msg = EmailMultiAlternatives(subject, message, email_from, [emails_string])
        msg.attach_alternative('Transactions.pdf', "application/pdf")
        msg.send()
        messages.success(request, 'Transaction has been shared via email successfully..!')
        return redirect('view_vendor_details',pk)
    
    

def payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)        
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            normalized_data = terms.replace(" ", "")
            pay_tm = add_space_before_first_digit(normalized_data)
            ptr = Company_Payment_Term(term_name=pay_tm, days=day, company=dash_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=dash_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)    
            

#---------------------------------------End----------------------------------------------------------------  


def check_term_exist(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

    if request.method == 'GET':
       term_name = request.GET.get('term_name', None)
       if term_name:
            normalized_data = term_name.replace(" ", "")
            term_name_processed = add_space_before_first_digit(normalized_data)
            exists = Company_Payment_Term.objects.filter(
                    term_name=term_name_processed,
                    company=dash_details
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})
        
def check_email_exist(request):
    if request.method == 'GET':
       vendoremail = request.GET.get('vendor_email', None)

       if vendoremail:
          
            exists = Vendor.objects.filter(
                    vendor_email=vendoremail
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})
        
def check_work_phone_exist(request):
    if request.method == 'GET':
       wPhone = request.GET.get('w_Phone', None)

       if wPhone:
          
            exists = Vendor.objects.filter(
                    phone=wPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})
        
        
def check_phonenumber_exist(request):
    if request.method == 'GET':
       mPhone = request.GET.get('m_Phone', None)

       if mPhone:
          
            exists = Vendor.objects.filter(
                    mobile=mPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})
        
        
def add_space_before_first_digit(data):
    for index, char in enumerate(data):
        if char.isdigit():
            return data[:index] + ' ' + data[index:]
    return data
    
    
# -------------------------------Zoho Modules section--------------------------------
 
# Check Pan Number Exist or Not
def check_pan(request):
    if request.method == 'POST':
        panNumber = request.POST.get('panNumber')
        pan_exists = Vendor.objects.filter(pan_number=panNumber).exists()

        if pan_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'}) 
        
# Check GST Number Exist or Not
def check_gst(request):
    if request.method == 'POST':
        gstNumber = request.POST.get('gstNumber')
        gst_exists = Vendor.objects.filter(gst_number=gstNumber).exists()
       
        if gst_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'})
        
        
# -------------------------------Zoho Modules section--------------------------------
 
def sort_vendor(request,selectId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        vendor_obj = Vendor.objects.get(id=pk)
        vendor_objs = Vendor.objects.filter(company=dash_details)

        if selectId == 0:
            vendor_objs=Vendor.objects.filter(company=dash_details)
        if selectId == 1:
            vendor_objs=Vendor.objects.filter(company=dash_details).order_by('first_name')
        if selectId == 2:
            vendor_objs=Vendor.objects.filter(company=dash_details).order_by('opening_balance')
           
        
        vendor_comments=Vendor_comments_table.objects.filter(vendor=vendor_obj)
        vendor_history=VendorHistory.objects.filter(vendor=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/vendor/vendor_detailsnew.html',content)


def vendor_status_change(request,statusId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        vendor_obj = Vendor.objects.get(id=pk)
        vendor_objs = Vendor.objects.filter(company=dash_details)

        if statusId == 0:
            vendor_objs=Vendor.objects.filter(company=dash_details)
        if statusId == 1:
            vendor_objs=Vendor.objects.filter(company=dash_details,vendor_status='Active').order_by('-id')
        if statusId == 2:
            vendor_objs=Vendor.objects.filter(company=dash_details,vendor_status='Inactive').order_by('-id')
           
        
        vendor_comments=Vendor_comments_table.objects.filter(vendor=vendor_obj)
        vendor_history=VendorHistory.objects.filter(vendor=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/vendor/vendor_detailsnew.html',content)
    
#End


#---------------- Zoho Final Attendance - Meenu Shaju - Start--------------------
    
def get_days_in_month(target_year, target_month):
    _, days_in_month = monthrange(target_year, target_month)
    days = [day for day in range(1, days_in_month + 1)]
    return days
    
def calculate_leave_count(employee, target_month, target_year):
    return Attendance.objects.filter(employee=employee, date__month =target_month, date__year=target_year).count()
    
def calculate_holiday_count(company, target_month, target_year):
    _, last_day = monthrange(target_year, target_month)
    first_day_of_month = datetime(target_year, target_month, 1)
    last_day_of_month = datetime(target_year, target_month, last_day) + timedelta(days=1)  # Add one day to include the entire end day

    holidays = Holiday.objects.filter(
        start_date__lt=last_day_of_month,
        end_date__gte=first_day_of_month,
        company=company,
    )

    count = 0
    for day in range(1, last_day + 1):
        target_date = datetime(target_year, target_month, day)
        if holidays.filter(start_date__lte=target_date, end_date__gte=target_date).exists():
            count += 1

    return count

   
def company_attendance_list(request):
        
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=staff.company,employee__in=payroll_employee.objects.filter(status='Active'))
            allmodules= ZohoModules.objects.get(company=staff.company,status='New')
                
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=company,employee__in=payroll_employee.objects.filter(status='Active'))
            allmodules= ZohoModules.objects.get(company=company,status='New')
             

        consolidated_entries = defaultdict(list)
        MONTH_NAMES = {
                    1: 'January',
                    2: 'February',
                    3: 'March',
                    4: 'April',
                    5: 'May',
                    6: 'June',
                    7: 'July',
                    8: 'August',
                    9: 'September',
                    10: 'October',
                    11: 'November',
                    12: 'December'
                }


        for item in items:
            target_month = item.date.month
            target_year = item.date.year
            employee_id = item.employee.id

            leave_count = calculate_leave_count(item.employee, target_month, target_year)

            existing_entry = next(
                (
                    entry
                    for entry in consolidated_entries[employee_id]
                    if entry['target_month'] == target_month and entry['target_year'] == target_year
                ),
                None,
            )

            if existing_entry:
                existing_entry['leave'] += leave_count
            else:
               
                entry = {
                    'employee': item.employee,
                    'target_month': target_month,
                    'target_month_name': MONTH_NAMES.get(target_month, ''),
                    'target_year': target_year,
                    'working_days': len(get_days_in_month(target_year, target_month)),
                    'holidays': calculate_holiday_count(item.company, target_month, target_year),
                    'leave': leave_count,
                    'work_days': len(get_days_in_month(target_year, target_month)) - calculate_holiday_count(item.company, target_month, target_year) - leave_count,
                    'total_leave': leave_count,
                }

                consolidated_entries[employee_id].append(entry)

        all_entries = []
        for employee_id, entries in consolidated_entries.items():
            for entry in entries:
                all_entries.append(entry)
        employee_ids = [entry['employee'].id for entries in consolidated_entries.values() for entry in entries]
        employee_ids = [int(id) for id in employee_ids]  # Convert IDs to integers
        request.session['employee_ids'] = employee_ids
        print(employee_ids)

        return render(request, 'zohomodules/Attendance/company_attendance_list.html', {
            'all_entries': all_entries,
            'month_name': MONTH_NAMES,
            'allmodules': allmodules
        })
            
def company_mark_attendance(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            employee = payroll_employee.objects.filter(login_details=log_details,status='Active')
            allmodules= ZohoModules.objects.get(company=company)
            bloods=Bloodgroup.objects.all
            return render(request,'zohomodules/Attendance/company_mark_attendance.html',{'staffs':employee,'blood':bloods,'allmodules':allmodules})
        if log_details.user_type=='Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            
            employee = payroll_employee.objects.filter(company=staff.company,status='Active')
            allmodules= ZohoModules.objects.get(company=staff.company)
            bloods = Bloodgroup.objects.all()
            return render(request,'zohomodules/Attendance/company_mark_attendance.html',{'staffs':employee,'blood':bloods,'allmodules':allmodules})
        
def add_attendance(request):
        if request.method == 'POST':
            emp_id = request.POST['employee']
        date = request.POST['date']
        status = request.POST['status']
        reason = request.POST['reason']

        if 'login_id' in request.session:
            log_id = request.session['login_id']
            log_details = LoginDetails.objects.get(id=log_id)

            if log_details.user_type == 'Company':
                employee = get_object_or_404(payroll_employee, id=emp_id, login_details=log_details)
                company = CompanyDetails.objects.get(login_details=log_details)
            elif log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                employee = get_object_or_404(payroll_employee, id=emp_id, company=staff.company)
                company = staff.company

            is_holiday = Holiday.objects.filter(company=company, start_date__lte=date, end_date__gte=date).exists()

            if is_holiday:
                messages.warning(request, 'Selected date is a company holiday.')
                return redirect('company_mark_attendance')
            

            attendance, created = Attendance.objects.get_or_create(
                employee=employee,
                date=date,
                defaults={'status': status, 'reason': reason, 'company': company, 'login_details': log_details}
            )

            if not created:
                # Update the existing attendance if it already exists for the specified date
                attendance.status = status
                attendance.reason = reason
           
                
            history=Attendance_History(company=company,login_details=log_details,attendance=attendance,date=date,action='Created')
            history.save()
            attendance.save()
            
            return redirect('company_attendance_list')

def attendance_calendar(request, employee_id, target_year, target_month):
    calendar_data = {
        'employee_id': employee_id,
        'target_year': target_year,
        'target_month': target_month,
       
    }
    comment = Attendance_comment.objects.filter(month=target_month,year=target_year,employee=employee_id)
    history = Attendance_History.objects.filter(date__month=target_month,date__year=target_year,attendance__employee=employee_id)
    
# Sort the combined list based on the date of the history or attendance entry
    

    employee_attendance = Attendance.objects.filter(
        employee_id=employee_id,
        date__year=target_year,
        date__month=target_month
    ).values('status', 'date')  # Fetch only the required fields
    
    employee=payroll_employee.objects.get(id=employee_id)
    target_month = max(1, min(target_month, 12))

# Calculate the next month and year if target_month is December
    next_month = 1 if target_month == 12 else target_month + 1
    next_year = target_year + 1 if target_month == 12 else target_year

# Construct the date strings for the start and end of the month
    start_date = datetime(target_year, target_month, 1).date()
    end_date = datetime(next_year, next_month, 1).date() - timedelta(days=1)
    
    holidays = Holiday.objects.filter(
    Q(company=employee.company) & (
    (Q(start_date__lte=end_date) & Q(end_date__gte=start_date)))  # Holidays overlapping the target month
    
)   
    for holiday in holidays:
        holiday.end_date += timedelta(days=1)
    
    
    # for getting atendance list
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=staff.company)
            
            allmodules= ZohoModules.objects.get(company=staff.company,status='New')
                
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=company)
            allmodules= ZohoModules.objects.get(company=company,status='New')

        
             

        consolidated_entries = defaultdict(list)

        for item in items:
            target_month = item.date.month
            target_year = item.date.year
            employee_id = item.employee.id

            leave_count = calculate_leave_count(item.employee, target_month, target_year)

            existing_entry = next(
                (
                    entry
                    for entry in consolidated_entries[employee_id]
                    if entry['target_month'] == target_month and entry['target_year'] == target_year
                ),
                None,
            )

            if existing_entry:
                existing_entry['leave'] += leave_count
            else:
                MONTH_NAMES = {
                    1: 'January',
                    2: 'February',
                    3: 'March',
                    4: 'April',
                    5: 'May',
                    6: 'June',
                    7: 'July',
                    8: 'August',
                    9: 'September',
                    10: 'October',
                    11: 'November',
                    12: 'December'
                }

                entry = {
                    'employee': item.employee,
                    'target_month': target_month,
                    'target_month_name': MONTH_NAMES.get(target_month, ''),
                    'target_year': target_year,
                    'working_days': len(get_days_in_month(target_year, target_month)),
                    'holidays': calculate_holiday_count(item.company, target_month, target_year),
                    'leave': leave_count,
                    'work_days': len(get_days_in_month(target_year, target_month)) - calculate_holiday_count(item.company, target_month, target_year) - leave_count,
                    'total_leave': leave_count,
                }

                consolidated_entries[employee_id].append(entry)

        all_entries = []
        for employee_id, entries in consolidated_entries.items():
            for entry in entries:
                
                all_entries.append(entry)
    
   
    
    return render(request, 'zohomodules/Attendance/attendance_calendar.html', {'emp_attendance': employee_attendance,'holiday':holidays,'entries':all_entries,'employee':employee,'comments':comment,'calendar_data':calendar_data,'history':history,'allmodules':allmodules})

def attendance_add_comment(request):
    if request.method == 'POST':
        if 'login_id' not in request.session:
            return JsonResponse({'error': 'User not logged in'}, status=401)

        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)

        employee_id = request.POST.get('employee')
        employee = payroll_employee.objects.get(id=employee_id)
        target_month = request.POST.get('target_month')
        target_year = request.POST.get('target_year')
        comment_text = request.POST.get('comment')

        # Create the comment object
        if comment_text:  # Check if comment text is provided
            # Create the comment object
            comment = Attendance_comment(
                comment=comment_text,
                employee=employee,
                month=target_month,
                year=target_year,
                company=company,
                login_details=log_details
            )
            comment.save()

            return JsonResponse({'message': 'Comment added successfully'})
        else:
            return JsonResponse({'error': 'Comment text is required'}, status=400)  # Return an error response if comment text is empty

    return JsonResponse({'error': 'Invalid request method'}, status=405)
    
def delete_attendance_comment(request,id):
    comment = Attendance_comment.objects.get(id=id)    
    comment.delete()  
    return redirect('attendance_calendar', employee_id=comment.employee.id, target_year=comment.year, target_month=comment.month)      
                
def attendance_overview(request, employee_id, target_month, target_year):  
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        employee = payroll_employee.objects.get(id=employee_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=staff.company,date__month=target_month,date__year=target_year,employee=employee)
            allmodules= ZohoModules.objects.get(company=staff.company)    
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=company,date__month=target_month,date__year=target_year,employee=employee)
            allmodules= ZohoModules.objects.get(company=company) 
        
        target_month = max(1, min(target_month, 12))
        target_month = int(target_month)

# Calculate the next month and year if target_month is December
        next_month = 1 if target_month == 12 else target_month + 1
        next_year = target_year + 1 if target_month == 12 else target_year

# Construct the date strings for the start and end of the month
        start_date = datetime(target_year, target_month, 1).date()
        end_date = datetime(next_year, next_month, 1).date() - timedelta(days=1)
        MONTH_NAMES = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December'
}
        
       
        target_month_name = MONTH_NAMES[target_month]

    
# Filter holidays that fall within the target month and year
        days_in_month = get_days_in_month(target_year, target_month)
        current_url = request.build_absolute_uri()

    # Calculate the leave count for the employee
        leave_count = calculate_leave_count(employee, target_month, target_year)

    # Calculate the holiday count for the company
        holiday_count = calculate_holiday_count(employee.company, target_month, target_year)

    # Calculate the working days
        working_days = len(days_in_month) - leave_count - holiday_count

        return render(request,'zohomodules/Attendance/attendance_overview.html',{'current_url': current_url,'items':items,'employee': employee,'tm':target_month,'target_month': target_month_name,'target_year': target_year,'leave_count': leave_count,'holiday_count': holiday_count,'working_days': working_days,'allmodules':allmodules})

def attendance_pdf(request,employee_id,target_month,target_year) :
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        employee = payroll_employee.objects.get(id=employee_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=staff.company,employee=employee,date__month=target_month,date__year=target_year)
                
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            items = Attendance.objects.filter(company=company,employee=employee,date__month=target_month,date__year=target_year)
       
        
        target_month = max(1, min(target_month, 12))
        target_month = int(target_month)

# Calculate the next month and year if target_month is December
        next_month = 1 if target_month == 12 else target_month + 1
        next_year = target_year + 1 if target_month == 12 else target_year

# Construct the date strings for the start and end of the month
        start_date = datetime(target_year, target_month, 1).date()
        end_date = datetime(next_year, next_month, 1).date() - timedelta(days=1)
        
        MONTH_NAMES = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December'
}
        
       
        target_month_name = MONTH_NAMES[target_month]

    

        days_in_month = get_days_in_month(target_year, target_month)

    
        leave_count = calculate_leave_count(employee, target_month, target_year)

    
        holiday_count = calculate_holiday_count(employee.company, target_month, target_year)

    
        working_days = len(days_in_month) - leave_count - holiday_count

        template_path = 'zohomodules/Attendance/attendance_pdf.html'
    context = {
        'items': items,
        'employee': employee,
        'target_month': target_month_name,
        'target_year': target_year,
        'leave_count': leave_count,
        'holiday_count': holiday_count,
        'working_days': working_days
    }

    html = get_template(template_path).render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=attendance.pdf'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response
    
def attendance_email(request,employee_id,target_month,target_year):
    if request.method == 'POST':
        try:
            emails_string = request.POST['email_ids']

                    # Split the string by commas and remove any leading or trailing whitespace
            emails_list = [email.strip() for email in emails_string.split(',')]
            email_message = request.POST['email_message']
            if 'login_id' in request.session:
                log_id = request.session['login_id']
                if 'login_id' not in request.session:
                    return redirect('/')
                log_details = LoginDetails.objects.get(id=log_id)
                employee = payroll_employee.objects.get(id=employee_id)

                if log_details.user_type == 'Staff':
                    staff = StaffDetails.objects.get(login_details=log_details)
                    company=staff.company
                    items = Attendance.objects.filter(company=company,employee=employee,date__month=target_month,date__year=target_year)
                        
                elif log_details.user_type == 'Company':
                    company = CompanyDetails.objects.get(login_details=log_details)
                    items = Attendance.objects.filter(company=company,employee=employee,date__month=target_month,date__year=target_year)
            
                
                target_month = max(1, min(target_month, 12))
                target_month = int(target_month)

        
                next_month = 1 if target_month == 12 else target_month + 1
                next_year = target_year + 1 if target_month == 12 else target_year

        
                start_date = datetime(target_year, target_month, 1).date()
                end_date = datetime(next_year, next_month, 1).date() - timedelta(days=1)
                
                MONTH_NAMES = {
            1: 'January',
            2: 'February',
            3: 'March',
            4: 'April',
            5: 'May',
            6: 'June',
            7: 'July',
            8: 'August',
            9: 'September',
            10: 'October',
            11: 'November',
            12: 'December'
        }
                
            
                target_month_name = MONTH_NAMES[target_month]

            
       
                days_in_month = get_days_in_month(target_year, target_month)

            
                leave_count = calculate_leave_count(employee, target_month, target_year)

           
                holiday_count = calculate_holiday_count(employee.company, target_month, target_year)

           
                working_days = len(days_in_month) - leave_count - holiday_count
                context = {
            'items': items,
            'company':company,
            'employee': employee,
            'target_month': target_month_name,
            'target_year': target_year,
            'leave_count': leave_count,
            'holiday_count': holiday_count,
            'working_days': working_days
        }
                template_path = 'zohomodules/Attendance/attendance_pdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                subject = f"Attendance Details - {company.company_name}"
                email = f"Hi,\nPlease find the attached file for -{employee.first_name} {employee.last_name}. \n{email_message}\n\n--\nRegards,\n{company.company_name}\n{company.address}\n{company.city} - {company.state}\n{company.contact}"
                email_from = settings.EMAIL_HOST_USER

        
                msg = EmailMultiAlternatives(subject, email, email_from, emails_list)
                msg.attach(f'{employee.first_name}_{employee.last_name}_Attendance_Details.pdf', pdf, "application/pdf")
                
                # Send the email
                msg.send()

                messages.success(request, 'Statement has been shared via email successfully..!')
                return redirect(attendance_overview, employee_id, target_month, target_year)

        except Exception as e:
            print(f"Error sending email: {e}")
            messages.error(request, 'An error occurred while sending the email. Please try again later.')
            return redirect(attendance_overview, employee_id, target_month, target_year)
        
def attendance_edit(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company=CompanyDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=company,status='New')
            
            employee = payroll_employee.objects.filter(login_details=log_details,status='Active')
            
        if log_details.user_type=='Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            allmodules=ZohoModules.objects.get(company=staff.company)
            
            employee = payroll_employee.objects.filter(company=staff.company,status='Active')
            
        attendance=Attendance.objects.get(id=id)
        target_month = attendance.date.month
        target_year = attendance.date.year
        return render(request,'zohomodules/Attendance/attendance_edit.html',{'item':attendance,'employee':employee,'tm':target_month,'ty':target_year,'allmodules':allmodules})

def edit_attendance(request,id):
    if request.method =='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
            log_details = LoginDetails.objects.get(id=log_id)
        
            if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                
            
            elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
            
            
            ename = request.POST['employee']
            emp = payroll_employee.objects.get(id=ename)
            date = request.POST['date']
            status = request.POST['status']
            reason = request.POST['reason']
            attendance = get_object_or_404(Attendance, id=id)
            employee_id = attendance.employee.id
            target_month = attendance.date.month
            target_year = attendance.date.year
            attendance.employee=emp
            attendance.date=date
            attendance.status=status
            attendance.reason=reason
            is_holiday = Holiday.objects.filter(company=company, start_date__lte=date, end_date__gte=date).exists()

            if is_holiday:
                    messages.warning(request, 'Selected date is a company holiday.')
                    return redirect('attendance_edit',id=id)
                
            attendance.save()
                
            history = Attendance_History(company=company,login_details=log_details,attendance=attendance,date=date,action='Edited')
            history.save()
            
            
            messages.success(request,'Leave edited successfully!!')
            return redirect('attendance_overview',employee_id,target_month,target_year)
        
def attendance_delete(request,id):
    item = Attendance.objects.get(id=id)
    employee_id = item.employee.id
    target_month = item.date.month
    target_year = item.date.year
    item.delete()
    return redirect('attendance_overview',employee_id,target_month,target_year)

def attendance_create_employee(request):
     if request.method == 'POST':
        # Get login_id from session
        log_id = request.session.get('login_id')
        if not log_id:
            return redirect('/')
        
        
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            
            company = dash_details.company
            print(company)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
        
        
        # Extract data from POST request
        title=request.POST['title']
        fname=request.POST['fname']
        lname=request.POST['lname']
        alias=request.POST['alias']
        joindate=request.POST['joindate']
        salarydate=request.POST['salary']
        saltype=request.POST['saltype']
        if (saltype == 'Fixed'):
            salary=request.POST['fsalary']
        else:
            salary=request.POST['vsalary']
        image=request.FILES.get('file')
        amountperhr=request.POST['amnthr']
        workhr=request.POST['hours'] 
        empnum=request.POST['empnum']
        if payroll_employee.objects.filter(emp_number=empnum,company=company).exists():
            return JsonResponse({'status': 'error', 'message': 'empnum_exists'}, status=400)
            
        designation = request.POST['designation']
        location=request.POST['location']
        gender=request.POST['gender']
        dob=request.POST['dob']
        blood=request.POST['blood']
        fmname=request.POST['fm_name']
        sname=request.POST['s_name']        
        add1=request.POST['address']
        add2=request.POST['address2']
        address=add1+" "+add2
        padd1=request.POST['paddress'] 
        padd2=request.POST['paddress2'] 
        paddress= padd1+padd2
        phone=request.POST['phone']
        ephone=request.POST['ephone']
        result_set1 = payroll_employee.objects.filter(company=company,Phone=phone)
        result_set2 = payroll_employee.objects.filter(company=company,emergency_phone=ephone)
        if result_set1:
            messages.error(request,'phone no already exists')
            return redirect('company_mark_attendance')
        if result_set2:
            messages.error(request,'phone no already exists')
            return redirect('company_mark_attendance')
        email=request.POST['email']
        result_set = payroll_employee.objects.filter(company=company,email=email)
        if result_set:
            messages.error(request,'email already exists')
            return redirect('company_mark_attendance')
        isdts=request.POST['tds']
        attach=request.FILES.get('attach')
        if isdts == '1':
            istdsval=request.POST['pora']
            if istdsval == 'Percentage':
                tds=request.POST['pcnt']
            elif istdsval == 'Amount':
                tds=request.POST['amnt']
        else:
                istdsval='No'
                tds = 0
        itn=request.POST['itn']
        an=request.POST['an']
        if payroll_employee.objects.filter(Aadhar=an,company=company):
                messages.error(request,'Aadhra number already exists')
                return redirect('company_mark_attendance')   
        uan=request.POST['uan'] 
        pfn=request.POST['pfn']
        pran=request.POST['pran']
        age=request.POST['age']
        bank=request.POST['bank']
        accno=request.POST['acc_no']       
        ifsc=request.POST['ifsc']       
        bname=request.POST['b_name']       
        branch=request.POST['branch']
        ttype=request.POST['ttype']
        try:
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                                emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                                amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                                UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=company,login_details=log_details)
            payroll.save()
            history=employee_history(company=company,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            new_employee_id = payroll.id  
            new_employee_name = f"{fname} {lname}"
            
            
            data = {
                'status': 'success',
                'employee_id': new_employee_id,
                'employee_name': new_employee_name
            }
            
            
            return JsonResponse(data)
        except Exception as e:
           
            error_message = str(e)
            return JsonResponse({'status': 'error', 'message': error_message})
            
def attendance_employee_dropdown(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company=CompanyDetails.objects.get(login_details=log_details)
            options = {}
            option_objects = payroll_employee.objects.filter(company=company,status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
            
        if log_details.user_type=='Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            options = {}
            option_objects = payroll_employee.objects.filter(company=staff.company,status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
           
            
def attendance_add_blood(request):
     if request.method == "POST":
        blood = request.POST.get('blood')

        # Check if the blood group already exists
        existing_entry = Bloodgroup.objects.filter(Blood_group=blood).first()

        if existing_entry:
            # Blood group already exists, return an appropriate message
            return JsonResponse({'blood': blood, 'message': 'Blood group already exists'})

        # Blood group doesn't exist, create a new entry
        Bloodgroup.objects.create(Blood_group=blood)
        return JsonResponse({'blood': blood, 'message': 'Blood group saved successfully'})

     return JsonResponse({'message': 'Invalid request method'}, status=400)
     
def attendance_import(request):
    if request.method == 'POST' and 'file' in request.FILES:
        if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
            log_details = LoginDetails.objects.get(id=log_id)

            if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                    
            elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)

            excel_file = request.FILES['file']
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                Employee_No, date, status, reason = row
                if not any(row):
                  continue 
                    
                for employee in payroll_employee.objects.filter(emp_number=Employee_No, company=company):
                        leave_exists = Attendance.objects.filter(employee=employee, company=company, date=date).exists()
                        
                        if not leave_exists:
                            attendance = Attendance.objects.create(
                                employee=employee,
                                company=company,
                                login_details=log_details,
                                date=date,
                                status=status,
                                reason=reason
                            )
                        
                    

                        
                            history = Attendance_History.objects.create(
                                company=company,
                                login_details=log_details,
                                attendance=attendance,
                                date=date,
                                action='Created'
                            )
               
                
                
            return redirect('company_attendance_list')

    return HttpResponse("No file uploaded or invalid request method")

#---------------- Zoho Final Attendance - Meenu Shaju - End--------------------
# ------------------------------- GOKUL KRISHNA UR -----------------------------------------
def SalaryDetailsListPage(request):
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                employees=SalaryDetails.objects.filter(company=dash_details.company)
                context = {
                    'employees':employees,
                    'allmodules': allmodules,
                    'details': dash_details,    
                    }
                return render(request,'zohomodules/SalaryDetails/SalaryDetailsListPage.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            employees=SalaryDetails.objects.filter(company=dash_details)
            company = CompanyDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            context = {
                'employees':employees,
                'allmodules': allmodules,
                'company':company,
                'details': dash_details,    
            }
        
        return render (request, 'zohomodules/SalaryDetails/SalaryDetailsListPage.html',context )
     
def EditSalaryDetails(request,id):
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        current_year = datetime.now().year
        end_year = current_year + 5
        current_date = datetime.now().strftime('%Y-%m-%d')
        range_year_list = range(current_year, end_year + 1)
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
            employees=payroll_employee.objects.filter(company=company)
            salary=SalaryDetails.objects.get(id=id)
            formatted_join_date = salary.employee.joindate.strftime('%Y-%m-%d')
            formatted_salary_date= salary.salary_date.strftime('%Y-%m-%d')
            attendance = Attendance.objects.filter(employee=salary.employee.id)
            holyday =Holiday.objects.filter(company=company)
            total_rows = 0
            holyday_rows = 0
            for attendance in attendance:
                total_rows += 1
            for holyday in holyday:
                holyday_rows += 1
            context = {
              
                'employees':employees,
                'range_year_list':range_year_list,
                'current_date':current_date,
                'salary':salary,
                'formatted_join_date':formatted_join_date,
                'formatted_salary_date':formatted_salary_date,
                'total_rows':total_rows,
                'holyday_rows':holyday_rows,
             
            }
            return render(request,'zohomodules/SalaryDetails/EditSalaryDetails.html',context) 
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
            employees=payroll_employee.objects.filter(company=company)
            salary=SalaryDetails.objects.get(id=id)
            formatted_join_date = salary.employee.joindate.strftime('%Y-%m-%d')
            formatted_salary_date= salary.salary_date.strftime('%Y-%m-%d')
            
            attendance = Attendance.objects.filter(employee=salary.employee.id)
            holyday =Holiday.objects.filter(company=company)
            total_rows = 0
            holyday_rows = 0
            for attendance in attendance:
                total_rows += 1
            for holyday in holyday:
                holyday_rows += 1
            context = {
              
                'employees':employees,
                'range_year_list':range_year_list,
                'current_date':current_date,
                'salary':salary,
                'formatted_join_date':formatted_join_date,
                'formatted_salary_date':formatted_salary_date,
                'total_rows':total_rows,
                'holyday_rows':holyday_rows,
           
            }
  
        return render(request,'zohomodules/SalaryDetails/EditSalaryDetails.html',context)

def SalaryDetailsOverViewPageWithId(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                allmodules= ZohoModules.objects.get(company=dash_details.company)
                employees=SalaryDetails.objects.filter(company=dash_details.company )
                employee =SalaryDetails.objects.get(company=dash_details.company,id=id)
                attendance =Attendance.objects.filter(company=dash_details.company,employee=employee.employee)
                holydays= Holiday.objects.filter(company=dash_details.company)
                total_rows = 0
                holyday_rows= 0
                for attendance in attendance:  
                    total_rows += 1
                for holyday in holydays:
                    holyday_rows += 1
                comment= CommentSalaryDetails.objects.filter(company=dash_details.company,salary_details=id)
                history = HistorySalaryDetails.objects.filter(company=dash_details.company,salary_details=id )
                total = employee.total_amount
                total_percentage = employee.basic_salary + employee.conveyance_allowance + employee.other_allowance + employee.hra
                print('total_percentage',total_percentage)
                basic_salary_amount = (employee.basic_salary / total_percentage) * total
                conveyance_allowance_amount = (employee.conveyance_allowance / total_percentage) * total
                hra_amount = (employee.hra / total_percentage) * total
                other_allowance_amount = (employee.other_allowance / total_percentage) * total
                salary = SalaryDetails.objects.filter(company=dash_details.company).first()
                basic_salary_deduction = basic_salary_amount - salary.basic_salary
                print('deduction',basic_salary_deduction)
                allowance_amounts = {
                    "Basic_Salary": basic_salary_amount,
                    "Conveyance_Allowance": conveyance_allowance_amount,
                    "HRA": hra_amount,
                    "Other_Allowance": other_allowance_amount,
                    'basic_salary_deduction':basic_salary_deduction,
                }
                print("Corrected Basic Salary Amount:", basic_salary_amount)
                print("Corrected Conveyance Allowance Amount:", conveyance_allowance_amount)
                print("Corrected HRA Amount:", hra_amount)
                print("Corrected Other Allowance Amount:", other_allowance_amount)
                
                return render(request,'zohomodules/SalaryDetails/SalaryDetailsOVerViewPage.html', {'details':dash_details,'allowance_amounts':allowance_amounts,'allmodules':allmodules,'details':dash_details, 'history':history,'comment':comment,'employees':employees,'allmodules': allmodules,'employee':employee,'attendance':total_rows,'holyday':holyday_rows})
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            employees=SalaryDetails.objects.filter(company=dash_details )
            employee =SalaryDetails.objects.get(company=dash_details,id=id)
            attendance =Attendance.objects.filter(company=dash_details,employee=employee.employee )
            holydays= Holiday.objects.filter(company=dash_details)
            total_rows = 0
            holyday_rows= 0
            for attendance in attendance:   
                total_rows += 1
            for holyday in holydays:
                holyday_rows += 1
            comment= CommentSalaryDetails.objects.filter(company=dash_details,salary_details=id)
            history = HistorySalaryDetails.objects.filter(company=dash_details,salary_details=id )
            total = employee.total_amount
            total_percentage = employee.basic_salary + employee.conveyance_allowance + employee.other_allowance + employee.hra
            print('total_percentage',total_percentage)
            basic_salary_amount = (employee.basic_salary / total_percentage) * total
            conveyance_allowance_amount = (employee.conveyance_allowance / total_percentage) * total
            hra_amount = (employee.hra / total_percentage) * total
            other_allowance_amount = (employee.other_allowance / total_percentage) * total
            salary = SalaryDetails.objects.filter(company=dash_details).first()
            basic_salary_deduction = basic_salary_amount - salary.basic_salary
            allowance_amounts = {
                "Basic_Salary": basic_salary_amount,
                "Conveyance_Allowance": conveyance_allowance_amount,
                "HRA": hra_amount,
                "Other_Allowance": other_allowance_amount,
                "basic_salary_deduction":basic_salary_deduction,
            }
            print("Corrected Basic Salary Amount:", basic_salary_amount)
            print("Corrected Conveyance Allowance Amount:", conveyance_allowance_amount)
            print("Corrected HRA Amount:", hra_amount)
            print("Corrected Other Allowance Amount:", other_allowance_amount)
            return render (request, 'zohomodules/SalaryDetails/SalaryDetailsOVerViewPage.html', {'allowance_amounts':allowance_amounts,'details':dash_details,'employee':employee,'comment':comment,'employees':employees,'allmodules': allmodules, 'attendance':total_rows,'holyday':holyday_rows} )

def CreateSalaryDetailsFunction(request):
    if request.method == 'POST':
        if 'login_id' not in request.session:
            return redirect('/')
        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
        salary_date_str = request.POST.get('salary_date')
        salary_date = datetime.strptime(salary_date_str, '%Y-%m-%d').date()
        year = request.POST.get('year')
        other_cuttings = request.POST.get('other_cuttings')
        holiday = request.POST.get('Holiday')
        month = request.POST.get('month')
        description = request.POST.get('Description')
        employee_id = request.POST.get('employee_id')
        print('emplyeeeeeeeee',employee_id)
        employee= payroll_employee.objects.get(emp_number=employee_id)
        
        casual_leave = request.POST.get('Casual_leave')
        working_day = request.POST.get('working_days')
        bonus = request.POST.get('bonus')
        calculated_salary = request.POST.get('calculated_salary')
        basic_salary = float(request.POST.get('Basic_Salary'))
        conveyance_allowance = float(request.POST.get('Conveyance_Allowance'))
        hra = float(request.POST.get('HRA'))
        other_allowance = float(request.POST.get('Other_Allowance'))
        attendance = Attendance.objects.filter(employee_id=employee_id, company=company).first()
        draft = request.POST.get('draft')
        save = request.POST.get('save')

        total = basic_salary + conveyance_allowance + hra + other_allowance
        basic_salary_percentage = (basic_salary / total) * 100
        conveyance_allowance_percentage = (conveyance_allowance / total) * 100
        hra_percentage = (hra / total) * 100
        other_allowance_percentage = (other_allowance / total) * 100
        total_percentage = basic_salary_percentage + conveyance_allowance_percentage + hra_percentage + other_allowance_percentage
        percentage_difference = 100 - total_percentage
        basic_salary_percentage += percentage_difference * (basic_salary_percentage / total_percentage)
        conveyance_allowance_percentage += percentage_difference * (conveyance_allowance_percentage / total_percentage)
        hra_percentage += percentage_difference * (hra_percentage / total_percentage)
        other_allowance_percentage += percentage_difference * (other_allowance_percentage / total_percentage)

        # total = basic_salary + conveyance_allowance + hra + other_allowance
        # basic_salary_percentage = round((basic_salary / total) * 100)
        # conveyance_allowance_percentage = round((conveyance_allowance / total) * 100)
        # hra_percentage = round((hra / total) * 100)
        # other_allowance_percentage = round((other_allowance / total) * 100)
        
        if draft:   
            DraftorSave = "Draft"
        else:   
            DraftorSave = "Save"
        
        SalaryDetails.objects.create(
            year=year,
            other_cuttings=other_cuttings,
            employee=employee,
            month=month,
            attendance=attendance,
            casual_leave=casual_leave,
            description=description,
            add_bonus=bonus,
            salary=calculated_salary,
            salary_date=salary_date,
            basic_salary=basic_salary_percentage,
            hra=hra_percentage,
            other_allowance=other_allowance_percentage,
            conveyance_allowance=conveyance_allowance_percentage,
            DraftorSave=DraftorSave,
            company=company,
            holiday=holiday,
            total_amount=total,total_working_days=working_day
        )


        salary = SalaryDetails.objects.filter(company=company)
        ids=salary.last()

        HistorySalaryDetails.objects.create(
        login_details=log_details,company=company,salary_details=ids,
                action='CREATED'
        )
        
        return redirect('SalaryDetailsOverViewPageWithId', ids.id )
    
    return render(request, 'zohomodules/SalaryDetails/CreateSalaryDetails.html')

def SharePayslipMail(request, id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
        elif log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        else:
            return HttpResponse('<script>alert("Invalid user type!");window.location="/"</script>')
        
        employee = SalaryDetails.objects.get(company=dash_details, id=id)
        attendance = Attendance.objects.filter(company=dash_details, employee=employee.employee)
        holydays = Holiday.objects.filter(company=dash_details)
        total_rows = attendance.count()
        holyday_rows = holydays.count()
        
        total = employee.total_amount
        total_percentage = employee.basic_salary + employee.conveyance_allowance + employee.other_allowance + employee.hra
        basic_salary_amount = (employee.basic_salary / total_percentage) * total
        conveyance_allowance_amount = (employee.conveyance_allowance / total_percentage) * total
        hra_amount = (employee.hra / total_percentage) * total
        other_allowance_amount = (employee.other_allowance / total_percentage) * total
        salary = SalaryDetails.objects.filter(company=dash_details).first()
        basic_salary_deduction = basic_salary_amount - salary.basic_salary
        allowance_amounts = {
            "Basic_Salary": basic_salary_amount,
            "Conveyance_Allowance": conveyance_allowance_amount,
            "HRA": hra_amount,
            "Other_Allowance": other_allowance_amount,
            "basic_salary_deduction":basic_salary_deduction,
        }
        
        my_subject = "PAYSLIP"
        emails_string = request.POST.get('email_ids', '')
        emails_list = [email.strip() for email in emails_string.split(',')]

        html_message = render_to_string('zohomodules/SalaryDetails/payslip_pdf.html', {'allowance_amounts': allowance_amounts, 'details': dash_details, 'employee': employee, 'attendance': total_rows, 'holyday': holyday_rows}) 
        plain_message = strip_tags(html_message)
        
        pdf_content = BytesIO()
        pisa_document = pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content) 
        pdf_content.seek(0)
        
        filename = f'Payslip_{dash_details.company_name}.pdf'
        
        subject = "Payslip"
  
        
        message = EmailMultiAlternatives(
            subject=subject,
            body=plain_message,
            to=emails_list,
        )
        
        message.attach_alternative(html_message, "text/html")
        message.attach(filename, pdf_content.read(), 'application/pdf')

        try:
            message.send()
            return HttpResponse('<script>alert("Report has been shared successfully!");window.location="/SalaryDetailsOverViewPageWithId/' + str(id) + '"</script>')
        except Exception as e:
            return HttpResponse('<script>alert("Failed to send email!");window.location="/SalaryDetailsOverViewPageWithId/' + str(id) + '"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/SalaryDetailsOverViewPage/' + str(id) + '"</script>')

     

def CreateSalaryDetails(request): 
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        current_year = datetime.now().year
        end_year = current_year + 5
        current_date = datetime.now().strftime('%Y-%m-%d')
        range_year_list = range(current_year, end_year + 1)
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
            employees=payroll_employee.objects.filter(company=dash_details.company)
            salary = SalaryDetails.objects.filter(company=dash_details.company).first()
            salary = SalaryDetails.objects.filter(company=dash_details).first()
            if salary is not None:
                total = salary.total_amount
                total_percentage = salary.basic_salary + salary.conveyance_allowance + salary.other_allowance + salary.hra
                print('total_percentage',total_percentage)
                basic_salary_amount = (salary.basic_salary / total_percentage) * total
                conveyance_allowance_amount = (salary.conveyance_allowance / total_percentage) * total
                hra_amount = (salary.hra / total_percentage) * total
                other_allowance_amount = (salary.other_allowance / total_percentage) * total
            else:
                basic_salary_amount = None
                conveyance_allowance_amount =None
                hra_amount = None
                other_allowance_amount =None

       
            context = {
                'employees':employees,
                'allmodules': allmodules,
                'range_year_list':range_year_list,
                'current_date':current_date,
                'details':dash_details,
                'salary':salary,
                "Basic_Salary": basic_salary_amount,
                "Conveyance_Allowance": conveyance_allowance_amount,
                "HRA": hra_amount,
                 "Other_Allowance": other_allowance_amount
            }
            return render (request, 'zohomodules/SalaryDetails/CreateSalaryDetails.html',context )
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            employees=payroll_employee.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            
            salary = SalaryDetails.objects.filter(company=dash_details).first()
            if salary is not None:
                total = salary.total_amount
                total_percentage = salary.basic_salary + salary.conveyance_allowance + salary.other_allowance + salary.hra
                print('total_percentage',total_percentage)
                basic_salary_amount = (salary.basic_salary / total_percentage) * total
                conveyance_allowance_amount = (salary.conveyance_allowance / total_percentage) * total
                hra_amount = (salary.hra / total_percentage) * total
                other_allowance_amount = (salary.other_allowance / total_percentage) * total
            else:
                basic_salary_amount = None
                conveyance_allowance_amount =None
                hra_amount = None
                other_allowance_amount =None
            context = {
              
                'employees':employees,
                'allmodules': allmodules,
                'range_year_list':range_year_list,
                'current_date':current_date,
                'details':dash_details,
                'salary':salary,
                "Basic_Salary": basic_salary_amount,
                "Conveyance_Allowance": conveyance_allowance_amount,
                "HRA": hra_amount,
                "Other_Allowance": other_allowance_amount,
            }
        
        return render (request, 'zohomodules/SalaryDetails/CreateSalaryDetails.html',context )
     

def addCommentSalaryDetails(request,id):
    if request.method == 'POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        company = CompanyDetails.objects.get(id=dash_details.company.id)
    elif log_details.user_type == 'Company':
        company = CompanyDetails.objects.get(login_details=log_details)
    salary = SalaryDetails.objects.get(id=id) 
    comments = request.POST.get('comments')
    CommentSalaryDetails.objects.create(comment=comments,company=company,employee=salary.employee,salary_details=salary)
    return redirect('SalaryDetailsOverViewPageWithId', salary.id)


def DeleteCommentSalaryDetails(request, id):
    comment = CommentSalaryDetails.objects.get(id=id)
    salary_id = comment.employee.salarydetails_set.first().id
    comment.delete()
    return redirect('SalaryDetailsOverViewPageWithId', salary_id)


 

     
def ImportSalaryDetails(request):
    if request.method == 'POST' and 'empfile' in request.FILES:
        if 'login_id' not in request.session:
            return redirect('/')
        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
        excel_file = request.FILES['empfile']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            emp_number,year, month ,salary_date, other_cuttings,casual_leave,description,bonus,DraftorSave,basic_salary,conveyance_allowance,hra,other_allowance= row
            employee=payroll_employee.objects.get(emp_number=emp_number)
            attendance = Attendance.objects.filter(employee=employee)
            total = basic_salary + conveyance_allowance + hra + other_allowance
            basic_salary_percentage = float((basic_salary / total) * 100)
            conveyance_allowance_percentage = float((conveyance_allowance / total) * 100)
            hra_percentage = float((hra / total) * 100)
            other_allowance_percentage = float((other_allowance / total) * 100)
            salary = employee.salary
            total_rows = 0
            for attendance in attendance:
                total_rows += 1
            days=calendar.monthrange(year, month)[1]
            totalWorkedDays = days - total_rows + casual_leave
            calculatedSalary = (salary / days) * totalWorkedDays + bonus - other_cuttings ; 
            SalaryDetails.objects.create(
                employee=employee,
                year=year,
                other_cuttings=other_cuttings,
                month=month,
                attendance=attendance,
                casual_leave=casual_leave,
                description=description,
                add_bonus=bonus,
                salary=calculatedSalary,
                salary_date=salary_date,
                basic_salary =basic_salary_percentage,
                hra =hra_percentage,
                other_allowance=other_allowance_percentage,
                conveyance_allowance=conveyance_allowance_percentage,
                DraftorSave =DraftorSave,
                company=company,
            )
            HistorySalaryDetails.objects.create(
            login_details=log_details,
         
                company=company,
                action='CREATED'
        )
        return redirect('SalaryDetailsListPage')




def addEmployeeFromSalaryDetails(request):
    if request.method == 'POST':
        if 'login_id' not in request.session:
            return redirect('/')
        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)

        title = request.POST.get('title')
        first_name = request.POST.get('first_name')
        lname = request.POST.get('lname')
        alias = request.POST.get('alias')
        joindate = request.POST.get('joindate')
        salarydate = request.POST.get('salary')
        saltype = request.POST.get('saltype')
        if saltype == 'Fixed':
            salary = request.POST.get('fsalary')
        else:
            salary = request.POST.get('vsalary')
        image = request.FILES.get('file')
        amountperhr = request.POST.get('amnthr')
        workhr = request.POST.get('hours')
        empnum = request.POST.get('empnum')
        # if payroll_employee.objects.filter(emp_number=empnum, company=company).exists():
        #     messages.info(request, 'Employee number already exists')
        #     return redirect('payroll_employee_create')
        designation = request.POST.get('designation')
        location = request.POST.get('location')
        gender = request.POST.get('gender')
        dob = request.POST.get('dob')
        blood = request.POST.get('blood')
        fmname = request.POST.get('fm_name')
        sname = request.POST.get('s_name')
        add1 = request.POST.get('address')
        add2 = request.POST.get('address2')
        address = add1 + " " + add2
        padd1 = request.POST.get('paddress')
        padd2 = request.POST.get('paddress2')
        paddress = padd1 + padd2
        phone = request.POST.get('phone')
        ephone = request.POST.get('ephone')
        # if payroll_employee.objects.filter(company=company, Phone=phone).exists():
        #     messages.error(request, 'Phone number already exists')
        #     return redirect('CreateSalaryDetails')
        # if payroll_employee.objects.filter(company=company, emergency_phone=ephone).exists():
        #     messages.error(request, 'Emergency phone number already exists')
        #     return redirect('CreateSalaryDetails')
        email = request.POST.get('email')
        # if payroll_employee.objects.filter(company=company, email=email).exists():
        #     messages.error(request, 'Email already exists')
        #     return redirect('CreateSalaryDetails')
        isdts = request.POST.get('tds')
        attach = request.FILES.get('attach')
        if isdts == '1':
            stdsval = request.POST.get('pora')
            if stdsval == 'Percentage':
                tds = request.POST.get('pcnt')
            elif stdsval == 'Amount':
                tds = request.POST.get('amnt')
            else:
                stdsval = 'No'
                tds = 0
        else:
            tds = 0

        itn = request.POST.get('itn')
        an = request.POST.get('an')
        # if payroll_employee.objects.filter(Aadhar=an, company=company).exists():
        #     messages.error(request, 'Aadhar number already exists')
        #     return redirect('CreateSalaryDetails')
        uan = request.POST.get('uan')
        pfn = request.POST.get('pfn')
        pran = request.POST.get('pran')
        age = request.POST.get('age')
        bank = request.POST.get('bank')
        accno = request.POST.get('acc_no')
        ifsc = request.POST.get('ifsc')
        bname = request.POST.get('b_name')
        branch = request.POST.get('branch')
        ttype = request.POST.get('ttype')

        payroll = payroll_employee(
            title=title, first_name=first_name, last_name=lname, alias=alias, image=image,
            joindate=joindate, salary_type=saltype, salary=salary, age=age,
            emp_number=empnum, designation=designation, location=location, gender=gender,
            dob=dob, blood=blood, parent=fmname, spouse_name=sname, workhr=workhr,
            amountperhr=amountperhr, address=address, permanent_address=paddress,
            Phone=phone, emergency_phone=ephone, email=email, Income_tax_no=itn, Aadhar=an,
            UAN=uan, PFN=pfn, PRAN=pran, uploaded_file=attach, isTDS=isdts,
            TDS_percentage=tds, salaryrange=salarydate, acc_no=accno, IFSC=ifsc,
            bank_name=bname, branch=branch, transaction_type=ttype,company=company  
             
        )

        payroll.save()
        history=employee_history(company=company,login_details=log_details, employee=payroll,Action='CREATED')
        history.save()


    return redirect('CreateSalaryDetails')

   

 


def EditSalaryDetailsFunction(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
            return redirect('/')
    if request.method == 'POST':
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
        salary= SalaryDetails.objects.get(id=id)
        salary_date_str = request.POST.get('salary_date')
        salary.salary_date = datetime.strptime(salary_date_str, '%Y-%m-%d').date()
        salary.year = request.POST.get('year')
        salary.other_cuttings = request.POST.get('other_cuttings')
        salary.month = request.POST.get('month')
        salary.description = request.POST.get('Description')
        salary.employee = request.POST.get('id')
        salary.casual_leave = request.POST.get('Casual_leave')
        salary.total_working_days = request.POST.get('working_day')
        salary.add_bonus = request.POST.get('bonus')
        employee_id = request.POST.get('employee_id')
        salary.employee = get_object_or_404(payroll_employee, id=employee_id)  
        # salary.salary= request.POST.get('calculated_salary')
        salary.save()
        existing_entry = HistorySalaryDetails.objects.filter(
            login_details=log_details,
            action='EDITED'
        ).exists()

        salary = SalaryDetails.objects.get(company=company,id=id)

        if not existing_entry:
            HistorySalaryDetails.objects.create(
                login_details=log_details,salary_details=salary,company=company,
                action='EDITED'
            )
        return redirect('SalaryDetailsOverViewPageWithId', salary.id)
    return render (request, 'zohomodules/SalaryDetails/SalaryDetailsOverViewPage.html' )      


def SalaryDetailsActiveAndInnactive(request, id):
    if request.method == 'GET':
        salary = SalaryDetails.objects.get(id=id)
        if salary.status == 'Active':
            salary.status = 'Inactive'
        else:
            salary.status = 'Active'
        salary.save()
    return redirect('SalaryDetailsOverViewPageWithId', salary.id)

 

def SalaryDetailsAddBloodGroup(request):
    if request.method == "POST":
        if 'login_id' not in request.session:
            return redirect('/')
        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            company = CompanyDetails.objects.get(id=dash_details.company.id)
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)
        blood = request.POST.get('blood')
        Bloodgroup.objects.create(Blood_group=blood, company=company)
        return JsonResponse({'blood': blood})
    return render(request, 'zohomodules/SalaryDetails/CreateSalaryDetails.html')

  
def SalaryDetailsConvert(request, id):
    if request.method == 'GET':
        salary = SalaryDetails.objects.get(id=id)
        if salary.DraftorSave == 'Draft':
            salary.DraftorSave = 'Save'
            salary.save()   
    return redirect('SalaryDetailsOverViewPage')


def custdata(request):
    cid = request.POST.get('id')
    employee = payroll_employee.objects.get(id=cid)
    attendance =Attendance.objects.filter(id=employee.id)
    holyday = Holiday.objects.filter(id=employee.id)
    total_rows = 0
    holyday_rows = 0
    for attendance in attendance:
        total_rows += 1
    for holyday in holyday:
        holyday_rows += 1
    data = {
        'email': employee.email,
        'join_date': employee.joindate.strftime('%Y-%m-%d'),
        'employee_id':employee.emp_number,
         'salary_id':employee.salary,
         'Designation_id':employee.designation,
         'total_rows':total_rows,
         'holyday':holyday_rows,
    }
    return JsonResponse(data)

def SalaryDetailsDelete (request,id):
    salary_details = SalaryDetails.objects.get(id=id)
    salary_details.delete()
    return redirect('CreateSalaryDetails')
 
# ------------------------------- GOKUL KRISHNA UR -----------------------------------------

#-------------------------employeeloan views by haripriya-------------#
def employee_listpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        pay= payroll_employee.objects.filter(company=dash_details)
        emp=EmployeeLoan.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        emp=EmployeeLoan.objects.filter(company=dash_details.company)
        pay= payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
                'details': dash_details,
                'emp':emp,
                'allmodules': allmodules,
                'log_id':log_details,
                'pay':pay,
                
        }
    return render(request,'zohomodules/employe_loan/employee_loan_list.html',content)   

def employeeloan_create(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    blood=Bloodgroup.objects.all()
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        pay = payroll_employee.objects.filter(company=dash_details,status='active')
        dur =LoanDuration.objects.filter(company=dash_details)
     
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        
        
        toda = date.today()
        tod = toda.strftime("%Y-%m-%d") 
    
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        pay = payroll_employee.objects.filter(company=dash_details.company,status='active')
        dur =LoanDuration.objects.filter(company=dash_details.company)
        bank=Banking.objects.filter(company=dash_details.company)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        
        toda = date.today()
        tod = toda.strftime("%Y-%m-%d") 
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'log_id':log_details,
            'pay':pay,
            'dur':dur,
            'tod':tod,
            'bank':bank,
            'blood':blood       
    }
    return render(request,'zohomodules/employe_loan/employee_loan_create.html',content)


def check_user_loan(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
  
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        emp = request.GET.get('emp',None)
        print('done')
        print(emp)
        data = {
            'is_tak': EmployeeLoan.objects.filter(Employee=emp).exists()
        }
        if data['is_tak']:
            data['error_message'] = 'Loan  already Taken.'
        return JsonResponse(data)
    elif log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        emp = request.GET.get('emp',None)
        print('done')
        print(emp)
        data = {
            'is_tak': EmployeeLoan.objects.filter(Employee=emp).exists()
        }
        if data['is_tak']:
            data['error_message'] = 'Loan  already Taken.'
        return JsonResponse(data)
    

def listemployee_loan(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        id = request.POST.get('id')
        cust = payroll_employee.objects.get(id = id, company = dash_details)
        email = cust.email
        employeeno = cust.emp_number
        joindate = cust.joindate
        amount = cust.salary
   
        print(email)
        print(employeeno)
       
        return JsonResponse({'email': email,'employeeno': employeeno,'joindate':joindate,'amount': amount},safe=False)

    elif log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        id = request.POST.get('id')
        cust = payroll_employee.objects.get(id = id, company = dash_details.company)
        email = cust.email
        employeeno = cust.emp_number
        joindate = cust.joindate
        amount = cust.salary
        
        print(email)
        print(employeeno)
       
        return JsonResponse({'email': email,'employeeno': employeeno,'joindate':joindate,'amount': amount},safe=False)    


def addemployeloan(request):
    if request.method=='POST':
        if 'login_id' in request.session:
            if request.session.has_key('login_id'):
                log_id = request.session['login_id']
            else:
                return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == "Company":
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            empid = request.POST['employee']
            print(empid)
            print('emploree')
            employee = payroll_employee.objects.get(id=empid)
            Loan_Amound = request.POST['Loan_Amount'] 
            loandate = request.POST['loandate'] 
            experydate = request.POST['experydate']
            cuttingPercentage = request.POST['cuttingPercentage']
            cuttinamount = request.POST['Cutingamount']
            duration = request.POST['loan_duration']
            cheque_id = request.POST['cheque_id']  
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
           
            
            try:
                file = request.FILES['file']
            except:
                file = '' 
            Note = request.POST['Note']
    

            data=EmployeeLoan(Employee=employee,LoanAmount=Loan_Amound,Loandate=loandate,Expiry_date=experydate,note=Note,file=file,company=dash_details,
            balance=Loan_Amound,upi_id=upi_id,cheque_number=cheque_id,bank_acc_number=bnk_id,payment_method=payment_method,login_details=log_details,duration=duration)
            

            if int(cuttingPercentage)==0 and int(cuttinamount)!=0:
                data.MonthlyCut_Amount=cuttinamount
                data.MonthlyCut_percentage=0
                data.Monthly_payment_type='No'         
            else: 
                data.MonthlyCut_percentage= cuttingPercentage  
                data.MonthlyCut_Amount = ((int(cuttingPercentage)/100)*int(Loan_Amound)) 
                data.Monthly_payment_type='Yes'   
            
            data.save()

            
            history=Employeeloan_history(company=dash_details,login_details=log_details, employeeloan=data, action='CREATED')
            history.save()
            EmployeeLoanRepayment.objects.create(
                emp=data,
                principal_amount = Loan_Amound,
                interest_amonut = 0,
                total_payment = 0,
                payment_date = loandate,
                balance = Loan_Amound,
                payment_method = payment_method,
                upi_id = upi_id,
                cheque_id = cheque_id,
                bank_id = bnk_id,
                particular = 'LOAN ISSUED',
                company=dash_details,
                logindetails=log_details,
                employee=employee
            )

        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details) 
            empid = request.POST['employee']
            print(empid)
            print('emploree')
            employee = payroll_employee.objects.get(id=empid,company=dash_details.company)
            Loan_Amound = request.POST['Loan_Amount'] 
            loandate = request.POST['loandate'] 
            experydate = request.POST['experydate']
            cuttingPercentage = request.POST['cuttingPercentage']
            cuttinamount = request.POST['Cutingamount']
            duration = request.POST['loan_duration']
            cheque_id = request.POST['cheque_id'] 
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
            
            try:
                file = request.FILES['file']
            except:
                file = '' 
            Note = request.POST['Note']
            
        

            data=EmployeeLoan(Employee=employee,LoanAmount=Loan_Amound,Loandate=loandate,Expiry_date=experydate,note=Note,file=file,company=dash_details.company,
            upi_id=upi_id,cheque_number=cheque_id,payment_method=payment_method,bank_acc_number=bnk_id,login_details=log_details,balance=Loan_Amound,duration=duration)
            
            if int(cuttingPercentage)==0 and int(cuttinamount)!=0:
                data.MonthlyCut_Amount=cuttinamount
                data.MonthlyCut_percentage=0
                data.Monthly_payment_type='No'         
            else: 
                data.MonthlyCut_percentage= cuttingPercentage  
                data.MonthlyCut_Amount = ((int(cuttingPercentage)/100)*int(Loan_Amound)) 
                data.Monthly_payment_type='Yes'
            
            data.save()
            
            history=Employeeloan_history(company=dash_details.company,login_details=log_details, employeeloan=data,action='CREATED')
            history.save()

            EmployeeLoanRepayment.objects.create(
                emp=data,
                principal_amount = Loan_Amound,
                interest_amonut = 0,
                total_payment = 0,
                payment_date = loandate,
                balance = Loan_Amound,
                payment_method = payment_method,
                upi_id = upi_id,
                cheque_id = cheque_id,
                bank_id = bnk_id,
                particular = 'LOAN ISSUED',
                company=dash_details.company,
                logindetails=log_details,
                employee=employee
            )

        return redirect('employee_listpage')

        
def createpayroll2(request):
    if request.method=='POST':
        print("hii")
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Company':    
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            title=request.POST['title']
            print(title)
            fname=request.POST['fname']
            print(fname)
            lname=request.POST['lname']
            print(lname)
            alias=request.POST['alias']
            print(alias)
            joindate=request.POST.get('joindate')
            print(joindate)
            salarydate=request.POST['salarydate']
            print(salarydate)

            saltype=request.POST['saltype']
            print(saltype)
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
                print(salary)
            else:
                salary=request.POST['vsalary']
                print(salary)
            
            image=request.FILES.get('file')
            print(image)
            amountperhr=request.POST['amnts']
            print(amountperhr)
            workhr=request.POST['hours'] 
            print(workhr)
            empnum=request.POST['empnum']
            print(empnum)
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details):
                messages.info(request,'employee number all ready exists')
                return redirect('employeeloan_create')
            designation = request.POST['designation']
            print(designation)
            location=request.POST['location']
            print(location)
            gender=request.POST['gender']
            print(gender)
            dob=request.POST['dob']
            print(dob)
            blood=request.POST['blood']
            print(blood)
            fmname=request.POST['fm_name']
            print(fmname)
            sname=request.POST['s_name']
            print(sname)        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            print(address)
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            print(paddress)
            phone=request.POST['phone']
            print(phone)
            ephone=request.POST['ephone']
            print(ephone)
        
            email=request.POST['email2']
            print(email)
            
            isdts=request.POST['tds']
            print(isdts)
            attach=request.FILES.get('attach')
            print(attach)
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']

            an=request.POST['an']
            print(itn)
            print(an)
           
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            print(uan)
            print(pfn)
           
            print(pran)
            
            print(age)
            
            print(bank)
            
            print(accno)
            print(ifsc)
            
            print(bname)
            
            print(branch)
            
            print(ttype)
           
           
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                            emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                            amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                            UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=company_details,login_details=log_details)
            print(payroll)
            payroll.save()
            
            history=employee_history(company=company_details,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            messages.info(request,'employee created')
            
        if log_details.user_type == 'Staff':
            company_details = StaffDetails.objects.get(login_details=log_details)
            title=request.POST['title']
            print(title)
            fname=request.POST['fname']
            print(fname)
            lname=request.POST['lname']
            print(lname)
            alias=request.POST['alias']
            print(alias)
            joindate=request.POST.get('joindate')
            print(joindate)
            salarydate=request.POST['salarydate']
            print(salarydate)

            saltype=request.POST['saltype']
            print(saltype)
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
                print(salary)
            else:
                salary=request.POST['vsalary']
                print(salary)
            
            image=request.FILES.get('file')
            print(image)
            amountperhr=request.POST['amnts']
            print(amountperhr)
            workhr=request.POST['hours'] 
            print(workhr)
            empnum=request.POST['empnum']
            print(empnum)
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details.company):
                messages.info(request,'employee number all ready exists')
                return redirect('employeeloan_create')
            designation = request.POST['designation']
            print(designation)
            location=request.POST['location']
            print(location)
            gender=request.POST['gender']
            print(gender)
            dob=request.POST['dob']
            print(dob)
            blood=request.POST['blood']
            print(blood)
            fmname=request.POST['fm_name']
            print(fmname)
            sname=request.POST['s_name']
            print(sname)        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            print(address)
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            print(paddress)
            phone=request.POST['phone']
            print(phone)
            ephone=request.POST['ephone']
            print(ephone)
        
            email=request.POST['email2']
            print(email)
            
            isdts=request.POST['tds']
            print(isdts)
            attach=request.FILES.get('attach')
            print(attach)
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']

            an=request.POST['an']
            print(itn)
            print(an)
           
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            print(uan)
            print(pfn)
           
            print(pran)
            
            print(age)
            
            print(bank)
            
            print(accno)
            print(ifsc)
            
            print(bname)
            
            print(branch)
            
            print(ttype)
           
           
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                            emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                            amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                            UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=company_details.company,login_details=log_details)
            print(payroll)
            payroll.save()
            
            history=employee_history(company=company_details.company,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            messages.info(request,'employee created')
            
        return redirect('employeeloan_create')
    

def employeeloan_details(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type =='Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        employee_loan=EmployeeLoan.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=EmployeeLoan.objects.get(id=id)
        pay= payroll_employee.objects.filter(company=dash_details)
        comment_data=employeeloan_comments.objects.filter(employee=id)
        history=Employeeloan_history.objects.filter(employeeloan=id)
        his=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details).last()
        name = his.login_details.first_name + ' ' + his.login_details.last_name 
        action = his.action
        his_date=his.Date
        repay=EmployeeLoanRepayment.objects.filter(emp=id)
        last_loan = EmployeeLoanRepayment.objects.filter(emp=id).last().balance
        loan_trans = EmployeeLoanRepayment.objects.filter(emp=id)
       
    if log_details.user_type =='Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        employee_loan=EmployeeLoan.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=EmployeeLoan.objects.get(id=id)
        pay= payroll_employee.objects.filter(company=dash_details.company)
        comment_data=employeeloan_comments.objects.filter(employee=id,company=dash_details.company)
        history=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details.company)
        his=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details.company).last()
        name = his.login_details.first_name + ' ' + his.login_details.last_name 
        action = his.action
        his_date=his.Date
        repay=EmployeeLoanRepayment.objects.filter(emp=id)
        last_loan = EmployeeLoanRepayment.objects.filter(emp=id).last().balance
        loan_trans = EmployeeLoanRepayment.objects.filter(emp=id)
       
    content = {
                'details': dash_details,
                'employee_loan':employee_loan,
                'p':p,
                'allmodules': allmodules,
                'comment':comment_data,
                'history':history,
                'log_id':log_details,
                'pay':pay,
                'his':his,
                'name':name,
                'action':action,
                'his_date':his_date,
                'last_loan':last_loan,
                'repay':repay,
                'loan_trans':loan_trans,
                'state':'0'
                
        }
  
    return render(request,'zohomodules/employe_loan/employeeloan_overview.html',content)


def add_emploan_comment(request,id):                                                               

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=employeeloan_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comments']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    empo=EmployeeLoan.objects.get(id=id)
                    com.employee=empo
                    
                    com.save()
                    return redirect('employeeloan_details',id)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=employeeloan_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comments']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            empo=EmployeeLoan.objects.get(id=id)
            com.employee=empo
                    
            com.save()
            return redirect('employeeloan_details',id)
    return redirect('employeeloan_details',id)


def delete_emploan_comment(request,ph,pr):                                                              
    empo=employeeloan_comments.objects.filter(id=ph)
    empo.delete()
    ac=EmployeeLoan.objects.get(id=pr)
    
    return redirect(employeeloan_details,ac.id)


def employeeloan_repayment_pageload(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        today = datetime.now().strftime('%Y-%m-%d')
        emp=EmployeeLoan.objects.filter(company=dash_details)
        employee=EmployeeLoan.objects.get(id=id)
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        emp=EmployeeLoan.objects.filter(company=dash_details.company)
        
        employee=EmployeeLoan.objects.get(id=id)
        today = datetime.now().strftime('%Y-%m-%d')
        bank=Banking.objects.filter(company=dash_details.company)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
       
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
                'details': dash_details,
                'emp':emp,
                'allmodules': allmodules,
                'log_id':log_details,
                'employee':employee,
                'today':today,
                'bank':bank,
                

             
        }
    return render(request,'zohomodules/employe_loan/employee_loan_repayment.html',content)   


def add_repayment(request,id):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            pamnt=float(request.POST.get('principal'))
            interest=float(request.POST.get('interest'))
            pdate=request.POST.get('date')
            pmethod=request.POST.get('payment_method')
            upi_id=request.POST.get('upi_id')
            cheque_id=request.POST.get('cheque_id')
            bank_id=request.POST.get('bnk_id')
            total=float(request.POST.get('total'))
            employ = request.POST.get('emp')
            print(employ)
            employee = payroll_employee.objects.get(id=employ)

            repay = EmployeeLoanRepayment.objects.filter(emp=id).last()
            print(repay)
            balance = repay.balance
            
            if float(pamnt) > float(balance):
                messages.info(request, 'Paid Amount is Greater Than Balance!!!')
                return redirect('repayment_view',id)
            else:
                bal = float(balance) - float(pamnt)

            rep = EmployeeLoanRepayment(employee=employee,principal_amount=pamnt,interest_amonut=interest,payment_date=pdate,payment_method=pmethod,total_payment = total,
                                cheque_id=cheque_id,upi_id=upi_id,bank_id=bank_id,balance=bal,particular='EMI PAID',emp=repay.emp,logindetails=log_details,company=dash_details)
            rep.save()

            return redirect('employeeloan_trans',id)
            
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            pamnt=float(request.POST.get('principal'))
            interest=float(request.POST.get('interest'))
            pdate=request.POST.get('date')
            pmethod=request.POST.get('payment_method')
            upi_id=request.POST.get('upi_id')
            cheque_id=request.POST.get('cheque_id')
            bank_id=request.POST.get('bnk_id')
            total=float(request.POST.get('total'))
            employ = request.POST.get('emp')
            print(employ)
            employee = payroll_employee.objects.get(id=employ)

            repay = EmployeeLoanRepayment.objects.filter(emp=id).last()
            print(repay)
            balance = repay.balance
            
            if float(pamnt) > float(balance):
                messages.info(request, 'Paid Amount is Greater Than Balance!!!')
                return redirect('repayment_view',id)
            else:
                bal = float(balance) - float(pamnt)

            rep = EmployeeLoanRepayment(employee=employee,principal_amount=pamnt,interest_amonut=interest,payment_date=pdate,payment_method=pmethod,total_payment = total,
                                cheque_id=cheque_id,upi_id=upi_id,bank_id=bank_id,balance=bal,particular='EMI PAID',emp=repay.emp,logindetails=log_details,company=dash_details.company)
            rep.save()  

            return redirect('employeeloan_trans',id)
                    
        
def add_newloan_pageload(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        today = datetime.now().strftime('%Y-%m-%d')
        emp=EmployeeLoan.objects.filter(company=dash_details)
        employee=EmployeeLoan.objects.get(id=id)
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]

        repay = EmployeeLoanRepayment.objects.filter(emp=id).last()
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        emp=EmployeeLoan.objects.filter(company=dash_details.company)
        
        employee=EmployeeLoan.objects.get(id=id)
        today = datetime.now().strftime('%Y-%m-%d')
        bank=Banking.objects.filter(company=dash_details.company)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        repay = EmployeeLoanRepayment.objects.filter(emp=id).last()
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
                'details': dash_details,
                'emp':emp,
                'allmodules': allmodules,
                'log_id':log_details,
                'employee':employee,
                'today':today,
                'bank':bank,
                'repay':repay,
               
             
        }
    return render(request,'zohomodules/employe_loan/employee_loan_newloan.html',content)                   

def add_newloan(request,id):
    if request.method == 'POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        employ = EmployeeLoan.objects.get(id=id)
        print(employ)
        em_id = employ.Employee.id
        employee = payroll_employee.objects.get(id=em_id)
    
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            principal = int(request.POST.get('remain_loan'))
            date = request.POST.get('adjdate')
            new_loan = request.POST.get('new')
            total = request.POST.get('amount')
            cheque_id = request.POST['cheque_id'] 
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
            
            rep = EmployeeLoanRepayment(employee=employee,principal_amount=principal,interest_amonut=0,payment_date=date,payment_method=payment_method,total_payment = new_loan,
                                cheque_id=cheque_id,upi_id=upi_id,bank_id=bnk_id,balance=total,particular='ADDITIONAL LOAN ISSUED',emp=employ,logindetails=log_details,company=dash_details)
            rep.save()
            employ.balance = total
            print(employ.LoanAmount)
            res = employ.LoanAmount + int(new_loan)
            employ.LoanAmount = res
            print(total)
            print(principal)
            print(res)
            employ.save()
            
            return redirect('employeeloan_trans',id) 

        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            principal = int(request.POST.get('remain_loan'))
            date = request.POST.get('adjdate')
            new_loan = request.POST.get('new')
            total = request.POST.get('amount')
            cheque_id = request.POST['cheque_id'] 
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
            
            
            rep = EmployeeLoanRepayment(employee=employee,principal_amount=principal,interest_amonut=0,payment_date=date,payment_method=payment_method,total_payment = new_loan,
                                cheque_id=cheque_id,upi_id=upi_id,bank_id=bnk_id,balance=total,particular='ADDITIONAL LOAN ISSUED',emp=employ,logindetails=log_details,company=dash_details.company)
            rep.save()
            employ.balance = total
            print(employ.LoanAmount)
            res = employ.LoanAmount + int(new_loan)
            employ.LoanAmount = res
            print(total)
            print(principal)
            print(res)
            employ.save()

            
            return redirect('employeeloan_trans',id)

                    
def delete_repayment(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        loan_id = repay.emp.id
        id_to_delete = repay.id
        repay.delete()

        entry_list = []
        gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=id_to_delete, emp=loan_id)
        lt_entrie = EmployeeLoanRepayment.objects.filter(id__lt=id_to_delete, emp=loan_id).last()

        if gt_entries:
            entry_list.append(lt_entrie)
            for g in gt_entries:
                entry_list.append(g)

            for i in range(1,len(entry_list)):
                if entry_list[i].particular == 'ADDITIONAL LOAN ISSUED':
                    entry_list[i].balance = float(entry_list[i-1].balance) + float(entry_list[i].principal_amount)
                else:
                    entry_list[i].balance = float(entry_list[i-1].balance) - float(entry_list[i].payment_made)
                entry_list[i].save()

        return redirect('employeeloan_trans',loan_id)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        loan_id = repay.emp.id
        id_to_delete = repay.id
        repay.delete()

        entry_list = []
        gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=id_to_delete, emp=loan_id)
        lt_entrie = EmployeeLoanRepayment.objects.filter(id__lt=id_to_delete, emp=loan_id).last()

        if gt_entries:
            entry_list.append(lt_entrie)
            for g in gt_entries:
                entry_list.append(g)

            for i in range(1,len(entry_list)):
                if entry_list[i].particular == 'ADDITIONAL LOAN ISSUED':
                    entry_list[i].balance = float(entry_list[i-1].balance) + float(entry_list[i].principal_amount)
                else:
                    entry_list[i].balance = float(entry_list[i-1].balance) - float(entry_list[i].principal_amount)
                entry_list[i].save()

        return redirect('employeeloan_trans',loan_id)
    

def edit_loanrepayment(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        
       
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        bank=Banking.objects.filter(company=dash_details.company)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    context= {'repay':repay,'details':dash_details,'log_details':log_details,'allmodules':allmodules,'bank':bank}
    return render(request,'zohomodules/employe_loan/edit_repaymentloan.html',context)

def save_edit_loanrepayment(request,id):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            repay = EmployeeLoanRepayment.objects.get(id=id)
            repay.principal_amount = request.POST.get('principal')
            repay.interest_amonut=request.POST.get('interest')
            repay.payment_date=request.POST.get('date')
            repay.total_payment=request.POST.get('total')
            repay.payment_method=request.POST.get('payment_method')
            repay.upi_id=request.POST.get('upi_id')
            repay.cheque_id=request.POST.get('cheque_id')
            repay.bank_id=request.POST.get('bnk_id')
            gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=repay.id, emp=repay.emp.id)
            lt_entrie = EmployeeLoanRepayment.objects.filter(id__lt=repay.id, emp=repay.emp.id).last()

            bal = float(lt_entrie.balance)- float(request.POST.get('principal'))
            repay.balance = bal
            repay.save()

            for entry in gt_entries:
                if entry.particular == 'ADDITIONAL LOAN ISSUED':
                    entry.balance = bal + float(entry.principal_amount)
                else:
                    entry.balance = bal - float(entry.principal_amount)
                entry.save()
                bal = entry.balance
            return redirect('employeeloan_trans',repay.emp.id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            repay = EmployeeLoanRepayment.objects.get(id=id)
            repay.principal_amount = request.POST.get('principal')
            repay.interest_amonut=request.POST.get('interest')
            repay.payment_date=request.POST.get('date')
            repay.total_payment=request.POST.get('total')
            repay.payment_method=request.POST.get('payment_method')
            repay.upi_id=request.POST.get('upi_id')
            repay.cheque_id=request.POST.get('cheque_id')
            repay.bank_id=request.POST.get('bnk_id')

            gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=repay.id, emp=repay.emp.id)
            lt_entrie = EmployeeLoanRepayment.objects.filter(id__lt=repay.id, emp=repay.emp.id).last()

            bal = float(lt_entrie.balance)- float(request.POST.get('principal'))
            repay.balance = bal
            repay.save()

            for entry in gt_entries:
                if entry.particular == 'ADDITIONAL LOAN ISSUED':
                    entry.balance = bal + float(entry.principal_amount)
                else:
                    entry.balance = bal - float(entry.principal_amount)
                entry.save()
                bal = entry.balance
            return redirect('employeeloan_trans',repay.emp.id)


def edit_additionalloan_pageloage(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        current_bal = float(repay.balance) - float(repay.principal_amount)
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        repay = EmployeeLoanRepayment.objects.get(id=id)
        current_bal = float(repay.balance) - float(repay.principal_amount)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
    context= {'repay':repay,'details':dash_details,'log_details':log_details,'current_bal':current_bal,'allmodules':allmodules,'bank':bank}
    return render(request,'zohomodules/employe_loan/edit_additionalloan.html',context)


def save_edit_additionalloan(request,id):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            repay = EmployeeLoanRepayment.objects.get(id=id)
            repay.principal_amount = request.POST.get('new')
            repay.payment_date=request.POST.get('adjdate')
            repay.balance=request.POST.get('amount')
            repay.payment_method=request.POST.get('payment_method')
            repay.upi_id=request.POST.get('upi_id')
            repay.cheque_id=request.POST.get('cheque_id')
            repay.bank_id=request.POST.get('bnk_id')
            repay.total_payment=float(request.POST.get('remain_loan'))
            repay.save()

            gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=repay.id, emp=repay.emp.id)
            bal = float(repay.balance)
            for entry in gt_entries:
                if entry.particular == 'ADDITIONAL LOAN ISSUED':
                    entry.balance = bal + float(entry.principal_amount)
                else:
                    entry.balance = bal - float(entry.principal_amount)
                entry.save()
                bal = entry.balance
            return redirect('employeeloan_trans',repay.emp.id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            repay = EmployeeLoanRepayment.objects.get(id=id)
            repay.principal_amount = request.POST.get('new')
            repay.payment_date=request.POST.get('adjdate')
            repay.balance=request.POST.get('amount')
            repay.payment_method=request.POST.get('payment_method')
            repay.upi_id=request.POST.get('upi_id')
            repay.cheque_id=request.POST.get('cheque_id')
            repay.bank_id=request.POST.get('bnk_id')
            repay.total_payment=float(request.POST.get('remain_loan'))
            repay.save()

            gt_entries = EmployeeLoanRepayment.objects.filter(id__gt=repay.id, emp=repay.emp.id)
            bal = float(repay.balance)
            for entry in gt_entries:
                if entry.particular == 'ADDITIONAL LOAN ISSUED':
                    entry.balance = bal + float(entry.principal_amount)
                else:
                    entry.balance = bal - float(entry.principal_amount)
                entry.save()
                bal = entry.balance
            return redirect('employeeloan_trans',repay.emp.id)


def shareloanemail(request,pk):
  
    if request.method == 'POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

            emails_string = request.POST['email']
            fdate = request.POST['fdate']
            edate = request.POST['ldate']
            print(fdate)
            print(edate)

            print(emails_string)

        
            emails_list = [email.strip() for email in emails_string.split(',')]
                
            p=EmployeeLoan.objects.get(id=pk)
            loan_trans = EmployeeLoanRepayment.objects.filter(emp=pk)
            if fdate and edate:
                loan_trans = EmployeeLoanRepayment.objects.filter(payment_date__gte=fdate, payment_date__lte=edate)
                            
            context = {'p':p,'loan_trans':loan_trans,'details':dash_details,'log_details':log_details}
            template_path = 'zohomodules/employe_loan/emploan_share.html'
            template = get_template(template_path)
            html  = template.render(context)
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
            pdf = result.getvalue()
           
            filename = f'{p.Employee.first_name}details - {p.id}.pdf'
            
            subject = f"{p.Employee.first_name}{p.Employee.last_name}  - {p.id}-details"
            email = EmailMultiAlternatives(subject, f"Hi,\nPlease find the attached employeeloan details - File-{p.Employee.first_name}{p.Employee.last_name} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
            email.attach(filename, pdf, "application/pdf")
            email.send(fail_silently=False)
            
            return redirect('employeeloan_details',pk)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            emails_string = request.POST['email']
            print(emails_string)

        
            emails_list = [email.strip() for email in emails_string.split(',')]
                
            p=EmployeeLoan.objects.get(id=pk)
            loan_trans = EmployeeLoanRepayment.objects.filter(emp=pk)
                            
            context = {'p':p,'loan_trans':loan_trans,'details':dash_details,'log_details':log_details}
            template_path = 'zohomodules/employe_loan/emploan_share.html'
            template = get_template(template_path)
            html  = template.render(context)
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
            pdf = result.getvalue()
           
            filename = f'{p.Employee.first_name}details - {p.id}.pdf'
            
            subject = f"{p.Employee.first_name}{p.Employee.last_name}  - {p.id}-details"
            email = EmailMultiAlternatives(subject, f"Hi,\nPlease find the attached employeeloan details - File-{p.Employee.first_name}{p.Employee.last_name} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
            email.attach(filename, pdf, "application/pdf")
            email.send(fail_silently=False)
            
            return redirect('employeeloan_details',pk)

def active_loan(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        l = get_object_or_404(EmployeeLoan, id=id)

    # Activate the bank account
        l.active = True
        l.save()
        
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        l = get_object_or_404(EmployeeLoan, id=id)

    # Activate the bank account
        l.active = True
        l.save()

    return redirect('employeeloan_details',id)

def inactive_loan(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        l = get_object_or_404(EmployeeLoan, id=id)

    # Activate the bank account
        l.active = False
        l.save()
     
        
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        l = get_object_or_404(EmployeeLoan, id=id)

    # Activate the bank account
        l.active = False
        l.save()

    return redirect('employeeloan_details',id) 


def deleteloan(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        employee = EmployeeLoan.objects.get(id=id)
        employee.delete()
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        employee = EmployeeLoan.objects.get(id=id)
        employee.delete()
    return redirect('employee_listpage')


def create_loan_duration(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        days=request.POST['days']
        durs=request.POST['durs']
        LoanDuration.objects.create(day=days,duration=durs,logindetails=log_details,company=dash_details)
       
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        days=request.POST['days']
        durs=request.POST['durs']
        LoanDuration.objects.create(day=days,duration=durs,logindetails=log_details,company=dash_details.company) 
    return JsonResponse({'success': 'LoanDuration Saved'})

def loan_duration(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
    else:
        return redirect('/')

    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(
            login_details=log_details,
            superadmin_approval=1,
            Distributor_approval=1
        )
        data = {}
        dur_objects = LoanDuration.objects.filter(company=dash_details)

        for dur in dur_objects:
            duration = f"{dur.day} {dur.duration}"
        data[dur.id] = [duration, f"{duration}"]

        return JsonResponse(data)


    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details) 
        dur = LoanDuration.objects.filter(company=dash_details.company)
        data = {}
        dur_objects = LoanDuration.objects.filter(company=dash_details.company)

        for dur in dur_objects:
            duration = f"{dur.day} {dur.duration}"
        data[dur.id] = [duration, f"{duration}"]

        return JsonResponse(data)


def edit_loan(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    blood=Bloodgroup.objects.all()
    log_details= LoginDetails.objects.get(id=log_id)
  
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        pay = payroll_employee.objects.filter(company=dash_details,status='active')
        emp= EmployeeLoan.objects.filter(company=dash_details)
        p=EmployeeLoan.objects.get(id=id)
        dur =LoanDuration.objects.filter(company=dash_details)
        bank=Banking.objects.filter(company=dash_details)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        loan_trans=EmployeeLoanRepayment.objects.get(emp=p.id,particular='LOAN ISSUED',company=dash_details)
        
        
       


    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        pay = payroll_employee.objects.filter(company=dash_details.company,status='active')
        emp= EmployeeLoan.objects.filter(company=dash_details.company)
        p=EmployeeLoan.objects.get(id=id)
        dur =LoanDuration.objects.filter(company=dash_details.company)
        bank=Banking.objects.filter(company=dash_details.company)
        for i in bank:
            i.last_digit = str(i.bnk_acno)[-4:]
        loan_trans=EmployeeLoanRepayment.objects.get(emp=p.id,particular='LOAN ISSUED',company=dash_details.company)
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'log_id':log_details,
            'pay':pay,
            'dur':dur,
            'emp':emp,
            'p':p,
            'loan_trans':loan_trans,
            'blood':blood,
            'bank':bank,
            

            
    }

    return render(request,'zohomodules/employe_loan/edit_employeeloan.html',content)


def update_Employeeloan(request,id):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type =='Company':
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            employ = EmployeeLoan.objects.get(id=id, company=company_details)

            Loan_Amount = request.POST.get('Loan_Amount')
            loandate = request.POST.get('loandate')
            experydate = request.POST.get('experydate')
            cuttingPercentage = request.POST.get('cuttingPercentage')
            cuttinamount = request.POST.get('Cutingamount')
            print(cuttingPercentage)
            Note = request.POST.get('Note')
            loan_duration= request.POST['loan_duration']
            cheque_id = request.POST['cheque_id'] 
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
            try:
                file = request.FILES['file']
            except:
                file = ''

            employ.Loandate = loandate
            employ.ExperyDate = experydate
            employ.note = Note
            employ.cheque_number = cheque_id
            employ.upi_id = upi_id
            employ.bank_acc_number=bnk_id
            employ.payment_method = payment_method
            employ.duration = loan_duration

            if int(cuttingPercentage) == 0 and int(cuttinamount) != 0:
                employ.MonthlyCut_Amount = cuttinamount
                employ.MonthlyCut_percentage =0
                employ.Monthly_payment_type = 'No'
            else:
                employ.MonthlyCut_percentage = cuttingPercentage
                employ.MonthlyCut_Amount = (int(cuttingPercentage) / 100) * int(Loan_Amount)
                employ.Monthly_payment_type = 'Yes'
            employ.LoanAmount = int(Loan_Amount)
            employ.balance = int(Loan_Amount)

            employ.save()

            repay = EmployeeLoanRepayment.objects.filter(emp=employ.id)

            bal = float(employ.LoanAmount)
            for r in repay:
                if r.particular == 'LOAN ISSUED':
                    r.principal_amount = bal
                    r.total_payment = bal
                    r.payment_date = loandate
                    r.payment_method = payment_method

                    r.cheque_id = cheque_id
                    r.upi_id = upi_id
                    r.bank_id = bnk_id
                    r.balance = bal
                    r.save()
                elif r.particular == 'ADDITIONAL LOAN ISSUED':
                    r.balance = bal + float(r.principal_amount)
                    bal = float(r.balance)
                    r.save()
                else :
                    r.balance = bal - float(r.principal_amount)
                    bal = float(r.balance)
                    r.save()
            
            history=Employeeloan_history(company=company_details,login_details=log_details, employeeloan=employ, action='EDITED')
            history.save()

            return redirect('employeeloan_details',id)
        if log_details.user_type == "Staff":
            company_details = StaffDetails.objects.get(login_details=log_details)   
            employ = EmployeeLoan.objects.get(id=id, company=company_details.company)
            
            Loan_Amount = request.POST.get('Loan_Amount')
            loandate = request.POST.get('loandate')
            experydate = request.POST.get('experydate')
            cuttingPercentage = request.POST.get('cuttingPercentage')
            cuttinamount = request.POST.get('Cutingamount')
            print(cuttingPercentage)
            Note = request.POST.get('Note')
            loan_duration= request.POST['loan_duration']
            cheque_id = request.POST['cheque_id'] 
            upi_id = request.POST['upi_id'] 
            bnk_id = request.POST['bnk_id'] 
            payment_method = request.POST['payment_method']
            try:
                file = request.FILES['file']
            except:
                file = ''

            employ.Loandate = loandate
            employ.ExperyDate = experydate
            employ.note = Note
            employ.cheque_number = cheque_id
            employ.upi_id = upi_id
            employ.bank_acc_number=bnk_id
            employ.payment_method = payment_method
            employ.duration = loan_duration

            if int(cuttingPercentage) == 0 and int(cuttinamount) != 0:
                employ.MonthlyCut_Amount = cuttinamount
                employ.MonthlyCut_percentage =0
                employ.Monthly_payment_type = 'No'
            else:
                employ.MonthlyCut_percentage = cuttingPercentage
                employ.MonthlyCut_Amount = (int(cuttingPercentage) / 100) * int(Loan_Amount)
                employ.Monthly_payment_type = 'Yes'
            employ.LoanAmount = int(Loan_Amount)
            employ.balance = int(Loan_Amount)

            employ.save()

            repay = EmployeeLoanRepayment.objects.filter(emp=employ.id)

            bal = float(employ.LoanAmount)
            for r in repay:
                if r.particular == 'LOAN ISSUED':
                    r.principal_amount = bal
                    r.total_payment = bal
                    r.payment_date = loandate
                    r.payment_method = payment_method

                    r.cheque_id = cheque_id
                    r.upi_id = upi_id
                    r.bank_id = bnk_id
                    r.balance = bal
                    r.save()
                elif r.particular == 'ADDITIONAL LOAN ISSUED':
                    r.balance = bal + float(r.principal_amount)
                    bal = float(r.balance)
                    r.save()
                else :
                    r.balance = bal - float(r.principal_amount)
                    bal = float(r.balance)
                    r.save()
                
            history=Employeeloan_history(company=company_details.company,login_details=log_details, employeeloan=employ, action='EDITED')
            history.save()

            return redirect('employeeloan_details',id) 

    
def bankdata(request):
    if 'login_id' in request.session:
            log_id = request.session['login_id']
    if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
        
    if log_details.user_type =='Company':
        company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        bank_id = request.GET.get('id')
        term = Banking.objects.get(id=bank_id,company=company_details)
        data = {'bank': term.bnk_acno}
        return JsonResponse(data)  
    
    if log_details.user_type == "Staff":
        company_details = StaffDetails.objects.get(login_details=log_details)
        bank_id = request.GET.get('id')
        term = Banking.objects.get(id=bank_id,company=company_details.company)
        data = {'bank': term.bnk_acno}
        return JsonResponse(data)


def bankdata1(request):
    if 'login_id' in request.session:
            log_id = request.session['login_id']
    if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
        
    if log_details.user_type =='Company':
        company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        bank_id = request.GET.get('id')
        term = Banking.objects.get(bnk_name=bank_id,company=company_details)
        data = {'bank': term.bnk_acno}
        return JsonResponse(data)  
    
    if log_details.user_type == "Staff":
        company_details = StaffDetails.objects.get(login_details=log_details)
        bank_id = request.GET.get('id')
        term = Banking.objects.get(bnk_name=bank_id,company=company_details.company)
        data = {'bank': term.bnk_acno}
        return JsonResponse(data)


def addloan_file(request,pk):
    if request.method == 'POST':
        data=request.FILES.get('file')
        emp=EmployeeLoan.objects.get(id=pk)
        if emp.file:
            try:
                                # Check if the file exists before removing it
                if os.path.exists(emp.file.path):
                    os.remove(emp.file.path)
            except Exception as e:
                messages.error(request,'file upload error')
                return redirect('employeeloan_details',pk)

                            # Assign the new file to payroll.image
            emp.file = data
            emp.save()
            messages.info(request,'file uploaded')
            return redirect('employeeloan_details',pk)
        else:
            emp.file = data
            emp.save()
        messages.info(request,'fil uploaded')
        return redirect('employeeloan_details',pk)


def import_employee_loan_details(request):
    if request.method == 'POST' and 'file' in request.FILES:
        if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
            log_details = LoginDetails.objects.get(id=log_id)

            if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                    
            elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
            excel_file = request.FILES['file']
            print(excel_file)
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                empnumber,loanamount,loandate,duration,Expirydate,paymentmethod,chequeno,upi_id,bankid,Monthlypayment_type,MonthlyCutpercentage,MonthlyCutAmount,note = row
                
                    
                employees=payroll_employee.objects.filter(emp_number=empnumber)
                for employee in employees:
                    existing_loan = EmployeeLoan.objects.filter(Employee=employee).first()
                    
                    if existing_loan:
                        # Display an error message
                        messages.error(request, f"Loan for employee {empnumber} on {loandate} already exists.")
                    else:
                        # Create a new employee loan
                        emp = EmployeeLoan.objects.create(
                            Employee=employee,
                            Loandate=loandate,
                            LoanAmount=loanamount,
                            duration=duration,
                            Expiry_date=Expirydate,
                            payment_method=paymentmethod,
                            cheque_number=chequeno,
                            upi_id=upi_id,
                            bank_acc_number=bankid,
                            Monthly_payment_type=Monthlypayment_type,
                            MonthlyCut_percentage=MonthlyCutpercentage,
                            MonthlyCut_Amount=MonthlyCutAmount,
                            note=note,
                            balance=loanamount,
                            company=company,
                            login_details=log_details
                        )

                        Employeeloan_history.objects.create(
                            login_details=log_details,
                            employeeloan=emp,
                            company=company,
                            action='CREATED'
                        )
                        EmployeeLoanRepayment.objects.create(
                            emp=emp,
                            principal_amount=loanamount,
                            interest_amonut=0,
                            total_payment=0,
                            payment_date=loandate,
                            balance=loanamount,
                            payment_method=paymentmethod,
                            upi_id=upi_id,
                            cheque_id=chequeno,
                            bank_id=bankid,
                            particular='LOAN ISSUED',
                            company=company,
                            logindetails=log_details,
                            employee=employee
                        )
                        
                        # Display a success message
                        messages.success(request, "Employee loan successfully created.")
            
            return redirect('employee_listpage')
        
        messages.error(request, "No file uploaded or invalid request method")
        return redirect('employee_listpage')

#-------------------------employeeloan views end -------------#

#--------------Customer-----------------#
#-------------------Arya E.R----------------------#

def customer(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')

       
        return render(request,'zohomodules/customer/create_customer.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details,'price_lists':price_lists}) 
    else:
        return redirect('/')  

def view_customer_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        data=Customer.objects.filter(company=comp_details)

        

        return render(request,'zohomodules/customer/customer_list.html',{'details':dash_details,'allmodules': allmodules,'data':data,'log_details':log_details}) 


    else:
        return redirect('/')
    

def add_customer(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        

       
        if request.method=="POST":
            customer_data=Customer()
            customer_data.login_details=log_details
            customer_data.company=comp_details
            customer_data.customer_type = request.POST.get('type')

            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            customer_data.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type=op_type
            else:
                customer_data.opening_balance_type='Opening Balance not selected'
    
            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            # customer_data.price_list=request.POST['plst']
            plst=request.POST.get('plst')
            if plst!=0:
                 customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'




            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'
    



           
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
            rdata=Customer_remarks_table()
            rdata.remarks=request.POST['remark']
            rdata.company=comp_details
            rdata.customer=vdata
            rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = CustomerContactPersons.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor)
                
        
            messages.success(request, 'Customer created successfully!')   

            return redirect('view_customer_list')
        
        else:
            messages.error(request, 'Some error occurred !')   

            return redirect('view_customer_list')


def check_customer_phonenumber_exist(request):
    if request.method == 'GET':
       mPhone = request.GET.get('m_Phone', None)

       if mPhone:
          
            exists = Customer.objects.filter(
                    customer_mobile=mPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def check_customer_work_phone_exist(request):
    if request.method == 'GET':
       wPhone = request.GET.get('w_Phone', None)

       if wPhone:
          
            exists = Customer.objects.filter(
                    customer_phone=wPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})   

def check_customer_email_exist(request):
    if request.method == 'GET':
       vendoremail = request.GET.get('vendor_email', None)

       if vendoremail:
          
            exists = Customer.objects.filter(
                    customer_email=vendoremail
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def customer_payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            normalized_data = terms.replace(" ", "")
            pay_tm = add_space_before_first_digit(normalized_data)
            ptr = Company_Payment_Term(term_name=pay_tm, days=day, company=dash_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=comp_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)   
            

def check_customer_term_exist(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

    if request.method == 'GET':
       term_name = request.GET.get('term_name', None)
       if term_name:
            normalized_data = term_name.replace(" ", "")
            term_name_processed = add_space_before_first_digit(normalized_data)
            exists = Company_Payment_Term.objects.filter(
                    term_name=term_name_processed,
                    company=comp_details
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})    

def customer_check_pan(request):
    if request.method == 'POST':
        panNumber = request.POST.get('panNumber')
        pan_exists = Customer.objects.filter(PAN_number=panNumber).exists()

        if pan_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'})  

def customer_check_gst(request):
    if request.method == 'POST':
        gstNumber = request.POST.get('gstNumber')
        gst_exists = Customer.objects.filter(GST_number=gstNumber).exists()
       
        if gst_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'}) 

def sort_customer_by_name(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
  
        data=Customer.objects.filter(login_details=log_details).order_by('first_name')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
            return redirect('/')    

def sort_customer_by_amount(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        data=Customer.objects.filter(login_details=log_details).order_by('opening_balance')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/')   


def view_customer_active(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        data=Customer.objects.filter(login_details=log_details,customer_status='Active').order_by('-id')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/') 

    
    
def view_customer_inactive(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        data=Customer.objects.filter(login_details=log_details,customer_status='Inactive').order_by('-id')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/') 


def import_customer_excel(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        if request.method == 'POST' :
       
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                            
                    vendorsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    comp_term=vendorsheet[18]
                    if comp_term:
                        normalized_data = comp_term.replace(" ", "")

                        pay_tm = add_space_before_first_digit(normalized_data)
                    else:
                        cpt =Company_Payment_Term.objects.filter(company=comp_details).first()
                        pay_tm = cpt.term_name
   
                    try:
                        com_term_obj=Company_Payment_Term.objects.get(company=comp_details,term_name=pay_tm)
                    except Company_Payment_Term.DoesNotExist:
                        com_term_obj= None
                    
                    opn_blc_str = vendorsheet[17]  # Assuming vendorsheet[15] is a string representing a decimal
                    if opn_blc_str:

                        opn_blc = float(opn_blc_str)
                    else:
                        opn_blc = 0.00    
                    

                    Vendor_object=Customer(customer_type=vendorsheet[0],title=vendorsheet[1],first_name=vendorsheet[2],last_name=vendorsheet[3],company_name=vendorsheet[4],customer_email=vendorsheet[5],customer_phone=vendorsheet[6],customer_mobile=vendorsheet[7],skype=vendorsheet[8],designation=vendorsheet[9],department=vendorsheet[10],website=vendorsheet[11],
                                         GST_treatement=vendorsheet[12],place_of_supply=vendorsheet[13],tax_preference=vendorsheet[14],currency=vendorsheet[15],opening_balance_type=vendorsheet[16],
                                         opening_balance=opn_blc,company_payment_terms=com_term_obj,billing_attention=vendorsheet[19],billing_country=vendorsheet[20],billing_address=vendorsheet[21],
                                         billing_city=vendorsheet[22],billing_state=vendorsheet[23],billing_pincode=vendorsheet[24],
                                         billing_mobile=vendorsheet[25],billing_fax=vendorsheet[26],shipping_attention=vendorsheet[27],shipping_country=vendorsheet[28],shipping_address=vendorsheet[29],shipping_city=vendorsheet[30],
                                         shipping_state=vendorsheet[31],shipping_pincode=vendorsheet[32],
                                         shipping_mobile=vendorsheet[33], shipping_fax=vendorsheet[34], remarks=vendorsheet[35],current_balance=opn_blc,customer_status="Active",company=comp_details,login_details=log_details)
                    Vendor_object.save()

    
                   
                messages.warning(request,'file imported')
                return redirect('view_customer_list')    

    
            messages.error(request,'File upload Failed!11')
            return redirect('view_customer_list')
        else:
            messages.error(request,'File upload Failed!11')
            return redirect('view_customer_list') 

def view_customer_details(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        vendor_obj=Customer.objects.get(id=pk)

        # Getting all vendor to disply on the left side of vendor_detailsnew page
        vendor_objs=Customer.objects.filter(company=comp_details)

        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
    content = {
                'details': dash_details,
               
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content)    

def sort_customer(request,selectId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        vendor_obj = Customer.objects.get(id=pk)
        vendor_objs = Customer.objects.filter(company=comp_details)

        if selectId == 0:
            vendor_objs=Customer.objects.filter(company=comp_details)
        if selectId == 1:
            vendor_objs=Customer.objects.filter(company=comp_details).order_by('first_name')
        if selectId == 2:
            vendor_objs=Customer.objects.filter(company=comp_details).order_by('opening_balance')
           
        
        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content) 

def customer_status_change(request,statusId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
    

        vendor_obj = Customer.objects.get(id=pk)
        vendor_objs = Customer.objects.filter(company=comp_details)

        if statusId == 0:
            vendor_objs=Customer.objects.filter(company=comp_details)
        if statusId == 1:
            vendor_objs=Customer.objects.filter(company=comp_details,customer_status='Active').order_by('-id')
        if statusId == 2:
            vendor_objs=Customer.objects.filter(company=comp_details,customer_status='Inactive').order_by('-id')
           
        
        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content)       

def delete_customers(request, pk):
    try:
        vendor_obj = Customer.objects.get(id=pk)

        vendor_obj.delete()
        return redirect('view_customer_list')  
    except Customer.DoesNotExist:
        return HttpResponseNotFound("Customer not found.")  

def customer_status(request,pk):
    vendor_obj = Customer.objects.get(id=pk)
    if vendor_obj.customer_status == 'Active':
        vendor_obj.customer_status ='Inactive'
    elif vendor_obj.customer_status == 'Inactive':
        vendor_obj.customer_status ='Active'
    vendor_obj.save()
    return redirect('view_customer_details',pk)         

def customer_add_comment(request,pk):
   if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        if request.method =='POST':
            comment_data=request.POST['comments']
       
            vendor_id= Customer.objects.get(id=pk) 
            vendor_obj=Customer_comments_table()
            vendor_obj.comment=comment_data
            vendor_obj.customer=vendor_id
            vendor_obj.company=comp_details
            vendor_obj.login_details= LoginDetails.objects.get(id=log_id)

            vendor_obj.save()
            return redirect('view_customer_details',pk)
   return redirect('view_customer_details',pk) 


def customer_delete_comment(request, pk):
    try:
        vendor_comment =Customer_comments_table.objects.get(id=pk)
        vendor_id=vendor_comment.customer.id
        vendor_comment.delete()
        return redirect('view_customer_details',vendor_id)  
    except Customer_comments_table.DoesNotExist:
        return HttpResponseNotFound("comments not found.") 

def add_customer_file(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        if request.method == 'POST':
            data=request.FILES.getlist('file')
            try:
                for doc in data:

                    vendor_obj=Customer_doc_upload_table()
                    
                    vendor_obj.document = doc
                    vendor_obj.login_details = log_details
                    vendor_obj.company = comp_details
                    vendor_obj.customer = Customer.objects.get(id=pk)
                    vendor_obj.save()
                
                messages.success(request,'File uploaded')
                return redirect('view_customer_details',pk) 
            except Customer_doc_upload_table.DoesNotExist:
                return redirect('view_customer_details',pk) 

def customer_shareemail(request,pk):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
    
        vendor_obj=Customer.objects.get(id=pk)

        context = {'vendor_obj':vendor_obj,'details':dash_details}
        if request.method == 'POST':
            try:
                emails_string = request.POST['email_ids']

                        # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                                                                                          
                template_path = 'zohomodules/customer/customermailoverview.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                subject = f"Transaction Details"
                email = f"Hi,\nPlease find the attached transaction details {vendor_obj.first_name} {vendor_obj.last_name}.\n"
                email_from = settings.EMAIL_HOST_USER

        
                msg = EmailMultiAlternatives(subject, email, email_from, emails_list)
                msg.attach(f'{vendor_obj.first_name}_{vendor_obj.last_name}_Transactions.pdf', pdf, "application/pdf")
                
                # Send the email
                msg.send()

                messages.success(request, 'Transaction has been shared via email successfully..!')
                return redirect('view_customer_details',pk)

            except Exception as e:
                print(f"Error sending email: {e}")
                messages.error(request, 'An error occurred while sending the email. Please try again later.')
                return redirect('view_customer_details',pk)  


def Customer_edit(request,pk):
   
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        customer_obj=Customer.objects.get(id=pk)

        customer_contact_obj=CustomerContactPersons.objects.filter(customer=customer_obj)  
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'customer_obj':customer_obj,
                'log_details':log_details,
                'customer_contact_obj':customer_contact_obj,
                'comp_payment_terms':comp_payment_terms,
                'price_lists': price_lists,
        }
    

        return render(request,'zohomodules/customer/edit_customer.html',content)
     else:
            return redirect('/')

def do_customer_edit(request,pk):
         
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method=="POST":
            customer_data=Customer.objects.get(id=pk)
            customer_data.login_details=log_details
            customer_data.company=comp_details
            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            
            cob=float(request.POST['opening_bal'])
            oc=float(customer_data.current_balance) 
            ob=float(customer_data.opening_balance) 


            if cob > ob:
                diffadd=cob-ob
                oc=oc + diffadd
                customer_data.current_balance=oc
                customer_data.opening_balance=cob
            elif cob < ob:
                diffadd=ob-cob
                oc=oc-diffadd
                customer_data.current_balance=oc
                customer_data.opening_balance=cob

            else:
                customer_data.current_balance=request.POST['opening_bal']   
       
            

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type=op_type
            else:
                customer_data.opening_balance_type='Opening Balance not selected'
            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            plst=request.POST.get('plst')
            if plst!=0:

                 customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'


            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'
            
           
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.save()


              # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Edited'
            vendor_history_obj.save()
    # .......................................................adding to remaks table.....................
            vdata=Customer.objects.get(id=customer_data.id)
            try:

                rdata=Customer_remarks_table.objects.get(customer=vdata)
                rdata.remarks=request.POST['remark']
                rdata.company=comp_details
                rdata.customer=vdata
                rdata.save()
            except Customer_remarks_table.DoesNotExist:
                remarks_obj= Customer_remarks_table()   
                remarks_obj.remarks=request.POST['remark']
                remarks_obj.company=comp_details
                remarks_obj.customer=vdata
                remarks_obj.save()


    #  ...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            person = request.POST.getlist('contact_person_id[]')
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department)==len(person):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department,person)
                    mapped2=list(mapped2)
                    for ele in mapped2:
                       
                        existing_instance = CustomerContactPersons.objects.filter(id=ele[9], company=comp_details, customer=vendor).first()
                        if existing_instance:
                            # Update the existing instance
                            existing_instance.title = ele[0]
                            existing_instance.first_name = ele[1]
                            existing_instance.last_name = ele[2]
                            existing_instance.email = ele[3]
                            existing_instance.work_phone  = ele[4]
                            existing_instance.mobile = ele[5]
                            existing_instance.skype = ele[6]
                            existing_instance.designation = ele[7]
                            existing_instance.department = ele[8]

                            # Update other fields

                            existing_instance.save()
                        else:
                            # Create a new instance
                            new_instance = CustomerContactPersons.objects.create(
                                title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor
                            )
            return redirect('view_customer_details',pk)  

#------------------------------------End----------------------------------------------#

#### Kesia  ####
def loan_listing(request):
  if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            
            
            loan_details = loan_account.objects.filter(company=company)

            # Calculate balance for each loan account
            for loan in loan_details:
                total_emis_paid = LoanRepayemnt.objects.filter(loan=loan, type='EMI paid').aggregate(total=Sum('principal_amount'))['total'] or 0
                total_additional_loan = LoanRepayemnt.objects.filter(loan=loan, type='Additional Loan').aggregate(total=Sum('principal_amount'))['total'] or 0
                loan.balance = loan.loan_amount - total_emis_paid + total_additional_loan
            
            context = {
                'details': dash_details,
                'allmodules': allmodules,
                'loan_details': loan_details,
                'log_details':log_details
                
            }
  return render(request,'zohomodules/loan_account/loan_listing.html',context)

def get_account_number(request, account_id):
    try:
        bank_account = BankAccount.objects.get(id=account_id,)
        account_number = bank_account.account_number
        return JsonResponse({'account_number': account_number})
    except BankAccount.DoesNotExist:
        return JsonResponse({'error': 'Bank account not found'}, status=404)
    
def full_account_number(request, bank_id):
    try:
        print('bank')
        # bank_id = request.GET.get('bank_id')
        # print(bank_id)
        # acc = Banking.objects.get(bnk_name=bank_id)
        acc = Banking.objects.get(pk=bank_id)
        data = {'bank':acc.bnk_acno}
        print(data)
        return JsonResponse(data)
    except Banking.DoesNotExist:
        return JsonResponse({'error': 'Banking record not found'}, status=404)

def add_loan(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                # Fetch company details
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                # Fetch staff details
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            banks = Banking.objects.values('id','bnk_name','bnk_acno','bnk_ifsc','status').filter(company=company)
            today_date=date.today()
            loaned_bank_account_ids = loan_account.objects.values_list('bank_holder_id', flat=True)
            context = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'bank_holder': BankAccount.objects.filter(company=company).exclude(id__in=loaned_bank_account_ids),
                    'loan_details': loan_account.objects.filter(company=company),
                    'banks': banks,
                    'today_date':today_date,
                    'log_details':log_details
                }
            

            if request.method == 'POST':
                # account_name = request.POST.get('customer_name')
                # customer_name = BankAccount.objects.get(id=account_name)
                account_name = request.POST.get('account_name')
                account_number = request.POST.get('account_number')
                loan_amount = request.POST.get('loan_amount')
                balance=request.POST.get('loan_amount')
                lender_bank = request.POST.get('lender_bank')
                loan_date = request.POST.get('loan_date')
                payment_method = request.POST.get('payment_method')
                if payment_method is not None and payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=payment_method)
                    payment_method = acc.bnk_name
                # else:
                #     print("payment_method is not a number")
                upi_id=request.POST.get('upi_id')
                cheque=request.POST.get('cheque_number')
                payment_accountnumber=request.POST.get('laccount_number')
                processing_method = request.POST.get('processing_method')
                if processing_method is not None and processing_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=processing_method)
                    processing_method = acc.bnk_name
                processing_upi=request.POST.get('pupi_id')
                processing_cheque=request.POST.get('pcheque_number')
                processing_acc=request.POST.get('paccount_number')
                interest = request.POST.get('interest')
                processing_fee = request.POST.get('processing_fee')
                description = request.POST.get('description')
                term=request.POST.get('terms')
                
                interest = float(interest) if interest else 0
                processing_fee = float(processing_fee) if processing_fee else 0
                loaned_bank_account_ids = loan_account.objects.values_list('bank_holder_id', flat=True)
                loan = loan_account(
                    company=company,
                    logindetails=log_details,
                    bank_holder_id=account_name,
                    account_number=account_number,
                    loan_amount=loan_amount,
                    balance=balance,
                    lender_bank=lender_bank,
                    loan_date=loan_date,
                    payment_method=payment_method,
                    upi_id=upi_id,
                    cheque=cheque,
                    payment_accountnumber=payment_accountnumber,
                    processing_method=processing_method,
                    processing_upi=processing_upi,
                    processing_cheque=processing_cheque,
                    processing_acc=processing_acc,
                    interest=interest,
                    processing_fee=processing_fee,
                    description=description,
                    term=term
                )
                loan.save()
                
                history=LoanAccountHistory.objects.create(
                    login_details=log_details,
                    company=dash_details,
                    loan=loan,
                    date=now().date(),
                    action='Created'
                )
                history.save()
                
                context = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'bank_holder': BankAccount.objects.filter(company=company).exclude(id__in=loaned_bank_account_ids),
                    'loan_details': loan_account.objects.filter(company=company),
                    'banks':banks,
                    'today_date':today_date,
                    'selected_account_name': account_name,
                    'loan': loan,
                    'log_details':log_details
                }

                
                
                return redirect('loan_listing')
            else:
               
                return render(request, 'zohomodules/loan_account/add_loan.html', context)
        else:
            
            return HttpResponse("Unauthorized access")
    else:
        return redirect('/')
    

def save_account_details(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                # Fetch company details
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else:
                # Fetch staff details
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            
            if request.method == 'POST':
               print('inside post')
               
               customer_name = request.POST.get('customer_name')
               alias = request.POST.get('alias')
               phone_number = request.POST.get('phone_number')
               email = request.POST.get('email')
               account_type = request.POST.get('account_type')
               bankname = request.POST.get('bankname')
               account_number = request.POST.get('baccount_no')
               ifsc_code = request.POST.get('ifsc_code')
               swift_code = request.POST.get('swift_code')
               branch_name = request.POST.get('branch_name')
               cheque_book_range = request.POST.get('cheque_book_range')
               enable_cheque_printing = request.POST.get('enable_cheque_printing')
               cheque_printing_configuration = request.POST.get('cheque_printing_configuration')
               mailing_name = request.POST.get('mailing_name')
               address = request.POST.get('address')
               country = request.POST.get('country')
               state = request.POST.get('state')
               pin = request.POST.get('pin')
               pan_number = request.POST.get('pan_number')
               registration_type = request.POST.get('registration_type')
               gst_num = request.POST.get('gst_num')
               alter_gst_details = request.POST.get('gst_alter_details')
               date = request.POST.get('date')
               amount_type = request.POST.get('amount_type', None)
               amount = request.POST.get('amount')
               amount = float(amount) if amount else 0 
               alter_gst_details = alter_gst_details if  alter_gst_details else False 
               print(amount)
               print(alter_gst_details)
               

               if BankAccount.objects.filter( Q(pan_number=pan_number),company=company).exists():
                   print("inside panbankaccount filter")
                   return JsonResponse({'status': 'error', 'message': 'pan_number'}) 
               if gst_num and BankAccount.objects.filter(  Q(gst_num=gst_num),company=company).exists():
                   print("inside bgstankaccount filter")
                   return JsonResponse({'status': 'error', 'message': 'gst_num'}) 
               if BankAccount.objects.filter(  Q(phone_number=phone_number),company=company).exists():
                   print("inside phbankaccount filter")
                   return JsonResponse({'status': 'error', 'message': 'phone_number'}) 
               if BankAccount.objects.filter( Q(account_number=account_number),company=company).exists():
                   print("inside accbankaccount filter")
                   return JsonResponse({'status': 'error', 'message': 'account_number'}) 
               print('outside bank account filter')        

            try:
                bank=BankAccount(
                customer_name=customer_name,
                alias=alias,
                phone_number=phone_number,
                email=email,
                account_type=account_type,
                bankname=bankname,
                account_number=account_number,
                ifsc_code=ifsc_code,
                swift_code=swift_code,
                branch_name=branch_name,
                cheque_book_range=cheque_book_range,
                enable_cheque_printing=enable_cheque_printing,
                cheque_printing_configuration=cheque_printing_configuration,
                mailing_name=mailing_name,
                address=address,
                country=country,
                state=state,
                pin=pin,
                pan_number=pan_number,
                registration_type=registration_type,
                gst_num=gst_num,
                alter_gst_details=alter_gst_details,
                date=date,
                amount_type=amount_type,
                amount=amount,
                company=company,
                login_details=log_details,
               
                
                    
                    )
                bank.save()
                    
                BankAccountHistory
                bank_history=BankAccountHistory.objects.create(
                            logindetails=log_details,
                            company=dash_details,
                            bank_holder=bank,
                            date=now().date(),
                            action='Created'
                        )
                bank_history.save()
                new_account_id = bank.id  
                new_account_name = customer_name
                data = {
                        'status': 'success',
                        'account_id': new_account_id,
                        'customer_name': new_account_name
                    }
                    
                print(data)
                print('created')
                return JsonResponse(data)
            except Exception as e:
                error_message = str(e)
                print(error_message)
                return JsonResponse({'status': 'error', 'message': error_message})
    
    
def holder_dropdown(request):
    print("start fuction")
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules =ZohoModules.objects.get(company=dash_details.company, status='New')

            loaned_bank_account_ids = loan_account.objects.values_list('bank_holder_id', flat=True)    
            print("inside holder")
            options = {}
            option_objects = BankAccount.objects.filter(company=company).exclude(id__in=loaned_bank_account_ids)
            for option in option_objects:
                options[option.id] = option.customer_name
                print(option.customer_name)
            return JsonResponse(options)

   
def overview(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else: 
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            
            today=date.today()
            today_date = today.strftime("%Y-%m-%d")

            # loan_info = get_object_or_404(loan_account, id=account_id, company=company)
            # account = loan_info.bank_holder
            account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.filter(bank_holder=account,company=company).first()
            repayment_details = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
            repayment_history = LoanRepaymentHistory.objects.filter(repayment__in=repayment_details,company=company)
            # repayment_history = LoanRepaymentHistory.objects.filter(repayment='3')
            
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            current_balance = loan_info.loan_amount  
            balances = [] 
            loan_side = loan_account.objects.filter(company=company) 
            for loan in loan_side:
                total_emis_paid = LoanRepayemnt.objects.filter(company=company,loan=loan, type='EMI paid').aggregate(total=Sum('principal_amount'))['total'] or 0
                total_additional_loan = LoanRepayemnt.objects.filter(company=company,loan=loan, type='Additional Loan').aggregate(total=Sum('principal_amount'))['total'] or 0
                loan.balance = loan.loan_amount - total_emis_paid + total_additional_loan 

            for repayment in repayment_details:
                if repayment.type == 'EMI paid':
                    current_balance -= repayment.principal_amount
                elif repayment.type == 'Additional Loan':
                    current_balance += repayment.principal_amount     
                balances.append(current_balance)

            overall_balance = current_balance
            repayment_details_with_balances = zip(repayment_details, balances)
            total_amount= loan_info.loan_amount + loan_info.interest


            history=LoanAccountHistory.objects.filter(loan=loan_info,company=company)
            comment=Comments.objects.filter(loan=loan_info,company=company)

            context = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'log_id':log_details,
                    'account':account,
                    'loan_info':loan_info,
                    'repayment_details': repayment_details,
                    'repayment_details_with_balances': repayment_details_with_balances,
                    'overall_balance': overall_balance, 
                    'total_amount':total_amount,
                    'history':history,
                    'loan_side':loan_side,
                    'today_date':today_date,
                    'repayment_history':repayment_history,
                    'comment':comment,
                    'banks':banks,
                    
                    'account_id':account_id,
                    'loanpage':'0'
                    
                    
                     }          
    
            return render(request,'zohomodules/loan_account/overview.html',context)
        
from django.http import JsonResponse

def update_status(request, account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else: 
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        try:
            bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan = loan_account.objects.get(bank_holder=bank_account,company=company)     
            if loan.status == 'Active':
               loan.status = 'Inactive'
            else:
              loan.status = 'Active'   
            loan.save()       
            return redirect('overview',account_id=account_id)
        except loan_account.DoesNotExist:
         return render(request, 'zohomodules/loan_account/overview.html', {'message': 'Loan account does not exist'})



def repayment_due_form(request, account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
                
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            if request.method == 'POST':
                principal_amount = request.POST.get('principal_amount')
                interest_amount = request.POST.get('interest_amount')
                payment_method=request.POST.get('payment_method')
                if payment_method is not None and payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=payment_method)
                    payment_method = acc.bnk_name
                upi_id=request.POST.get('upi_id')
                cheque=request.POST.get('cheque_number')
                account_number=request.POST.get('acc_no')
                date = request.POST.get('date')
                principal_amount = float(principal_amount) if principal_amount else 0
                interest_amount = float(interest_amount) if interest_amount else 0
                
                total_amount = principal_amount + interest_amount
                # total_amount = float(principal_amount) + float(interest_amount)
                type = 'EMI paid'
                print(payment_method)
                repayment = LoanRepayemnt(
                    login_details=login_details,
                    company=company,
                    principal_amount=principal_amount,
                    interest_amount=interest_amount,
                    payment_method=payment_method,
                    upi_id=upi_id,
                    cheque=cheque,
                    account_number=account_number,
                    payment_date=date,
                    total_amount=total_amount,
                    type = type
                )
                
                
                bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
                loan = loan_account.objects.get(bank_holder=bank_account,company=company)
                repayment.loan = loan
                repayment.save()
                repayment_history=LoanRepaymentHistory.objects.create(
                    login_details=login_details,
                    company=company,
                    repayment=repayment,
                    date=now().date(),
                    action='Created'
                )
                repayment_history.save()
                # url = reverse('overview', kwargs={'account_id': account_id}) + '?Transaction=True'
                # return redirect(url)
                return redirect('transactoverview', account_id=account_id)
            else:
                today_date = dt.today()
                
                
                return render(request, 'zohomodules/loan_account/overview.html', { 'details': dash_details, 'allmodules': allmodules,  'today_date': today_date,'account_id': account_id,'banks':banks,'repayment_history':repayment_history,'login_details':login_details})
    return redirect('/')

def new_loan(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
                
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            today_date = dt.today()
            if request.method == 'POST':
                principal_amount = request.POST.get('principal_amount')
                interest_amount = request.POST.get('interest_amount')
                payment_method=request.POST.get('payment_method')
                if payment_method is not None and payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=payment_method)
                    payment_method = acc.bnk_name
                upi_id=request.POST.get('upi_id')
                cheque=request.POST.get('cheque_number')
                account_number=request.POST.get('acc_num')
                date = request.POST.get('date')
                principal_amount = float(principal_amount) if principal_amount else 0
                interest_amount = float(interest_amount) if interest_amount else 0
                
                # total_amount = principal_amount + interest_amount
                total_amount = request.POST.get('total_amount')
                type = 'Additional Loan'
                
                repayment = LoanRepayemnt(
                    login_details=login_details,
                    company=company,
                    principal_amount=principal_amount,
                    interest_amount=interest_amount,
                    payment_method=payment_method,
                    upi_id=upi_id,
                    cheque=cheque,
                    account_number=account_number,
                    payment_date=date,
                    total_amount=total_amount,
                    type = type
                )
                bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
                loan = loan_account.objects.get(bank_holder=bank_account,company=company)
                
                repayment.loan = loan
                repayment.save()

                repayment_history=LoanRepaymentHistory.objects.create(
                    login_details=login_details,
                    company=company,
                    repayment=repayment,
                    date=now().date(),
                    action='Created'
                )
                repayment_history.save()
                
                return redirect('transactoverview', account_id=account_id)    

            context={
                'allmodules':allmodules,
                'details': dash_details,
                'today_date': today_date,
                'account_id': account_id,
                'banks':banks,
                'repayment_history':repayment_history,
                'login_details': login_details
                
            }
            return render(request, 'zohomodules/loan_account/overview.html',context)
    return redirect('/')

def edit_loanaccount(request, account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')

            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)
            bank_holder=BankAccount.objects.filter(company=company)

            bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan = loan_account.objects.get(bank_holder=bank_account,company=company)


            return render(request, 'zohomodules/loan_account/edit_loan.html', {'account': loan, 'details':dash_details,'bank_holder':bank_holder, 'user_type': user_type, 'allmodules': allmodules,'banks':banks,'login_details':login_details})

    

def edit_loantable(request, account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')

            bank_account = BankAccount.objects.get(id=account_id,company=company)
            loan = loan_account.objects.get(bank_holder=bank_account,company=company)
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            if request.method == 'POST':
                
                loan.bank_holder.customer_name = request.POST.get('account_name')
                loan.loan_amount = request.POST.get('loan_amount')
                loan.lender_bank = request.POST.get('lender_bank')
                loan.loan_date = request.POST.get('loan_date')
                loan.payment_method = request.POST.get('payment_method')
                if loan.payment_method is not None and loan.payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=loan.payment_method)
                    loan.payment_method = acc.bnk_name
                loan.upi_id = request.POST.get('upi_id')
                loan.cheque = request.POST.get('cheque_number')
                loan.payment_accountnumber = request.POST.get('account_number')
                loan.term = request.POST.get('terms')
                loan.processing_method = request.POST.get('processing_method')
                if loan.processing_method is not None and loan.processing_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=loan.processing_method)
                    loan.processing_method = acc.bnk_name
                loan.processing_upi = request.POST.get('p_upi_id')
                loan.processing_cheque = request.POST.get('p_cheque_number')
                loan.processing_acc = request.POST.get('p_account_number')
                # loan.interest = request.POST.get('interest')
                # loan.processing_fee = request.POST.get('processing_fee')
                loan.description = request.POST.get('description')
                interest = request.POST.get('interest')
                processing_fee = request.POST.get('processing_fee')
                interest = float(interest) if interest else 0
                processing_fee = float(processing_fee) if processing_fee else 0

                loan.interest = interest
                loan.processing_fee = processing_fee
                loan.save()

                history=LoanAccountHistory.objects.create(
                    login_details=login_details,
                    company=company,
                    loan=loan,
                    date=now().date(),
                    action='Edited'
                )
                history.save()
                
               
                return redirect('overview', account_id=account_id)  

        return render(request, 'zohomodules/loan_account/edit_loan.html', {'loan': loan, 'details': dash_details, 'allmodules': allmodules,'history':history,'banks':banks,'account_id':account_id})



def calculate_overall_balance(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.filter(bank_holder=account,company=company).first()
            repayment_details = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
            current_balance = loan_info.loan_amount
            balances = [] 
            loan_side = loan_account.objects.all() 
            for loan in loan_side:
              total_emis_paid = LoanRepayemnt.objects.filter(company=company,loan=loan, type='EMI paid').aggregate(total=Sum('total_amount'))['total'] or 0
              total_additional_loan = LoanRepayemnt.objects.filter(company=company,loan=loan, type='Additional Loan').aggregate(total=Sum('total_amount'))['total'] or 0
              loan.balance = loan.loan_amount - total_emis_paid + total_additional_loan 

            for repayment in repayment_details:
                if repayment.type == 'EMI paid':
                    current_balance -= repayment.total_amount
                elif repayment.type == 'Additional Loan':
                    current_balance += repayment.total_amount     
                balances.append(current_balance)

            overall_balance = current_balance
            return overall_balance

def edit_repayment(request, repayment_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            repayment = get_object_or_404(LoanRepayemnt, id=repayment_id,company=company)
            account_id = repayment.loan.bank_holder_id 
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)
            
           
            if request.method == 'POST':
                principal_amount = request.POST.get('principal_amount')
                interest_amount = request.POST.get('interest_amount')
                payment_method = request.POST.get('payment_method')
                if payment_method is not None and payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=payment_method)
                    payment_method = acc.bnk_name
                upi_id = request.POST.get('upi_id')
                cheque = request.POST.get('cheque_number')
                account_number=request.POST.get('acc_no')
                payment_date = request.POST.get('date')
                principal_amount = float(principal_amount) if principal_amount else 0
                interest_amount = float(interest_amount) if interest_amount else 0
                
                # total_amount = principal_amount + interest_amount
                total_amount = request.POST.get('total_amount')
                type = 'EMI paid' 
                print(repayment.payment_method)
                repayment.principal_amount = principal_amount
                repayment.interest_amount = interest_amount
                repayment.payment_method = payment_method
                repayment.upi_id = upi_id
                repayment.cheque = cheque
                repayment.account_number=account_number
                repayment.payment_date = payment_date
                repayment.total_amount = total_amount
                repayment.type = type
        
                repayment.save()
                
                return redirect('transactoverview' ,account_id=account_id)
            else:
                repayment_history=LoanRepaymentHistory.objects.create(
                    login_details=login_details,
                    company=company,
                    repayment=repayment,
                    date=now().date(),
                    action='Edited'
                )
                repayment_history.save()
                
                
                return render(request, 'zohomodules/loan_account/edit_repayment.html', {'repayment': repayment,'details': dash_details,  'allmodules': allmodules, 'repayment_history':repayment_history,'banks':banks,'login_details':login_details})


def edit_additional_loan(request, repayment_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            repayment = get_object_or_404(LoanRepayemnt, id=repayment_id,company=company)
            account_id = repayment.loan.bank_holder_id 
            current_balance=calculate_overall_balance(request,account_id)
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)
    
            if request.method == 'POST':
                principal_amount = request.POST.get('principal_amount')
                interest_amount = request.POST.get('interest_amount')
                payment_method = request.POST.get('payment_method')
                if payment_method is not None and payment_method.isdigit():
                    print("payment_method is a number")
                    acc = Banking.objects.get(pk=payment_method)
                    payment_method = acc.bnk_name
                upi_id = request.POST.get('upi_id')
                cheque = request.POST.get('cheque_number')
                account_number=request.POST.get('acc_num')
                payment_date = request.POST.get('date')
                principal_amount = float(principal_amount) if principal_amount else 0
                interest_amount = float(interest_amount) if interest_amount else 0
                
                total_amount = principal_amount + interest_amount
                # total_amount = request.POST.get('total_amount')
                type = 'Additional Loan'
        
                repayment.principal_amount = principal_amount
                repayment.interest_amount = interest_amount
                repayment.payment_method = payment_method
                repayment.upi_id = upi_id
                repayment.cheque = cheque
                repayment.account_number=account_number
                repayment.payment_date = payment_date
                repayment.total_amount = total_amount
                repayment.type = type
        
                repayment.save()
                
                return redirect('transactoverview',account_id=account_id)
            else:
                hist=LoanRepaymentHistory.objects.create(
                    login_details=login_details,
                    company=company,
                    repayment=repayment,
                    date=now().date(),
                    action='Edited'
                )
                hist.save()
                return render(request, 'zohomodules/loan_account/edit_additional_loan.html', {'repayment': repayment,'details': dash_details,  'allmodules': allmodules,'overall_balance':current_balance,'hist':hist,'banks':banks,'login_details':login_details})            
            
from django.template.loader import render_to_string

def share_email(request, account_id):
    try:
        if request.method == 'POST':
            emails_string = request.POST['email']
            emails_list = [email.strip() for email in emails_string.split(',')]
            if 'login_id' in request.session:
                log_id = request.session['login_id']
                if 'login_id' not in request.session:
                    return redirect('/')
        
            login_details = LoginDetails.objects.get(id=log_id)
            user_type = login_details.user_type

            if user_type in ['Company', 'Staff']:
                if user_type == 'Company':
                    dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                    company=dash_details
                    allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                else:
                    dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                    company=dash_details.company
                    allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')


                emails_string = request.POST['email']
                emails_list = [email.strip() for email in emails_string.split(',')]
                bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
                loan_info = loan_account.objects.get(bank_holder=bank_account,company=company)

                repayment_details = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
                current_balance = loan_info.loan_amount  
                balances = [] 
                for repayment in repayment_details:
                  if repayment.type == 'EMI paid':
                      current_balance -= repayment.total_amount
                  elif repayment.type == 'Additional Loan':
                      current_balance += repayment.total_amount     
                      balances.append(current_balance)
                overall_balance = current_balance
                total_amount= loan_info.loan_amount + loan_info.interest

                context = {
                'loan_info': loan_info,
                'repayment_details': repayment_details,
                'repayment_details_with_balances': zip(repayment_details, balances),
                'overall_balance': overall_balance, 
                'total_amount': total_amount,
                'details': dash_details,  
                'allmodules': allmodules,
                'login_details':login_details
               }
                template_path = 'zohomodules/loan_account/mailoverview.html'
                template = get_template(template_path)
                html  = template.render(context)
                # html_content = render_to_string('zohomodules/loan_account/mailoverview.html', context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()

                filename = f'{loan_info.bank_holder.customer_name}-statement - {loan_info.id}.pdf'
                subject = f"{loan_info.bank_holder.customer_name} - {loan_info.id}- statement"
                email=EmailMultiAlternatives(subject, f"Hi,\nPlease find the attached statement - File-{loan_info.bank_holder.customer_name}  .\n--\nRegards,\n",from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf") 
                email.send(fail_silently=False)

                messages.success(request, 'Loan Statement has been shared successfully..!')
                return redirect('statementoverview', account_id)
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect('statementoverview', account_id) 
    

def adding_comment(request, account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')

            bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.get(bank_holder=bank_account,company=company)
            if request.method == 'POST':

                comment = request.POST.get('comments')

                comm=Comments.objects.create(
                    login_details=login_details,
                    loan=loan_info,
                    company=company,
                    comment=comment
                )
                comm.save()
                context={'details': dash_details,  'allmodules': allmodules}
                return redirect('overview',account_id=account_id)

        return render(request, 'zohomodules/loan_account/overview.html', context) 
    
def delete_comment(request, comment_id,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')   
              
            comment = get_object_or_404(Comments, id=comment_id,company=company)
            comment.delete()

            context={'details': dash_details,  'allmodules': allmodules,'account_id':account_id}

            return redirect('overview',account_id=account_id)
        return render(request, 'zohomodules/loan_account/overview.html',context) 
    

def delete_repaymenttable(request, id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')

            repayment = get_object_or_404(LoanRepayemnt, id=id,company=company)
            repayment.delete()
            account_id = repayment.loan.bank_holder_id
            context={'details': dash_details,  'allmodules': allmodules}

            return redirect('transactoverview',account_id=account_id)
        return render(request, 'zohomodules/loan_account/overview.html',context)
    
def delete_loan(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        login_details = LoginDetails.objects.get(id=log_id)
        user_type = login_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=login_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            else:
                dash_details = StaffDetails.objects.get(login_details=login_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')

            bank_account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.get(bank_holder=bank_account,company=company)
            transactions = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
           
            context={'details': dash_details,  'allmodules': allmodules,'loanaccount': loan_info}
            if transactions.exists():
                 messages.error(request, 'This account can be deleted as it has done some transactions !!')
                 return redirect('transactoverview', account_id=account_id)
            else:
                loan_info.delete()
                return redirect('loan_listing')
        return render(request, 'zohomodules/loan_account/overview.html',context)
#End

def employeeloan_trans(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type =='Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        employee_loan=EmployeeLoan.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=EmployeeLoan.objects.get(id=id)
        pay= payroll_employee.objects.filter(company=dash_details)
        comment_data=employeeloan_comments.objects.filter(employee=id)
        history=Employeeloan_history.objects.filter(employeeloan=id)
        his=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details).last()
        name = his.login_details.first_name + ' ' + his.login_details.last_name 
        action = his.action
        his_date=his.Date
        repay=EmployeeLoanRepayment.objects.filter(emp=id)
        last_loan = EmployeeLoanRepayment.objects.filter(emp=id).last().balance
        loan_trans = EmployeeLoanRepayment.objects.filter(emp=id)
       
    if log_details.user_type =='Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        employee_loan=EmployeeLoan.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=EmployeeLoan.objects.get(id=id)
        pay= payroll_employee.objects.filter(company=dash_details.company)
        comment_data=employeeloan_comments.objects.filter(employee=id,company=dash_details.company)
        history=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details.company)
        his=Employeeloan_history.objects.filter(employeeloan=id,company=dash_details.company).last()
        name = his.login_details.first_name + ' ' + his.login_details.last_name 
        action = his.action
        his_date=his.Date
        repay=EmployeeLoanRepayment.objects.filter(emp=id)
        last_loan = EmployeeLoanRepayment.objects.filter(emp=id).last().balance
        loan_trans = EmployeeLoanRepayment.objects.filter(emp=id)
       
    content = {
                'details': dash_details,
                'employee_loan':employee_loan,
                'p':p,
                'allmodules': allmodules,
                'comment':comment_data,
                'history':history,
                'log_id':log_details,
                'pay':pay,
                'his':his,
                'name':name,
                'action':action,
                'his_date':his_date,
                'last_loan':last_loan,
                'repay':repay,
                'loan_trans':loan_trans,
                'state':'1'
                
        }
  
    return render(request,'zohomodules/employe_loan/employeeloan_overview.html',content)
    
def transactoverview(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else: 
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            
            today=date.today()
            today_date = today.strftime("%Y-%m-%d")

            # loan_info = get_object_or_404(loan_account, id=account_id, company=company)
            # account = loan_info.bank_holder
            account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.filter(bank_holder=account,company=company).first()
            repayment_details = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
            repayment_history = LoanRepaymentHistory.objects.filter(repayment__in=repayment_details,company=company)
            # repayment_history = LoanRepaymentHistory.objects.filter(repayment='3')
            
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            current_balance = loan_info.loan_amount  
            balances = [] 
            loan_side = loan_account.objects.filter(company=company) 
            for loan in loan_side:
                total_emis_paid = LoanRepayemnt.objects.filter(company=company,loan=loan, type='EMI paid').aggregate(total=Sum('principal_amount'))['total'] or 0
                total_additional_loan = LoanRepayemnt.objects.filter(company=company,loan=loan, type='Additional Loan').aggregate(total=Sum('principal_amount'))['total'] or 0
                loan.balance = loan.loan_amount - total_emis_paid + total_additional_loan 

            for repayment in repayment_details:
                if repayment.type == 'EMI paid':
                    current_balance -= repayment.principal_amount
                elif repayment.type == 'Additional Loan':
                    current_balance += repayment.principal_amount     
                balances.append(current_balance)

            overall_balance = current_balance
            repayment_details_with_balances = zip(repayment_details, balances)
            total_amount= loan_info.loan_amount + loan_info.interest


            history=LoanAccountHistory.objects.filter(loan=loan_info,company=company)
            comment=Comments.objects.filter(loan=loan_info,company=company)

            context = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'log_id':log_details,
                    'account':account,
                    'loan_info':loan_info,
                    'repayment_details': repayment_details,
                    'repayment_details_with_balances': repayment_details_with_balances,
                    'overall_balance': overall_balance, 
                    'total_amount':total_amount,
                    'history':history,
                    'loan_side':loan_side,
                    'today_date':today_date,
                    'repayment_history':repayment_history,
                    'comment':comment,
                    'banks':banks,
                    
                    'account_id':account_id,
                    'loanpage':'1'
                    
                    
                     }          
    
            return render(request,'zohomodules/loan_account/overview.html',context)
            
def statementoverview(request,account_id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        
        log_details = LoginDetails.objects.get(id=log_id)
        user_type = log_details.user_type

        if user_type in ['Company', 'Staff']:
            if user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details, superadmin_approval=1, Distributor_approval=1)
                company=dash_details
                allmodules = ZohoModules.objects.get(company=dash_details, status='New')
                
            else: 
                dash_details = StaffDetails.objects.get(login_details=log_details, company_approval=1)
                company=dash_details.company
                allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
            
            today=date.today()
            today_date = today.strftime("%Y-%m-%d")

            # loan_info = get_object_or_404(loan_account, id=account_id, company=company)
            # account = loan_info.bank_holder
            account = get_object_or_404(BankAccount, id=account_id,company=company)
            loan_info = loan_account.objects.filter(bank_holder=account,company=company).first()
            repayment_details = LoanRepayemnt.objects.filter(loan=loan_info,company=company)
            repayment_history = LoanRepaymentHistory.objects.filter(repayment__in=repayment_details,company=company)
            # repayment_history = LoanRepaymentHistory.objects.filter(repayment='3')
            
            banks = Banking.objects.values('id','bnk_name','bnk_acno','status').filter(company=company)

            current_balance = loan_info.loan_amount  
            balances = [] 
            loan_side = loan_account.objects.filter(company=company) 
            for loan in loan_side:
                total_emis_paid = LoanRepayemnt.objects.filter(company=company,loan=loan, type='EMI paid').aggregate(total=Sum('principal_amount'))['total'] or 0
                total_additional_loan = LoanRepayemnt.objects.filter(company=company,loan=loan, type='Additional Loan').aggregate(total=Sum('principal_amount'))['total'] or 0
                loan.balance = loan.loan_amount - total_emis_paid + total_additional_loan 

            for repayment in repayment_details:
                if repayment.type == 'EMI paid':
                    current_balance -= repayment.principal_amount
                elif repayment.type == 'Additional Loan':
                    current_balance += repayment.principal_amount     
                balances.append(current_balance)

            overall_balance = current_balance
            repayment_details_with_balances = zip(repayment_details, balances)
            total_amount= loan_info.loan_amount + loan_info.interest


            history=LoanAccountHistory.objects.filter(loan=loan_info,company=company)
            comment=Comments.objects.filter(loan=loan_info,company=company)

            context = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'log_id':log_details,
                    'account':account,
                    'loan_info':loan_info,
                    'repayment_details': repayment_details,
                    'repayment_details_with_balances': repayment_details_with_balances,
                    'overall_balance': overall_balance, 
                    'total_amount':total_amount,
                    'history':history,
                    'loan_side':loan_side,
                    'today_date':today_date,
                    'repayment_history':repayment_history,
                    'comment':comment,
                    'banks':banks,
                    
                    'account_id':account_id,
                    'loanpage':'2'
                    
                    
                     }          
    
            return render(request,'zohomodules/loan_account/overview.html',context)
            
def list_godown(request):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            godown_obj = Godown.objects.filter(company = dash_details)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'godown_obj':godown_obj
            }
        
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            godown_obj = Godown.objects.filter(company = dash_details.company)
            allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'godown_obj':godown_obj
            }

        return render(request, 'godown/godown_list.html', context)
    
def add_godown(request):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            item_obj = Items.objects.filter(company = dash_details)
            units = Unit.objects.filter(company = dash_details)
            accounts = Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'item_obj':item_obj,
            'units':units,
            'accounts':accounts
            }
        
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
            item_obj = Items.objects.filter(company = dash_details.company)
            units = Unit.objects.filter(company = dash_details.company)
            accounts = Chart_of_Accounts.objects.filter(company=dash_details.company)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'item_obj':item_obj,
            'units':units,
            'accounts':accounts
            }

        return render(request, 'godown/add_godown.html', context)
    
def add_godown_func(request):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method == 'POST':
                date = request.POST.get('Date')
                item = request.POST.get('Item')
                gname = request.POST.get('Gname')
                gaddress = request.POST.get('Gaddress')
                stock = request.POST.get('Stock')
                distance = request.POST.get('Distance')
                item_obj = Items.objects.get(id=item)
                action = request.POST.get('save')
                godown = Godown(date=date,
                                item=item_obj,
                                stock_keeping=stock,
                                godown_name=gname,
                                godown_address=gaddress,
                                distance=distance,
                                stock_in_hand = item_obj.current_stock,
                                hsn = item_obj.hsn_code,
                                login_details=log_details,
                                company = company,
                                action = action)
                godown.save()

                godown_history = GodownHistory(company = company,
                                               login_details=log_details,
                                               godown=godown,
                                               date=date,
                                               action='Created')
                godown_history.save()


        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method == 'POST':
                date = request.POST.get('Date')
                item = request.POST.get('Item')
                gname = request.POST.get('Gname')
                gaddress = request.POST.get('Gaddress')
                stock = request.POST.get('Stock')
                distance = request.POST.get('Distance')
                item_obj = Items.objects.get(id=item)
                action = request.POST.get('save')
                godown = Godown(date=date,
                                item=item_obj,
                                stock_keeping=stock,
                                godown_name=gname,
                                godown_address=gaddress,
                                distance=distance,
                                stock_in_hand = item_obj.current_stock,
                                hsn = item_obj.hsn_code,
                                login_details=log_details,
                                company = company,
                                action = action)
                godown.save()

                godown_history = GodownHistory(company = company,
                                               login_details=log_details,
                                               godown=godown,
                                               date=date,
                                               action='Created')
                godown_history.save()

        
        messages.success(request,'Added Successfully')
        return redirect('add_godown')
    
def overview_page(request,pk):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            godown_obj = Godown.objects.filter(company = dash_details)
            p = Godown.objects.get(id = pk)
            godown_history = GodownHistory.objects.filter(godown=p)
            comment = GodownComments.objects.filter(godown=p)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'godown_obj':godown_obj,
            'p':p,
            'godown_history':godown_history,
            'comment':comment
            }
        
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            godown_obj = Godown.objects.filter(company = dash_details.company)
            allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
            p = Godown.objects.get(id = pk)
            godown_history = GodownHistory.objects.filter(godown=p)
            comment = GodownComments.objects.filter(godown=p)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'godown_obj':godown_obj,
            'p':p,
            'godown_history':godown_history,
            'comment':comment
            }

        return render(request, 'godown/overview_page.html', context)

def edit_godown(request,pk):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            item_obj = Items.objects.filter(company = dash_details)
            units = Unit.objects.filter(company = dash_details)
            godown_obj = Godown.objects.get(id=pk)
            accounts = Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'item_obj':item_obj,
            'units':units,
            'accounts':accounts,
            'godown_obj':godown_obj
            }
        
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
            item_obj = Items.objects.filter(company = dash_details.company)
            units = Unit.objects.filter(company = dash_details.company)
            godown_obj = Godown.objects.get(id=pk)
            accounts = Chart_of_Accounts.objects.filter(company=dash_details.company)
            context = {
            'details': dash_details,
            'log_details':log_details,
            'dash_details':dash_details,
            'allmodules':allmodules,
            'item_obj':item_obj,
            'units':units,
            'accounts':accounts,
            'godown_obj':godown_obj
            }

        return render(request, 'godown/edit_godown.html', context)
    
def edit_godown_func(request):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method == 'POST':
                godown_id = request.POST.get('godown_id')
                date = request.POST.get('Date')
                item = request.POST.get('Item')
                gname = request.POST.get('Gname')
                gaddress = request.POST.get('Gaddress')
                stock = request.POST.get('Stock')
                distance = request.POST.get('Distance')
                item_obj = Items.objects.get(id=item)
                godown = Godown.objects.get(id=godown_id)
                godown.date=date
                godown.item=item_obj
                godown.stock_keeping=stock
                godown.godown_name=gname
                godown.godown_address=gaddress
                godown.distance=distance
                godown.stock_in_hand = item_obj.current_stock
                godown.hsn = item_obj.hsn_code
                godown.login_details=log_details
                godown.company = company

                godown.save()

                godown_history = GodownHistory(company = company,
                                               login_details=log_details,
                                               godown=godown,
                                               date=date,
                                               action='Edited')
                godown_history.save()

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method == 'POST':
                godown_id = request.POST.get('godown_id')
                date = request.POST.get('Date')
                item = request.POST.get('Item')
                gname = request.POST.get('Gname')
                gaddress = request.POST.get('Gaddress')
                stock = request.POST.get('Stock')
                distance = request.POST.get('Distance')
                item_obj = Items.objects.get(id=item)
                godown = Godown.objects.get(id=godown_id)
                godown.date=date
                godown.item=item_obj
                godown.stock_keeping=stock
                godown.godown_name=gname
                godown.godown_address=gaddress
                godown.distance=distance
                godown.stock_in_hand = item_obj.current_stock
                godown.hsn = item_obj.hsn_code
                godown.login_details=log_details
                godown.company = company

                godown.save()

                godown_history = GodownHistory(company = company,
                                               login_details=log_details,
                                               godown=godown,
                                               date=date,
                                               action='Edited')
                godown_history.save()
        
        messages.success(request,'Edited Successfully')
        return redirect('list_godown')
    
def newitem(request):

    return render(request,'godown/try.html')


def change_status(request, pk):

    godown_obj = Godown.objects.get(id=pk)
    if godown_obj.status == 'Active':
        godown_obj.status='Inactive'
    else:
        godown_obj.status='Active'
    godown_obj.save()
    return redirect('overview_page',pk=pk)

def change_action(request, pk):

    godown_obj = Godown.objects.get(id=pk)
    godown_obj.action='Adjusted'
    godown_obj.save()
    return redirect('overview_page',pk=pk)

def AddComment(request,pk):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        godown = Godown.objects.get(id=pk)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            
            if request.method == 'POST':
                comments = request.POST.get('comments')
                comment = GodownComments(
                                login_details=log_details,
                                company = company,
                                godown = godown,
                                comment = comments)
                comment.save()


        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method == 'POST':
                comments = request.POST.get('comments')
                comment = GodownComments(
                                login_details=log_details,
                                company = company,
                                godown = godown,
                                comment = comments)
                comment.save()
        
        messages.success(request,'Added Comment Successfully')
        return redirect('overview_page',pk=pk)
    
def AddFile(request, pk):

    godown_obj = Godown.objects.get(id=pk)
    if request.method == 'POST':
        file = request.FILES.get('file')
        godown_obj.file=file
        godown_obj.save()
    messages.success(request,'Added File Successfully')
    return redirect('overview_page',pk=pk)

def file_download(request,pk):
    godown_obj= Godown.objects.get(id=pk)
    file = godown_obj.file
    response = FileResponse(file)
    response['Content-Disposition'] = f'attachment; filename="{file.name}"'
    return response

def ShareEmail(request,pk):
    try:
            if request.method == 'POST':
                emails_string = request.POST['email']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                print(emails_list)
                p=Godown.objects.get(id=pk)
                        
                context = {'p':p}
                template_path = 'godown/overview_page.html'
                template = get_template(template_path)
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'{p.godown_name}details - {p.id}.pdf'
                subject = f"{p.godown_name}{p.godown_address}  - {p.id}-details"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached godown details - File-{p.godown_name}{p.godown_address} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)
                messages.success(request, 'over view page has been shared via email successfully..!')
                return redirect('overview_page',pk=pk)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('overview_page',pk=pk)
    
    
def DeleteComment(request,pk):
        
    comment = GodownComments.objects.get(id=pk)
    comment.delete()

    messages.success(request,'Deleted Comment Successfully')
    return redirect('overview_page',pk=pk)

def Add_Item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.Date=date.today()
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track = request.POST.get("trackState",None)
            track_state_value = request.POST.get("trackstate", None)

# Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0

            
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if Items.objects.filter(item_name=item_name, company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('add_godown')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('add_godown')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                messages.success(request,'Item Added Successfully !!!')
                return redirect('add_godown')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track_state_value = request.POST.get("trackState", None)

            track_state_value = request.POST.get("trackstate", None)

            # Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0
            minstock=request.POST.get("minimum_stock",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
        
        

        
            if Items.objects.filter(item_name=item_name,company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('add_godown')
                
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('add_godown')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                messages.success(request,'Item Added Successfully !!!')
                return redirect('add_godown')
    return redirect('add_godown')

def godownmodal_unit(request):
    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method=='POST':
                units =request.POST.get('unit_name')
                
                unit_obj = Unit(unit_name=units,
                        company=company)
                unit_obj.save()
                return JsonResponse({"message": "success"})
        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method=='POST':
                units =request.POST.get('unit_name')
            
                unit_obj = Unit(unit_name=units,
                        company=company)
                unit_obj.save()
                return JsonResponse({"message": "success"})
        return redirect('add_godown')

def godownunit_dropdown(request):

    options = {}
    option_objects = Unit.objects.all()
    for option in option_objects:
        options[option.id] = [option.unit,option.id]

    return JsonResponse(options)
    
def AddAccount(request):
    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method=='POST':
                account_type =request.POST.get('account_type')
                account_name =request.POST.get('account_name')
                account_code =request.POST.get('account_code')
                account_number =request.POST.get('account_number')
                description =request.POST.get('description')
            
                accounts = Chart_of_Accounts(account_type=account_type,
                                             account_name=account_name,
                                             description=description,
                                             account_number=account_number,
                                             account_code=account_code,
                                             company=company,
                                             login_details=log_details)
                accounts.save()
                messages.success(request,'Account Added Successfully !!!')
        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method=='POST':
                account_type =request.POST.get('account_type')
                account_name =request.POST.get('account_name')
                account_code =request.POST.get('account_code')
                
                description =request.POST.get('description')
            
                accounts = Chart_of_Accounts(account_type=account_type,
                                             account_name=account_name,
                                             description=description,
                                            
                                             account_code=account_code,
                                             company=company,
                                             login_details=log_details)
                accounts.save()
                messages.success(request,'Account Added Successfully !!!')
        return redirect('add_godown')
    

def Add_Item_Edit(request,pk):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.Date=date.today()
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track = request.POST.get("trackState",None)
            track_state_value = request.POST.get("trackstate", None)

# Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0

            
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if Items.objects.filter(item_name=item_name, company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('edit_godown',pk=pk)
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('edit_godown',pk=pk)
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                messages.success(request,'Item Added Successfully !!!')
                return redirect('edit_godown',pk=pk)
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track_state_value = request.POST.get("trackState", None)

            track_state_value = request.POST.get("trackstate", None)

            # Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0
            minstock=request.POST.get("minimum_stock",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
        
        

        
            if Items.objects.filter(item_name=item_name,company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('edit_godown',pk=pk)
                
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('edit_godown',pk=pk)
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                messages.success(request,'Item Added Successfully !!!')
                return redirect('edit_godown',pk=pk)
    return redirect('edit_godown',pk=pk)

def godownmodal_unit_edit(request,pk):
    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method=='POST':
                units =request.POST.get('unit_name')
            
                unit_obj = Unit(unit_name=units,
                            company=company)
                unit_obj.save()
                

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method=='POST':
                units =request.POST.get('unit_name')
                
        
                unit_obj = Unit(unit_name=units,
                        company=company)
                unit_obj.save()
               
        return redirect('edit_godown',pk)
    
    
def Add_Account_Edit(request,pk):
    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            if request.method=='POST':
                account_type =request.POST.get('account_type')
                account_name =request.POST.get('account_name')
                account_code =request.POST.get('account_code')
                account_number =request.POST.get('account_number')
                description =request.POST.get('description')
            
                accounts = Chart_of_Accounts(account_type=account_type,
                                             account_name=account_name,
                                             description=description,
                                             account_number=account_number,
                                             account_code=account_code,
                                             company=company,
                                             login_details=log_details)
                accounts.save()
                messages.success(request,'Account Added Successfully !!!')
        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
            if request.method=='POST':
                account_type =request.POST.get('account_type')
                account_name =request.POST.get('account_name')
                account_code =request.POST.get('account_code')
                
                description =request.POST.get('description')
            
                accounts = Chart_of_Accounts(account_type=account_type,
                                             account_name=account_name,
                                             description=description,
                                            
                                             account_code=account_code,
                                             company=company,
                                             login_details=log_details)
                accounts.save()
                messages.success(request,'Account Added Successfully !!!')
        return redirect('edit_godown',pk=pk)


def delete_godown(request,pk):

    godown = Godown.objects.get(id=pk)
    godown.delete()
    messages.success(request,'Deleted Successfully')
    return redirect('list_godown')


#--------------------------Meenu Shaju -Delivery Challan--------------

def challan_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        dc = Delivery_challan.objects.all()
    
    
        return render(request,'zohomodules/Delivery-challan/challan_list.html',{'d_challan':dc})

def delivery_challan(request):
     
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        
        customer=Customer.objects.filter(company=comp_details,customer_status='Active')
        item=Items.objects.filter(company=comp_details,activation_tag='Active')
        
        
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')
        last_reference = Delivery_challan.objects.filter(company=comp_details).last()
        if last_reference:
            next_reference_number = last_reference.reference_number + 1
        else:
            next_reference_number = 1
        last_challan = Delivery_challan.objects.filter(company=comp_details).last()

        

        if last_challan:
            challan_number = last_challan.challan_number
            prefix = ''.join(filter(str.isalpha, challan_number))
            numeric_part = ''.join(filter(str.isdigit, challan_number))
            
            if numeric_part:
                next_numeric_part = str(int(numeric_part) + 1).zfill(len(numeric_part))
            else:
                next_numeric_part = '001'  # If no numeric part exists, start with '001'
            
            next_challan_number = f'{prefix}{next_numeric_part}'
        else:
            next_challan_number = ''
        return render(request,'zohomodules/Delivery-challan/new_challan.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details,'price_lists':price_lists,'customer':customer,'item':item,'reference_number':next_reference_number,'challan_number':next_challan_number}) 
     

       
     


def get_customer_data(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
       
        data = {
            'email': customer.customer_email,
            'billing_address': customer.billing_address,
            'gsttype':customer.GST_treatement,
            'gstnumber':customer.GST_number,
            'place':customer.place_of_supply
           
        }
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
    


def get_item_data(request, item_id):
    try:
        item = Items.objects.get(id=item_id)
       
        data = {
            'hsn': item.hsn_code,
            'rate': item.purchase_price,
            'intratax':item.intrastate_tax,
            'intertax':item.interstate_tax,
            'company_state':item.company.state,
            'stock': item.current_stock 
            
           
        }
        return JsonResponse(data)
    except item.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
        
def challan_add_customer(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        

       
        if request.method=="POST":
            customer_dat=Customer()
            customer_dat.login_details=log_details
            customer_dat.company=comp_details
            customer_dat.customer_type = request.POST.get('type')

            customer_dat.title = request.POST.get('salutation')
            customer_dat.first_name=request.POST['first_name']
            customer_dat.last_name=request.POST['last_name']
            customer_dat.company_name=request.POST['company_name']
            customer_dat.customer_display_name=request.POST['v_display_name']
            customer_dat.customer_email=request.POST['vendor_email']
            customer_dat.customer_phone=request.POST['w_phone']
            customer_dat.customer_mobile=request.POST['m_phone']
            customer_dat.skype=request.POST['skype_number']
            customer_dat.designation=request.POST['designation']
            customer_dat.department=request.POST['department']
            customer_dat.website=request.POST['website']
            customer_dat.GST_treatement=request.POST['gst']
            customer_dat.customer_status="Active"
            customer_dat.remarks=request.POST['remark']
            customer_dat.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_dat.PAN_number=request.POST['pan_number']
                customer_dat.GST_number="null"
            else:
                customer_dat.GST_number=request.POST['gst_number']
                customer_dat.PAN_number=request.POST['pan_number']

            customer_dat.place_of_supply=request.POST['source_supply']
            customer_dat.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_dat.opening_balance_type=op_type
            else:
                customer_dat.opening_balance_type='Opening Balance not selected'
    
            customer_dat.opening_balance=request.POST['opening_bal']
            customer_dat.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            # customer_data.price_list=request.POST['plst']
            plst=request.POST.get('plst')
            if plst!=0:
                 customer_dat.price_list=plst
            else:
                customer_dat.price_list='Price list not selected'




            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_dat.portal_language=plang
            else:
                customer_dat.portal_language='Portal language not selected'

            customer_dat.facebook=request.POST['fbk']
            customer_dat.twitter=request.POST['twtr']
            customer_dat.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_dat.customer_type=type
            else:
                customer_dat.customer_type='Customer type not selected'
    



           
            customer_dat.billing_attention=request.POST['battention']
            customer_dat.billing_country=request.POST['bcountry']
            customer_dat.billing_address=request.POST['baddress']
            customer_dat.billing_city=request.POST['bcity']
            customer_dat.billing_state=request.POST['bstate']
            customer_dat.billing_pincode=request.POST['bzip']
            customer_dat.billing_mobile=request.POST['bphone']
            customer_dat.billing_fax=request.POST['bfax']
            customer_dat.shipping_attention=request.POST['sattention']
            customer_dat.shipping_country=request.POST['s_country']
            customer_dat.shipping_address=request.POST['saddress']
            customer_dat.shipping_city=request.POST['scity']
            customer_dat.shipping_state=request.POST['sstate']
            customer_dat.shipping_pincode=request.POST['szip']
            customer_dat.shipping_mobile=request.POST['sphone']
            customer_dat.shipping_fax=request.POST['sfax']
            customer_dat.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_dat
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            vdata=Customer.objects.get(id=customer_dat.id)
            vendor=vdata
            rdata=Customer_remarks_table()
            rdata.remarks=request.POST['remark']
            rdata.company=comp_details
            rdata.customer=vdata
            rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Customer.objects.get(id=customer_dat.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = CustomerContactPersons.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor)
                
            data = {
            "message": "success",
            'customer_id': customer_dat.id
            
            }
            return JsonResponse(data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400) 


def challan_customer_dropdown(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company=CompanyDetails.objects.get(login_details=log_details)
            options = {}
            option_objects =Customer.objects.filter(company=company,customer_status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
            
        if log_details.user_type=='Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            options = {}
            option_objects = Customer.objects.filter(company=staff.company,customer_status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
           


            
               
       
    

            
def challan_customer_payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            normalized_data = terms.replace(" ", "")
            pay_tm = add_space_before_first_digit(normalized_data)
            ptr = Company_Payment_Term(term_name=pay_tm, days=day, company=comp_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=comp_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)   
        

def challan_term_dropdown(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details = LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            company=CompanyDetails.objects.get(login_details=log_details)
            options = {}
            option_objects =Company_Payment_Term.objects.filter(company=company,status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
            
        if log_details.user_type=='Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            options = {}
            option_objects = Company_Payment_Term.objects.filter(company=staff.company,status='Active')
            for option in option_objects:
                full_name = f"{option.first_name} {option.last_name}"
                options[option.id] = full_name

            return JsonResponse(options)
           
            



def challan_customer_check_gst(request):
    if request.method == 'POST':
        gstNumber = request.POST.get('gstNumber')
        gst_exists = Customer.objects.filter(GST_number=gstNumber).exists()
       
        if gst_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'}) 


def challan_check_customer_term_exist(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

    if request.method == 'GET':
       term_name = request.GET.get('term_name', None)
       if term_name:
            normalized_data = term_name.replace(" ", "")
            term_name_processed = add_space_before_first_digit(normalized_data)
            exists = Company_Payment_Term.objects.filter(
                    term_name=term_name_processed,
                    company=comp_details
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})    


def challan_check_customer_email_exist(request):
    if request.method == 'GET':
       vendoremail = request.GET.get('vendor_email', None)

       if vendoremail:
          
            exists = Customer.objects.filter(
                    customer_email=vendoremail
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 


def challan_check_customer_work_phone_exist(request):
    if request.method == 'GET':
       wPhone = request.GET.get('w_Phone', None)

       if wPhone:
          
            exists = Customer.objects.filter(
                    customer_phone=wPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})  
    

def challan_check_customer_phonenumber_exist(request):
    if request.method == 'GET':
       mPhone = request.GET.get('m_Phone', None)

       if mPhone:
          
            exists = Customer.objects.filter(
                    customer_mobile=mPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 
    


def add_delivery_challan(request):
     if request.method == 'POST':
        if 'login_id' in request.session:
            if request.session.has_key('login_id'):
                log_id = request.session['login_id']
            
            else:
                return redirect('/')
        
            log_details= LoginDetails.objects.get(id=log_id)
            if log_details.user_type=='Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

            else:    
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                comp_details=CompanyDetails.objects.get(login_details=log_details)

            if 'draft' in request.POST:  
                cname = request.POST['customerName'] 
                customer = Customer.objects.get(id=cname)
                place_of_supply = request.POST['placeOfSupply']
                dc_number = request.POST['deliveryChallan']
                ref_number = request.POST['referenceNumber']
                dc_date = request.POST['deliveryChallanDate']
                dc_type = request.POST['challanType']
                description = request.POST['note']
                file = request.FILES.get('file')
                item_lists = request.POST.getlist('item[]')
                
                hsn_list = request.POST.getlist('hsn[]')
                quantity_list = request.POST.getlist('quantity[]')
                rate_list = request.POST.getlist('rate[]')
                dicount_list = request.POST.getlist('discount[]')
                tax_list = request.POST.getlist('tax[]')
                amount_list = request.POST.getlist('amount[]')
                subtotal = request.POST['subtotal']
                igst = request.POST['igst']
                cgst = request.POST['cgst']
                sgst = request.POST['sgst']
                taxamount = request.POST['taxAmount']
                shipping = request.POST['shippingCharge']
                adjustment = request.POST['adjustment']
                grand_total = request.POST['total']
                advance = request.POST['advance']
                balance = request.POST['balance']

                
                product = request.POST.getlist("item[]")
                quantity = [int(qty) for qty in request.POST.getlist("quantity[]")]
                total_texts = request.POST.getlist("amount[]")
                total = [float(value) for value in total_texts]
                discount = [float(disc) for disc in request.POST.getlist("discount[]")]
                hsn = [int(code) for code in request.POST.getlist("hsn[]")]
                rate = [float(r) for r in request.POST.getlist("rate[]")]
                tax = [float(t) for t in request.POST.getlist("tax[]")]

                if '0' in quantity:
                    messages.info(request, 'Quantity of one item is 0')
                    return redirect('delivery_challan')

                if all(int(qty) > 0 for qty in quantity):
                    dc = Delivery_challan(
                        login_details=log_details,
                        company=comp_details,
                        customer=customer,
                        place_of_supply=place_of_supply,
                        challan_date=dc_date,
                        reference_number=ref_number,
                        challan_number=dc_number,
                        challan_type=dc_type,
                        description=description,
                        document=file,
                        sub_total=subtotal,
                        igst=igst,
                        cgst=cgst,
                        sgst=sgst,
                        tax_amount=taxamount,
                        shipping_charge=shipping,
                        adjustment=adjustment,
                        grand_total=grand_total,
                        advance=advance,
                        balance=balance,
                        status='Draft'
                    )
                    
                    dc.save()

                    if len(product) == len(quantity) == len(discount) == len(total) == len(hsn) == len(tax) == len(rate):
                    

    
                    

                        group = zip(product, hsn, quantity, rate, tax, discount, total)

                        try:
                            mapped = list(group)
                            print(mapped)
                        except Exception as e:
                            print("Exception occurred during conversion:", e)

                        print(mapped)
                        print('HI')

                        for itemsNew in mapped:
                            item_id = int(itemsNew[0])  
                            item_instance = Items.objects.get(id=item_id)
                            itemsTable = Delivery_challan_item(
                                item=item_instance, 
                                hsn=itemsNew[1], 
                                quantity=itemsNew[2], 
                                price=itemsNew[3], 
                                tax_rate=itemsNew[4], 
                                discount=itemsNew[5], 
                                total=itemsNew[6], 
                                delivery_challan=dc, 
                                login_details=log_details, 
                                company=comp_details
                            )
                            itemsTable.save()
                    dc_reference = Delivery_challan_reference(
                        login_details=log_details,
                        company=comp_details,
                        reference_number=ref_number
                    )
                    print('afer afer loop')
                    dc_reference.save()

                    current_date = date.today()
                    dc_history = Delivery_challan_history(
                        login_details=log_details,
                        company=comp_details,
                        delivery_challan=dc,
                        date=current_date,
                        action='Created'
                    )
                    dc_history.save()

                    
                    return redirect('challan_list')

            if 'save' in request.POST:  
                cname = request.POST['customerName'] 
                customer = Customer.objects.get(id=cname)
                place_of_supply = request.POST['placeOfSupply']
                dc_number = request.POST['deliveryChallan']
                ref_number = request.POST['referenceNumber']
                dc_date = request.POST['deliveryChallanDate']
                dc_type = request.POST['challanType']
                description = request.POST['note']
                file = request.FILES.get('file')
                item_lists = request.POST.getlist('item[]')
                
                hsn_list = request.POST.getlist('hsn[]')
                quantity_list = request.POST.getlist('quantity[]')
                rate_list = request.POST.getlist('rate[]')
                dicount_list = request.POST.getlist('discount[]')
                tax_list = request.POST.getlist('tax[]')
                amount_list = request.POST.getlist('amount[]')
                subtotal = request.POST['subtotal']
                igst = request.POST['igst']
                cgst = request.POST['cgst']
                sgst = request.POST['sgst']
                taxamount = request.POST['taxAmount']
                shipping = request.POST['shippingCharge']
                adjustment = request.POST['adjustment']
                grand_total = request.POST['total']
                advance = request.POST['advance']
                balance = request.POST['balance']

                
                product = request.POST.getlist("item[]")
                quantity = [int(qty) for qty in request.POST.getlist("quantity[]")]
                total_texts = request.POST.getlist("amount[]")
                total = [float(value) for value in total_texts]
                discount = [float(disc) for disc in request.POST.getlist("discount[]")]
                hsn = [int(code) for code in request.POST.getlist("hsn[]")]
                rate = [float(r) for r in request.POST.getlist("rate[]")]
                tax = [float(t) for t in request.POST.getlist("tax[]")]

                if '0' in quantity:
                    messages.info(request, 'Quantity of one item is 0')
                    return redirect('delivery_challan')

                if all(int(qty) > 0 for qty in quantity):
                    dc = Delivery_challan(
                        login_details=log_details,
                        company=comp_details,
                        customer=customer,
                        place_of_supply=place_of_supply,
                        challan_date=dc_date,
                        reference_number=ref_number,
                        challan_number=dc_number,
                        challan_type=dc_type,
                        description=description,
                        document=file,
                        sub_total=subtotal,
                        igst=igst,
                        cgst=cgst,
                        sgst=sgst,
                        tax_amount=taxamount,
                        shipping_charge=shipping,
                        adjustment=adjustment,
                        grand_total=grand_total,
                        advance=advance,
                        balance=balance,
                        status='Save'
                    )
                    
                    dc.save()

                    if len(product) == len(quantity) == len(discount) == len(total) == len(hsn) == len(tax) == len(rate):
                    

    
                    

                        group = zip(product, hsn, quantity, rate, tax, discount, total)

                        try:
                            mapped = list(group)
                            print(mapped)
                        except Exception as e:
                            print("Exception occurred during conversion:", e)

                        print(mapped)
                        print('HI')

                        for itemsNew in mapped:
                            item_id = int(itemsNew[0])  
                            item_instance = Items.objects.get(id=item_id)
                            itemsTable = Delivery_challan_item(
                                item=item_instance, 
                                hsn=itemsNew[1], 
                                quantity=itemsNew[2], 
                                price=itemsNew[3], 
                                tax_rate=itemsNew[4], 
                                discount=itemsNew[5], 
                                total=itemsNew[6], 
                                delivery_challan=dc, 
                                login_details=log_details, 
                                company=comp_details
                            )
                            itemsTable.save()
                    dc_reference = Delivery_challan_reference(
                        login_details=log_details,
                        company=comp_details,
                        reference_number=ref_number
                    )
                    print('afer afer loop')
                    dc_reference.save()

                    current_date = date.today()
                    dc_history = Delivery_challan_history(
                        login_details=log_details,
                        company=comp_details,
                        delivery_challan=dc,
                        date=current_date,
                        action='Created'
                    )
                    dc_history.save()

                    
                    return redirect('challan_list')
                
        return redirect('/')

def challan_overview(request,id):
    if 'login_id' in request.session:
            if request.session.has_key('login_id'):
                log_id = request.session['login_id']
            
            else:
                return redirect('/')
        
            log_details= LoginDetails.objects.get(id=log_id)
            if log_details.user_type=='Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

            else:    
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                comp_details=CompanyDetails.objects.get(login_details=log_details)

            challan = Delivery_challan.objects.get(id=id)
            all_challan = Delivery_challan.objects.all()
            items = Delivery_challan_item.objects.filter(company=comp_details,delivery_challan=challan)
            comments = Delivery_challan_comment.objects.filter(company=comp_details,delivery_challan=challan)
            history = Delivery_challan_history.objects.filter(company=comp_details,delivery_challan=challan)
        
            return render(request,'zohomodules/Delivery-challan/challan_overview.html',{'challan':challan,'d_challan':all_challan,'items':items,'comments':comments,'history':history}) 
                
def convert_save(request,id):
    dc = Delivery_challan.objects.get(id=id)
    dc.status = 'Save'
    dc.save()
    return redirect(reverse('challan_overview', args=[id]))


def challan_edit(request,id):
    
   if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        
        customer=Customer.objects.filter(company=comp_details,customer_status='Active')
        item=Items.objects.filter(company=comp_details,activation_tag='Active')
        
        
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')
        
        

        
        dc = Delivery_challan.objects.get(id=id)
        dct = Delivery_challan_item.objects.filter(delivery_challan=dc,company=comp_details)
       
        
        return render(request,'zohomodules/Delivery-challan/challan_edit.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details,'price_lists':price_lists,'customer':customer,'item':item,'challan':dc,'citem':dct}) 
     
def edit_challan(request,id):
    if request.method == 'POST':
            if 'login_id' in request.session:
                if request.session.has_key('login_id'):
                    log_id = request.session['login_id']
                
                else:
                    return redirect('/')
            
                log_details= LoginDetails.objects.get(id=log_id)
                if log_details.user_type=='Staff':
                    dash_details = StaffDetails.objects.get(login_details=log_details)
                    comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

                else:    
                    dash_details = CompanyDetails.objects.get(login_details=log_details)
                    comp_details=CompanyDetails.objects.get(login_details=log_details)

                  
                cname = request.POST['customerName'] 
                customer = Customer.objects.get(id=cname)
                place_of_supply = request.POST['placeOfSupply']
                dc_number = request.POST['deliveryChallan']
                ref_number = request.POST['referenceNumber']
                dc_date = request.POST['deliveryChallanDate']
                dc_type = request.POST['challanType']
                description = request.POST['note']
                file = request.FILES.get('file')
                item_lists = request.POST.getlist('item[]')
                    
                    
                    
                subtotal = request.POST['subtotal']
                igst = request.POST['igst']
                cgst = request.POST['cgst']
                sgst = request.POST['sgst']
                taxamount = request.POST['taxAmount']
                shipping = request.POST['shippingCharge']
                adjustment = request.POST['adjustment']
                grand_total = request.POST['total']
                advance = request.POST['advance']
                balance = request.POST['balance']

                    
                product = request.POST.getlist("item[]")
                quantity = [int(qty) for qty in request.POST.getlist("quantity[]")]
                total_texts = request.POST.getlist("amount[]")
                total = [float(value) for value in total_texts]
                discount = [float(disc) for disc in request.POST.getlist("discount[]")]
                hsn = [int(code) for code in request.POST.getlist("hsn[]")]
                rate = [float(r) for r in request.POST.getlist("rate[]")]
                tax = [float(t) for t in request.POST.getlist("tax[]")]

                if quantity<=0:
                        messages.info(request, 'Quantity of one item is 0')
                        return redirect('challan_edit',id=id)

                if all(int(qty) > 0 for qty in quantity):
                        dc = Delivery_challan(
                            login_details=log_details,
                            company=comp_details,
                            customer=customer,
                            place_of_supply=place_of_supply,
                            challan_date=dc_date,
                            reference_number=ref_number,
                            challan_number=dc_number,
                            challan_type=dc_type,
                            description=description,
                            document=file,
                            sub_total=subtotal,
                            igst=igst,
                            cgst=cgst,
                            sgst=sgst,
                            tax_amount=taxamount,
                            shipping_charge=shipping,
                            adjustment=adjustment,
                            grand_total=grand_total,
                            advance=advance,
                            balance=balance,
                            status='Save'
                        )
                        
                        dc.save()

                        if len(product) == len(quantity) == len(discount) == len(total) == len(hsn) == len(tax) == len(rate):
                        

        
                        

                            group = zip(product, hsn, quantity, rate, tax, discount, total)

                            try:
                                mapped = list(group)
                                print(mapped)
                            except Exception as e:
                                print("Exception occurred during conversion:", e)

                            

                            for itemsNew in mapped:
                                item_id = int(itemsNew[0])  
                                item_instance = Items.objects.get(id=item_id)
                                itemsTable = Delivery_challan_item(
                                    item=item_instance, 
                                    hsn=itemsNew[1], 
                                    quantity=itemsNew[2], 
                                    price=itemsNew[3], 
                                    tax_rate=itemsNew[4], 
                                    discount=itemsNew[5], 
                                    total=itemsNew[6], 
                                    delivery_challan=dc, 
                                    login_details=log_details, 
                                    company=comp_details
                                )
                                itemsTable.save()
                        dc_reference = Delivery_challan_reference(
                            login_details=log_details,
                            company=comp_details,
                            reference_number=ref_number
                        )
                        
                        dc_reference.save()

                        current_date = date.today()
                        dc_history = Delivery_challan_history(
                            login_details=log_details,
                            company=comp_details,
                            delivery_challan=dc,
                            date=current_date,
                            action='Edited'
                        )
                        dc_history.save()

                        
                        return redirect('challan_list')


def challan_add_comment(request):
    if request.method == 'POST':
        if 'login_id' not in request.session:
            return JsonResponse({'error': 'User not logged in'}, status=401)

        log_id = request.session['login_id']
        log_details = LoginDetails.objects.get(id=log_id)

        if log_details.user_type == 'Staff':
            staff = StaffDetails.objects.get(login_details=log_details)
            company = staff.company
        elif log_details.user_type == 'Company':
            company = CompanyDetails.objects.get(login_details=log_details)

        challan_id = request.POST.get('challan_id')
        challan = Delivery_challan.objects.get(id=challan_id)
        
        comment_text = request.POST.get('comment')
        current_date = date.today()
        
        if comment_text:  
            comment = Delivery_challan_comment(
                comment=comment_text,
                delivery_challan=challan,
                date = current_date,
                company=company,
                login_details=log_details
            )
            comment.save()

            return JsonResponse({'message': 'Comment added successfully'})
        else:
            return JsonResponse({'error': 'Comment text is required'}, status=400)  

    return JsonResponse({'error': 'Invalid request method'}, status=405)
def delete_challan_comment(request,id):
    comment = Delivery_challan_comment.objects.get(id=id)    
    comment.delete()  
    return redirect('challan_overview',id=comment.delivery_challan.id )       
                

                
                
